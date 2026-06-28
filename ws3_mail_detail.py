"""NetSort WS3 FCFL Customer Mail Detail XLS import (see Helper Files WS3 skill)."""

from __future__ import annotations

import csv
import logging
import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import db
from importer import resolve_xlsx_path

logger = logging.getLogger(__name__)

SKIP_COL0 = {
    "Customer Mail Details",
    "Name of Customer",
    "Profile Name",
    "Report:",
    "Entry:",
    "Sort:",
    "Date:",
}
RATE_TYPES = frozenset({"ThreeDigitAuto", "ADC Auto", "MXD ADC Auto", "Single Piece"})
TOTAL_LABELS = frozenset({"Profile Total", "Customer Total"})


def parse_date_from_filename(path: str) -> str | None:
    """YYYY-MM-DD from WS3_FCFL_CustomerMailDetail[_ ]M-D-YY.xls"""
    base = os.path.basename(path)
    m = re.search(r"(\d{1,2})-(\d{1,2})-(\d{2,4})\.(?:xls|xlsx)$", base, re.IGNORECASE)
    if not m:
        return None
    month, day, year = m.group(1), m.group(2), m.group(3)
    if len(year) == 2:
        year = "20" + year
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def parse_mail_id_date(mail_id: str | None) -> str | None:
    """
    Parse NetSort Mail ID (row 8 col BI) to YYYY-MM-DD.

    Input examples: '040626_F', '041726_F'
    """
    if mail_id is None:
        return None
    s = str(mail_id).strip()
    if not s:
        return None
    date_part = s.split("_", 1)[0].strip()
    if len(date_part) != 6 or not date_part.isdigit():
        return None
    month = date_part[0:2]
    day = date_part[2:4]
    yy = date_part[4:6]
    year = f"20{yy}"
    try:
        datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
    except ValueError:
        return None
    return f"{year}-{month}-{day}"


def read_bi8_mail_id(xlsx_path: str) -> str | None:
    """Mail ID string from cell BI8 (e.g. '040626_F')."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        v = ws["BI8"].value
    finally:
        wb.close()
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def read_mail_date_from_bi8(xlsx_path: str) -> str | None:
    """Authoritative WS3 mail_date derived from BI8 Mail ID (MMDDYY_F)."""
    mid = read_bi8_mail_id(xlsx_path)
    return parse_mail_id_date(mid)


def parse_currency(val: Any) -> float | None:
    if val is None:
        return None
    s = re.sub(r"^[^\d\-]+", "", str(val).strip()).replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val).strip().replace(",", "")))
    except ValueError:
        return None


def calc_cost_per_piece(postage_applied: float | None, num_pieces: int | None) -> float | None:
    if postage_applied is None or not num_pieces or num_pieces == 0:
        return None
    return round(float(postage_applied) / int(num_pieces), 4)


def calc_usps_cost_per_piece(postage_claimed: float | None, num_pieces: int | None) -> float | None:
    if postage_claimed is None or not num_pieces or num_pieces == 0:
        return None
    return round(float(postage_claimed) / int(num_pieces), 4)


def parse_ws3_xlsx(xlsx_path: str) -> tuple[str | None, str | None, dict[str, str], list[dict[str, Any]]]:
    """Returns mail_id, run_datetime, customers dict, detail rows."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    mail_id = None
    run_datetime = None
    customers: dict[str, str] = {}
    rows: list[dict[str, Any]] = []

    current_customer_name: str | None = None
    current_customer_code: str | None = None
    current_profile_name: str | None = None

    for raw_row in ws.iter_rows(values_only=True):
        c = raw_row

        def g(i: int) -> Any:
            try:
                return c[i]
            except (IndexError, TypeError):
                return None

        col0 = str(g(0) or "").strip()
        col1 = str(g(1) or "").strip()
        col5 = str(g(5) or "").strip()
        col7 = str(g(7) or "").strip()
        col19 = str(g(19) or "").strip()
        col21 = str(g(21) or "").strip()

        if g(48) in ("Mail ID :", "Mail ID:"):
            mail_id = str(g(57) or "").strip()
        if g(84) == "Date:" and g(90):
            run_datetime = str(g(90)).strip()

        if col5 == "Customer Mail Details":
            continue
        if col0 in SKIP_COL0:
            continue
        if col21 in TOTAL_LABELS:
            continue

        if (
            col1
            and str(g(14) or "").strip().startswith("$")
            and str(g(16) or "").strip() == "Metered"
        ):
            current_customer_name = col1
            current_customer_code = None
            current_profile_name = None
            continue

        if col7 and re.match(r"^\d{6}\s", col7):
            current_profile_name = col7
            m = re.match(r"^(\d{6})", col7)
            current_customer_code = m.group(1) if m else None
            if current_customer_code and current_customer_name:
                customers.setdefault(current_customer_code, current_customer_name)
            # No `continue`: this NetSort layout puts the first rate type on the
            # same row as the profile name, so fall through to the rate block.

        if col19 in RATE_TYPES and current_customer_code:
            is_single = col19 == "Single Piece"
            num_pieces = parse_int(g(66))
            pcs_accepted = parse_int(g(75)) if is_single else parse_int(g(70))
            if is_single:
                pcs_rejected = parse_int(g(75))
            else:
                pcs_rejected = (
                    (num_pieces - pcs_accepted)
                    if (num_pieces is not None and pcs_accepted is not None)
                    else None
                )
            if pcs_rejected is not None and pcs_rejected < 0:
                pcs_rejected = 0
            postage_applied = parse_currency(g(53))
            rows.append(
                {
                    "customer_code": current_customer_code,
                    "profile_name": current_profile_name or "",
                    "rate_type": col19,
                    "postage_claimed": parse_currency(g(40)),
                    "postage_applied": postage_applied,
                    "num_pieces": num_pieces,
                    "pcs_accepted": pcs_accepted,
                    "pcs_rejected": pcs_rejected,
                    "cost_per_piece": calc_cost_per_piece(postage_applied, num_pieces),
                }
            )

    wb.close()
    return mail_id, run_datetime, customers, rows


def _ensure_profile_id(conn: sqlite3.Connection, profile_name: str) -> int:
    conn.execute(
        "INSERT OR IGNORE INTO ws3_profiles (profile_name) VALUES (?)",
        (profile_name,),
    )
    r = conn.execute(
        "SELECT id FROM ws3_profiles WHERE profile_name = ?", (profile_name,)
    ).fetchone()
    return int(r[0])


def write_mail_detail_export_csv(conn: sqlite3.Connection, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "run_id",
                "mail_date",
                "mail_id",
                "customer_code",
                "customer_name",
                "profile_name",
                "rate_type",
                "postage_claimed",
                "postage_applied",
                "num_pieces",
                "pcs_accepted",
                "pcs_rejected",
                "cost_per_piece",
            ]
        )
        for row in conn.execute(
            """
            SELECT d.run_id, r.mail_date, r.mail_id,
                   d.customer_code, nc.customer_name,
                   p.profile_name, d.rate_type,
                   d.postage_claimed, d.postage_applied,
                   d.num_pieces, d.pcs_accepted, d.pcs_rejected, d.cost_per_piece
            FROM ws3_mail_detail d
            JOIN ws3_mail_runs r ON r.run_id = d.run_id
            JOIN ws3_netsort_customers nc ON nc.customer_code = d.customer_code
            JOIN ws3_profiles p ON p.id = d.profile_id
            ORDER BY r.mail_date, d.customer_code, d.id
            """
        ):
            w.writerow(list(row))


def process_ws3_mail_detail_file(
    xls_path: str,
    db_path: Path | str,
    csv_out_path: Path | str | None = None,
) -> dict[str, Any]:
    """
    Convert .xls to .xlsx, parse, insert into postage.db, refresh mail_detail_export.csv.
    Skips if file_name already in ws3_imports. Skips if (mail_date, mail_id) already exists.
    """
    db_path = Path(db_path)
    file_name = os.path.basename(xls_path)

    root = Path(__file__).resolve().parent
    if csv_out_path is None:
        csv_out_path = root / "reports" / "mail_detail_export.csv"
    else:
        csv_out_path = Path(csv_out_path)

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")

    dup = conn.execute(
        "SELECT 1 FROM ws3_imports WHERE file_name = ?", (file_name,)
    ).fetchone()
    if dup:
        conn.close()
        return {
            "skipped": True,
            "reason": "already_imported",
            "file_name": file_name,
        }

    xlsx_path = resolve_xlsx_path(xls_path)

    bi_mail_id = read_bi8_mail_id(xlsx_path)
    mail_date = parse_mail_id_date(bi_mail_id)
    if not mail_date:
        raise ValueError(
            f"Cannot determine WS3 mail date from BI8 Mail ID (expected MMDDYY_F): {file_name!r} BI8={bi_mail_id!r}"
        )

    mail_id, run_dt, customers, detail_rows = parse_ws3_xlsx(xlsx_path)
    mail_id = mail_id or ""
    if bi_mail_id and mail_id and bi_mail_id.strip() != mail_id.strip():
        logger.warning(
            "WS3 Mail ID mismatch: BI8=%r parse_ws3_xlsx=%r file=%r",
            bi_mail_id,
            mail_id,
            file_name,
        )

    existing_run = conn.execute(
        "SELECT run_id FROM ws3_mail_runs WHERE mail_date = ? AND mail_id = ?",
        (mail_date, mail_id),
    ).fetchone()
    if existing_run:
        conn.close()
        return {
            "skipped": True,
            "reason": "duplicate_mail_run",
            "file_name": file_name,
            "mail_date": mail_date,
            "mail_id": mail_id,
        }

    try:
        for code, name in customers.items():
            conn.execute(
                """
                INSERT OR IGNORE INTO ws3_netsort_customers (customer_code, customer_name)
                VALUES (?, ?)
                """,
                (code, name),
            )

        cur = conn.execute(
            """
            INSERT INTO ws3_mail_runs (mail_date, mail_id, run_datetime, source_file_name)
            VALUES (?, ?, ?, ?)
            """,
            (mail_date, mail_id, run_dt, file_name),
        )
        run_id = int(cur.lastrowid)

        insert_detail: list[tuple[Any, ...]] = []
        for r in detail_rows:
            pid = _ensure_profile_id(conn, r["profile_name"])
            insert_detail.append(
                (
                    run_id,
                    pid,
                    r["customer_code"],
                    r["rate_type"],
                    r["postage_claimed"],
                    r["postage_applied"],
                    r["num_pieces"],
                    r["pcs_accepted"],
                    r["pcs_rejected"],
                    r["cost_per_piece"],
                    calc_usps_cost_per_piece(r["postage_claimed"], r["num_pieces"]),
                )
            )

        conn.executemany(
            """
            INSERT INTO ws3_mail_detail (
                run_id, profile_id, customer_code, rate_type,
                postage_claimed, postage_applied, num_pieces,
                pcs_accepted, pcs_rejected, cost_per_piece, usps_cost_per_piece
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            insert_detail,
        )

        conn.execute(
            """
            INSERT INTO ws3_imports (file_name, mail_date, run_id, row_count)
            VALUES (?, ?, ?, ?)
            """,
            (file_name, mail_date, run_id, len(detail_rows)),
        )

        db.recompute_ws3_parent_rejects_for_mail_dates(conn, [mail_date])
        conn.commit()

        write_mail_detail_export_csv(conn, csv_out_path)
    except Exception:
        conn.rollback()
        conn.close()
        raise

    conn.close()
    return {
        "skipped": False,
        "file_name": file_name,
        "mail_date": mail_date,
        "mail_id": mail_id,
        "run_id": run_id,
        "rows_imported": len(detail_rows),
        "csv_path": str(csv_out_path),
    }
