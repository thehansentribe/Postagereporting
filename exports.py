"""Excel exports: postage invoice and parcel (BC Priority) report."""

from __future__ import annotations

from collections import defaultdict
import csv
from functools import cmp_to_key
import os
import re
import shutil
import sqlite3
import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from allocations import allocate_integer_proportional
import db

# Re-export from db (single source of truth).
POSTAGE_INVOICE_FLAT_MAIL_CLASSES = db.POSTAGE_INVOICE_FLAT_MAIL_CLASSES
POSTAGE_INVOICE_FLAT_MAIL_SQL_IN = db.POSTAGE_INVOICE_FLAT_MAIL_SQL_IN

DEFAULT_FLAT_RETAIL_BY_OZ: dict[int, float] = {
    1: 1.63,
    2: 1.9,
    3: 2.17,
    4: 2.44,
    5: 2.72,
    6: 3.0,
    7: 3.28,
    8: 3.56,
    9: 3.84,
    10: 4.14,
    11: 4.44,
    12: 4.74,
    13: 5.04,
}


def _retail_rate(rates: dict[int, float], oz: int) -> float:
    v = float(rates.get(oz) or 0.0)
    if v > 0:
        return v
    return float(DEFAULT_FLAT_RETAIL_BY_OZ.get(oz, 0.0))


def _row_value(row: Any, key: str, default: str = "") -> str:
    """sqlite3.Row has no .get; read key safely for export rows."""
    try:
        v = row[key]
    except (KeyError, TypeError, IndexError):
        return default
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _apply_bold_preserve_font(cell) -> None:
    f = cell.font
    if f:
        nf = copy(f)
        nf.bold = True
        cell.font = nf
    else:
        cell.font = Font(bold=True)


CUSTOMER_CONTACTS: dict[int, dict[str, str]] = {
    3901: {
        "contact_name": "Chris Torrez",
        "address1": "1133 S.W. Topeka Blvd.",
        "city_state_zip": "Topeka, KS 66629-0001",
        "phone": "785-291-8681",
        "fax": "785-291-8548",
        "email": "Chris.Torrez@bcbsks.com",
        "customer_id": "            1st 0012",
    }
}


def _sheet_title_for_date(file_date: str) -> str:
    dt = datetime.strptime(file_date, "%Y-%m-%d")
    return f"{dt.strftime('%b')} {dt.day} {dt.year}"


def _invoice_range_sheet_title(start_date: str, end_date: str) -> str:
    """Workbook sheet name for a date range (max 31 chars for Excel)."""
    s = datetime.strptime(start_date, "%Y-%m-%d")
    e = datetime.strptime(end_date, "%Y-%m-%d")
    if start_date == end_date:
        return _sheet_title_for_date(start_date)[:31]
    label = f"{s.strftime('%b')} {s.day} – {e.strftime('%b')} {e.day} {e.year}"
    return label[:31]


def _efd_for_oz(rates: dict[int, float], oz: int, discount: float) -> float:
    return round(max(0.0, _retail_rate(rates, oz) - discount), 4)


def _invoice_oz_parent_lines(
    wd: dict[str, Any],
    rates: dict[int, float],
    oz: int,
    discount: float,
) -> tuple[float, float]:
    """
    Aggregated parent row: I = max(0, efd - ws3), K = imb + ws3.
    Column J matches retail (same as column G): customer line = H*I + G*K.
    Returns (customer_line, retail_line) for that weight.
    """
    retail = float(_retail_rate(rates, oz))
    efd_h = round(max(0.0, retail - float(discount)), 4)
    efd_pieces = int(wd.get("efd_pieces", wd.get("pieces", 0)) or 0)
    imb = int(wd.get("reject_pieces", 0) or 0)
    ws3 = int(wd.get("ws3_pieces", 0) or 0)
    i_adj = max(0, efd_pieces - ws3)
    k_tot = imb + ws3
    cust = round(efd_h * float(i_adj) + retail * float(k_tot), 2)
    ret_line = round(retail * float(i_adj + k_tot), 2)
    return cust, ret_line


def _allocate_ws3_rejects_joint(
    cur: sqlite3.Cursor,
    scope_range_sql: str,
    scope_range_params: list[Any],
    parent_number: int,
    customer_number: int | None,
    ws3_total: int,
) -> tuple[dict[tuple[int, int], int], dict[int, int], dict[int, int]]:
    """
    Allocate WS3 reject pieces across (customer_number, weight_oz) using joint shares
    proportional to invoice flat volume (1CA5DFlt + other invoice flats) per cell.

    Returns:
      alloc: (customer_number, woz) -> ws3 piece count
      ws3_by_oz: woz -> count (1..13)
      ws3_by_dept: customer_number -> count
    """
    if ws3_total <= 0:
        return {}, {}, {}

    efd_cls = db.INVOICE_EFD_FLAT_MAIL_CLASS
    flat_rows = cur.execute(
        f"""
        SELECT c.customer_number, c.customer_name,
               CAST(ROUND(p.weight_oz) AS INTEGER) AS woz,
               SUM(CASE WHEN p.mail_class = '{efd_cls}' THEN p.pieces ELSE 0 END) AS efd_pieces,
               SUM(CASE WHEN p.mail_class <> '{efd_cls}' THEN p.pieces ELSE 0 END) AS rej_pieces
        FROM postage_data p
        JOIN customers c ON p.account_code = c.customer_number
        WHERE {scope_range_sql}
          AND p.weight_oz BETWEEN 1 AND 13
          AND p.mail_class IN {POSTAGE_INVOICE_FLAT_MAIL_SQL_IN}
        GROUP BY c.customer_number, c.customer_name, CAST(ROUND(p.weight_oz) AS INTEGER)
        """,
        scope_range_params,
    ).fetchall()

    pairs: list[tuple[int, int, float]] = []
    for row in flat_rows:
        cn = int(row["customer_number"])
        woz = int(row["woz"] or 0)
        if not (1 <= woz <= 13):
            continue
        w = float(int(row["efd_pieces"] or 0) + int(row["rej_pieces"] or 0))
        if w > 0:
            pairs.append((cn, woz, w))

    if not pairs:
        # No flat volume to apportion against: assign entire WS3 total to parent at 1 oz.
        alloc: dict[tuple[int, int], int] = {(parent_number, 1): ws3_total}
        return alloc, {1: ws3_total}, {parent_number: ws3_total}

    weights = [p[2] for p in pairs]
    amounts = allocate_integer_proportional(ws3_total, weights)
    alloc = {}
    ws3_by_oz: dict[int, int] = defaultdict(int)
    ws3_by_dept: dict[int, int] = defaultdict(int)
    for i, (cn, woz, _) in enumerate(pairs):
        a = int(amounts[i])
        if a:
            alloc[(cn, woz)] = a
            ws3_by_oz[woz] += a
            ws3_by_dept[cn] += a
    return alloc, dict(ws3_by_oz), dict(ws3_by_dept)


def _allocate_presort_rejects_by_day_efd_only(
    cur: sqlite3.Cursor,
    scope_range_sql: str,
    scope_range_params: list[Any],
    *,
    rejects_by_day: dict[str, int],
) -> tuple[dict[tuple[int, int], int], dict[int, int], dict[int, int]]:
    """
    Allocate presort rejects by day across (customer_number, weight_oz) using weights from that
    day's INVOICE_EFD_FLAT_MAIL_CLASS (`1CA5DFlt`) volume only.

    This makes multi-day ranges consistent with the sum of per-day allocations.
    """
    efd_cls = db.INVOICE_EFD_FLAT_MAIL_CLASS
    alloc: dict[tuple[int, int], int] = defaultdict(int)
    by_oz: dict[int, int] = defaultdict(int)
    by_dept: dict[int, int] = defaultdict(int)

    for d, total in (rejects_by_day or {}).items():
        day_total = int(total or 0)
        if day_total <= 0:
            continue
        day_rows = cur.execute(
            f"""
            SELECT c.customer_number,
                   CAST(ROUND(p.weight_oz) AS INTEGER) AS woz,
                   SUM(p.pieces) AS pcs
            FROM postage_data p
            JOIN customers c ON p.account_code = c.customer_number
            WHERE {scope_range_sql}
              AND p.file_date = ?
              AND p.mail_class = ?
              AND p.weight_oz BETWEEN 1 AND 13
            GROUP BY c.customer_number, CAST(ROUND(p.weight_oz) AS INTEGER)
            """,
            [*scope_range_params, str(d), efd_cls],
        ).fetchall()

        pairs: list[tuple[int, int, int]] = []
        for r in day_rows:
            cn = int(r["customer_number"])
            woz = int(r["woz"] or 0)
            pcs = int(r["pcs"] or 0)
            if pcs <= 0 or not (1 <= woz <= 13):
                continue
            pairs.append((cn, woz, pcs))

        if not pairs:
            continue

        cap = sum(p[2] for p in pairs)
        alloc_total = min(day_total, cap)
        if alloc_total <= 0:
            continue

        weights = [float(p[2]) for p in pairs]
        amounts = allocate_integer_proportional(int(alloc_total), weights)
        for (cn, woz, _), a in zip(pairs, amounts):
            aa = int(a or 0)
            if aa <= 0:
                continue
            alloc[(cn, woz)] += aa
            by_oz[woz] += aa
            by_dept[cn] += aa

    return dict(alloc), dict(by_oz), dict(by_dept)


def _cost_center_row_all_zeros(child: dict[str, Any]) -> bool:
    return (
        int(child.get("pieces") or 0) == 0
        and int(child.get("rejects") or 0) == 0
        and abs(float(child.get("cost") or 0.0)) < 1e-9
        and abs(float(child.get("savings") or 0.0)) < 1e-9
    )


def _cost_centers_flats_range(
    cur: sqlite3.Cursor,
    scope_range_sql: str,
    scope_range_params: list[Any],
    parent_number: int,
    customer_number: int | None,
    rates: dict[int, float],
    discount: float,
    *,
    ws3_by_oz: dict[int, int],
    ws3_alloc: dict[tuple[int, int], int],
    remove_zeros: bool = False,
) -> list[dict[str, Any]]:
    """
    Per-customer charges mirror the invoice weight grid: I = max(0, efd - ws3), K = imb + ws3,
    J = retail (same as G); costs per oz are scaled so they sum to the parent row customer total
    for that oz (same as H*I + G*K on the sheet). Savings = (pieces - rejects) * (1oz retail - 1oz FC),
    matching the invoice cost-center formula (C-D)*($G$14-$H$14).
    """
    efd_cls = db.INVOICE_EFD_FLAT_MAIL_CLASS
    flat_rows = cur.execute(
        f"""
        SELECT c.customer_number, c.customer_name,
               CAST(ROUND(p.weight_oz) AS INTEGER) AS woz,
               SUM(CASE WHEN p.mail_class = '{efd_cls}' THEN p.pieces ELSE 0 END) AS efd_pieces,
               SUM(CASE WHEN p.mail_class <> '{efd_cls}' THEN p.pieces ELSE 0 END) AS rej_pieces
        FROM postage_data p
        JOIN customers c ON p.account_code = c.customer_number
        WHERE {scope_range_sql}
          AND p.weight_oz BETWEEN 1 AND 13
          AND p.mail_class IN {POSTAGE_INVOICE_FLAT_MAIL_SQL_IN}
        GROUP BY c.customer_number, c.customer_name, CAST(ROUND(p.weight_oz) AS INTEGER)
        """,
        scope_range_params,
    ).fetchall()

    agg: dict[int, dict[str, Any]] = defaultdict(
        lambda: {"customer_name": "", "by_oz": defaultdict(lambda: {"efd": 0, "rej": 0})}
    )
    parent_oz: dict[int, dict[str, int]] = defaultdict(lambda: {"efd": 0, "imb": 0})
    for row in flat_rows:
        cn = int(row["customer_number"])
        woz = int(row["woz"] or 0)
        if not (1 <= woz <= 13):
            continue
        epc = int(row["efd_pieces"] or 0)
        rpc = int(row["rej_pieces"] or 0)
        agg[cn]["customer_name"] = _row_value(row, "customer_name")
        agg[cn]["by_oz"][woz]["efd"] += epc
        agg[cn]["by_oz"][woz]["rej"] += rpc
        parent_oz[woz]["efd"] += epc
        parent_oz[woz]["imb"] += rpc

    # Per (customer, oz) raw customer and retail lines (before scaling to parent row).
    raw_cust: dict[tuple[int, int], float] = {}
    raw_sum_oz: dict[int, float] = defaultdict(float)
    for row in flat_rows:
        cn = int(row["customer_number"])
        oz_i = int(row["woz"] or 0)
        if not (1 <= oz_i <= 13):
            continue
        e = int(row["efd_pieces"] or 0)
        rj = int(row["rej_pieces"] or 0)
        w3 = int(ws3_alloc.get((cn, oz_i), 0))
        i_d = max(0, e - w3)
        k_d = rj + w3
        retail = float(_retail_rate(rates, oz_i))
        efd_h = round(max(0.0, retail - float(discount)), 4)
        rc = efd_h * float(i_d) + retail * float(k_d)
        raw_cust[(cn, oz_i)] = rc
        raw_sum_oz[oz_i] += rc

    scale_oz: dict[int, float] = {}
    for oz_i in range(1, 14):
        po = parent_oz.get(oz_i)
        w3_oz = int(ws3_by_oz.get(oz_i, 0) or 0)
        wd_parent: dict[str, Any] = {
            "efd_pieces": int(po["efd"]) if po else 0,
            "reject_pieces": int(po["imb"]) if po else 0,
            "ws3_pieces": w3_oz,
        }
        l_parent, _ = _invoice_oz_parent_lines(wd_parent, rates, oz_i, discount)
        rs = float(raw_sum_oz.get(oz_i, 0.0))
        if rs > 1e-9:
            scale_oz[oz_i] = float(l_parent) / rs
        else:
            scale_oz[oz_i] = 0.0

    cust_sql = """
        SELECT customer_number, customer_name FROM customers WHERE parent_number = ?
    """
    cust_params: list[Any] = [parent_number]
    if customer_number is not None:
        cust_sql += " AND customer_number = ?"
        cust_params.append(customer_number)
    cust_sql += """
        UNION ALL
        SELECT customer_number, customer_name FROM customers WHERE customer_number = ?
        ORDER BY customer_number
    """
    cust_params.append(parent_number)

    ordered = cur.execute(cust_sql, cust_params).fetchall()

    out: list[dict[str, Any]] = []
    for row in ordered:
        cn = int(row["customer_number"])
        name = _row_value(row, "customer_name")
        data = agg.get(cn)
        if not data:
            out.append(
                {
                    "customer_number": cn,
                    "customer_name": name,
                    "pieces": 0,
                    "rejects": 0,
                    "cost": 0.0,
                    "savings": 0.0,
                }
            )
            continue
        by_oz = data["by_oz"]
        flat_pieces = 0
        rejects_acc = 0
        cost_acc = 0.0
        for oz, counts in by_oz.items():
            oz_i = int(oz)
            e = int(counts.get("efd") or 0)
            rj = int(counts.get("rej") or 0)
            w3 = int(ws3_alloc.get((cn, oz_i), 0))
            flat_pieces += e + rj
            rejects_acc += rj + w3
            sc = scale_oz.get(oz_i, 0.0)
            rc = float(raw_cust.get((cn, oz_i), 0.0))
            cost_acc += rc * sc

        cost_acc = round(cost_acc, 2)
        g1 = float(_retail_rate(rates, 1))
        h1 = float(_efd_for_oz(rates, 1, discount))
        savings_per_1oz = round(g1 - h1, 4)
        savings = round(float(flat_pieces - rejects_acc) * savings_per_1oz, 2)
        out.append(
            {
                "customer_number": cn,
                "customer_name": data["customer_name"] or name,
                "pieces": flat_pieces,
                "rejects": rejects_acc,
                "cost": cost_acc,
                "savings": savings,
            }
        )
    if remove_zeros:
        out = [c for c in out if not _cost_center_row_all_zeros(c)]
    return out


# Flats data grid XLSX — column keys/labels aligned with `postageTableColumns` in static/app.js.
def _flats_grid_header_keys_and_labels(
    hide_costs: bool, hide_customer_numbers: bool = True
) -> tuple[list[str], list[str]]:
    oz_keys = [f"oz_{i}" for i in range(13)] + ["oz_13", "oz_13plus"]
    keys: list[str] = ["date", "parent_name", "child_name"]
    labels: list[str] = ["Date", "Parent Name", "Child Name"]
    if not hide_customer_numbers:
        keys.append("child_number")
        labels.append("Child Number")
    keys.extend(["mail_class", *oz_keys, "total_qty"])
    labels.extend(
        [
            "Class",
            *[f"{i} oz" for i in range(13)],
            "13 oz",
            "13+ oz",
            "Total Qty",
        ]
    )
    if not hide_costs:
        keys.append("total_cost")
        labels.append("Total Cost")
    return keys, labels


_FLATS_GRID_ALL_KEYS = set(_flats_grid_header_keys_and_labels(False, False)[0])


def _flats_grid_cell_value(row: dict[str, Any], k: str) -> Any:
    v = row.get(k)
    if k == "total_cost":
        if v is None or v == "":
            return None
        return round(float(v), 2)
    if k.startswith("oz_") or k == "total_qty":
        return int(v or 0)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return v
    return v if v is not None else ""


def _flats_grid_cmp(key: str, mul: int):
    """Comparator aligned with `sortRows` in static/app.js (nulls last for mul=1)."""

    def cmp(a: dict[str, Any], b: dict[str, Any]) -> int:
        va, vb = a.get(key), b.get(key)
        if va is None and vb is None:
            return 0
        if va is None:
            return 1 * mul
        if vb is None:
            return -1 * mul
        if (
            isinstance(va, (int, float))
            and not isinstance(va, bool)
            and isinstance(vb, (int, float))
            and not isinstance(vb, bool)
        ):
            if va < vb:
                return -mul
            if va > vb:
                return mul
            return 0
        sa, sb = str(va), str(vb)
        try:
            na, nb = float(sa), float(sb)
            if na < nb:
                return -mul
            if na > nb:
                return mul
            return 0
        except ValueError:
            if sa < sb:
                return -mul
            if sa > sb:
                return mul
            return 0

    return cmp


def _sort_flats_grid_rows(
    rows: list[dict[str, Any]], sort_key: str, sort_dir: int
) -> list[dict[str, Any]]:
    key = sort_key if sort_key in _FLATS_GRID_ALL_KEYS else "date"
    mul = 1 if sort_dir >= 0 else -1
    return sorted(rows, key=cmp_to_key(_flats_grid_cmp(key, mul)))


def export_flats_data_grid_xlsx(
    start_date: str,
    end_date: str,
    parent_number: int | None = None,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
    consolidate: bool = False,
    remove_zeros: bool = False,
    hide_costs: bool = False,
    hide_customer_numbers: bool = True,
    allocate_presort_rejects: bool = False,
    sort_key: str = "date",
    sort_dir: int = 1,
) -> Path:
    """
    One-sheet workbook: postage dashboard rows (same columns as the former flats CSV export).
    """
    conn = db.get_connection()
    try:
        data = db.query_postage(
            conn,
            start_date,
            end_date,
            parent_number=parent_number,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            consolidate=consolidate,
            remove_zeros=remove_zeros,
            hide_costs=hide_costs,
            allocate_presort_rejects=allocate_presort_rejects,
        )
        rows_raw = list(data.get("rows") or [])
        rows = _sort_flats_grid_rows(rows_raw, sort_key, sort_dir)

        header_keys, headers = _flats_grid_header_keys_and_labels(
            hide_costs, hide_customer_numbers
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Flats"

        INT_FMT = "#,##0"
        CURR_FMT = "$#,##0.00"
        bold = Font(bold=True)

        for col, h in enumerate(headers, start=1):
            c = ws.cell(1, col, h)
            c.font = bold

        oz_key_set = {k for k in header_keys if k.startswith("oz_")}
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, k in enumerate(header_keys, start=1):
                val = _flats_grid_cell_value(row, k)
                cell = ws.cell(r_idx, c_idx, val)
                if k == "total_cost" and val is not None:
                    cell.number_format = CURR_FMT
                elif k in oz_key_set or k == "total_qty":
                    cell.number_format = INT_FMT

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(18, 12 + (2 if col <= 4 else 0))

        out = Path(
            tempfile.mkstemp(
                suffix=f"_Flats_Invoice_{start_date}_{end_date}.xlsx",
                prefix="flats_grid_",
            )[1]
        )
        wb.save(out)
        return out
    finally:
        conn.close()


def export_flats_data_grid_csv(
    start_date: str,
    end_date: str,
    parent_number: int | None = None,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
    consolidate: bool = False,
    remove_zeros: bool = False,
    hide_costs: bool = False,
    hide_customer_numbers: bool = True,
    allocate_presort_rejects: bool = False,
    sort_key: str = "date",
    sort_dir: int = 1,
) -> Path:
    """
    CSV export of the Flats data grid (same columns/order as the UI grid).
    """
    conn = db.get_connection()
    try:
        data = db.query_postage(
            conn,
            start_date,
            end_date,
            parent_number=parent_number,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            consolidate=consolidate,
            remove_zeros=remove_zeros,
            hide_costs=hide_costs,
            allocate_presort_rejects=allocate_presort_rejects,
        )
        rows_raw = list(data.get("rows") or [])
        rows = _sort_flats_grid_rows(rows_raw, sort_key, sort_dir)
        header_keys, headers = _flats_grid_header_keys_and_labels(
            hide_costs, hide_customer_numbers
        )

        out = Path(
            tempfile.mkstemp(
                suffix=f"_Flats_Report_{start_date}_{end_date}.csv",
                prefix="flats_grid_",
            )[1]
        )
        with open(out, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for row in rows:
                w.writerow([_flats_grid_cell_value(row, k) for k in header_keys])
        return out
    finally:
        conn.close()


def fill_postage_invoice_worksheet(
    ws,
    conn: sqlite3.Connection,
    parent_number: int,
    start_date: str,
    end_date: str,
    *,
    discount: float = 0.10,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
    remove_zeros: bool = False,
    hide_costs: bool = False,
    hide_savings: bool = False,
    set_sheet_title: str | None = None,
    sheet_idx: int = 1,
    empty_invoice_if_no_data: bool = False,
) -> bool:
    """
    Write postage invoice layout onto an existing worksheet.

    Returns True when a full invoice was written; False when only a no-data message
    was written (standalone export). When ``empty_invoice_if_no_data`` is True, writes
    a zeroed invoice layout so cross-sheet summary formulas still resolve.
    """
    cur = conn.cursor()

    rates: dict[int, float] = {}
    flat_view = db.get_flat_rate_costs(conn, as_of_date=end_date)
    for row in flat_view["rows"]:
        w = int(row["weight_not_over_oz"])
        rates[w] = float(row["rate_retail"] or 0)

    scope_range_sql, scope_range_params = db.postage_scope_where_clause(
        start_date,
        end_date,
        parent_number,
        customer_number,
        show_parents,
        show_main,
    )

    has_postage = cur.execute(
        f"""
        SELECT 1 FROM postage_data p
        JOIN customers c ON p.account_code = c.customer_number
        WHERE {scope_range_sql}
        LIMIT 1
        """,
        scope_range_params,
    ).fetchone()

    parent = cur.execute(
        "SELECT customer_name FROM customers WHERE customer_number = ?",
        (parent_number,),
    ).fetchone()
    parent_name = parent["customer_name"] if parent else f"Account {parent_number}"
    contact = CUSTOMER_CONTACTS.get(parent_number, {})

    if set_sheet_title is not None:
        ws.title = set_sheet_title[:31]

    if not has_postage and not empty_invoice_if_no_data:
        ws["A1"] = "No postage data in range for this parent."
        return False

    if not has_postage:
        weight_data: dict[int, dict[str, Any]] = {}
        other_pieces = 0
        other_cost = 0.0
        range_total_pieces = 0
        range_total_retail = 0.0
        children: list = []
    else:
        efd_cls = db.INVOICE_EFD_FLAT_MAIL_CLASS
        weight_data = {}
        for row in cur.execute(
            f"""
            SELECT CAST(ROUND(p.weight_oz) AS INTEGER) AS woz,
                   SUM(CASE WHEN p.mail_class = '{efd_cls}' THEN p.pieces ELSE 0 END) AS efd_pieces,
                   SUM(CASE WHEN p.mail_class <> '{efd_cls}' THEN p.pieces ELSE 0 END) AS reject_pieces,
                   SUM(p.total_cost) AS cost
            FROM postage_data p
            JOIN customers c ON p.account_code = c.customer_number
            WHERE {scope_range_sql}
              AND CAST(ROUND(p.weight_oz) AS INTEGER) BETWEEN 1 AND 13
              AND p.mail_class IN {POSTAGE_INVOICE_FLAT_MAIL_SQL_IN}
            GROUP BY CAST(ROUND(p.weight_oz) AS INTEGER)
            """,
            scope_range_params,
        ):
            woz = int(row["woz"] or 0)
            if not (1 <= woz <= 13):
                continue
            efd_pc = int(row["efd_pieces"] or 0)
            rej_pc = int(row["reject_pieces"] or 0)
            tc = float(row["cost"] or 0.0)
            if woz not in weight_data:
                weight_data[woz] = {
                    "efd_pieces": 0,
                    "reject_pieces": 0,
                    "pieces": 0,
                    "cost": 0.0,
                }
            weight_data[woz]["efd_pieces"] += efd_pc
            weight_data[woz]["reject_pieces"] += rej_pc
            weight_data[woz]["pieces"] += efd_pc + rej_pc
            weight_data[woz]["cost"] += tc

        other_row = cur.execute(
            f"""
            SELECT COALESCE(SUM(p.pieces), 0) AS pieces,
                   COALESCE(SUM(p.total_cost), 0) AS cost
            FROM postage_data p
            JOIN customers c ON p.account_code = c.customer_number
            WHERE {scope_range_sql}
              AND (p.weight_oz IS NULL OR p.weight_oz <= 13)
              AND (
                (p.mail_class IS NULL OR p.mail_class NOT IN {POSTAGE_INVOICE_FLAT_MAIL_SQL_IN})
                OR (
                  p.mail_class IN {POSTAGE_INVOICE_FLAT_MAIL_SQL_IN}
                  AND (p.weight_oz IS NULL OR p.weight_oz < 1)
                )
              )
            """,
            scope_range_params,
        ).fetchone()
        other_pieces = int(other_row["pieces"] or 0)
        other_cost = float(other_row["cost"] or 0.0)

        range_totals = db.query_postage(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            consolidate=False,
            remove_zeros=remove_zeros,
            hide_costs=False,
        )
        range_total_pieces = int(range_totals["total_pieces"] or 0)
        range_total_retail = float(range_totals.get("total_retail_cost") or 0.0)

        rejects_by_day = db.query_total_presort_reject_counts_by_day_for_invoice(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
        )
        ws3_alloc, ws3_by_oz, _ws3_by_dept = _allocate_presort_rejects_by_day_efd_only(
            cur,
            scope_range_sql,
            list(scope_range_params),
            rejects_by_day=rejects_by_day,
        )
        for oz, wc in ws3_by_oz.items():
            if int(wc or 0) <= 0:
                continue
            if oz not in weight_data:
                weight_data[oz] = {
                    "efd_pieces": 0,
                    "reject_pieces": 0,
                    "pieces": 0,
                    "cost": 0.0,
                }
            weight_data[oz]["ws3_pieces"] = int(wc)

        children = _cost_centers_flats_range(
            cur,
            scope_range_sql,
            list(scope_range_params),
            parent_number,
            customer_number,
            rates,
            discount,
            ws3_by_oz=ws3_by_oz,
            ws3_alloc=ws3_alloc,
            remove_zeros=remove_zeros,
        )

    period_end = datetime.strptime(end_date, "%Y-%m-%d")
    _write_invoice_sheet(
        ws,
        sheet_idx,
        period_end,
        start_date,
        end_date,
        parent_number,
        parent_name,
        contact,
        rates,
        weight_data,
        children,
        discount,
        other_pieces,
        other_cost,
        range_total_pieces,
        range_total_retail=range_total_retail,
        hide_costs=hide_costs,
        hide_savings=hide_savings,
    )
    return True


def export_postage_invoice(
    parent_number: int,
    start_date: str,
    end_date: str,
    discount: float = 0.10,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
    remove_zeros: bool = False,
    hide_costs: bool = False,
    hide_savings: bool = False,
) -> Path:
    conn = db.get_connection()
    try:
        wb = Workbook()
        wb.remove(wb.active)

        sheet_title = _invoice_range_sheet_title(start_date, end_date)
        ws = wb.create_sheet(title=sheet_title)
        wrote = fill_postage_invoice_worksheet(
            ws,
            conn,
            parent_number,
            start_date,
            end_date,
            discount=discount,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            remove_zeros=remove_zeros,
            hide_costs=hide_costs,
            hide_savings=hide_savings,
            set_sheet_title=sheet_title,
        )
        if not wrote:
            wb.remove(ws)
            ws = wb.create_sheet(title="No Data")
            ws["A1"] = "No postage data in range for this parent."

        out = Path(
            tempfile.mkstemp(
                suffix=f"_Postage_{parent_number}_{start_date}_{end_date}.xlsx",
                prefix="postage_",
            )[1]
        )
        wb.save(out)
        return out
    finally:
        conn.close()


def export_profit_report_xlsx(
    start_date: str,
    end_date: str,
    *,
    parent_number: int | None,
    customer_number: int | None,
    show_parents: bool,
    show_main: bool,
    flats_discount: float,
    flats_discount_efd: float = 0.23,
    efd_parcel_fee: float = 1.25,
    profit_account_ids: list[int] | None = None,
) -> Path:
    """
    Profit report workbook matching Helper Files/Profit Report Example.xlsx layout:

    Summary uses Excel formulas (B6–B10, flats table D–G, parcel mini-table, combined EFD/Lineage).
    ``efd_parcel_fee`` is Summary B10 (per-piece adder to billing on the Parcel Profit sheet column Y).
    Parcel Profit sheet is the full EFD Parcel Invoice (same as standalone EFD export).
    Flats Detail uses WS3 detail with sell-to = retail minus customer discount (preview parity).
    """
    RETAIL_RATE = 1.63
    discount_cust = max(0.0, float(flats_discount or 0.0))
    discount_efd = max(0.0, float(flats_discount_efd or 0.0))
    fee_efd = max(0.0, float(efd_parcel_fee or 0.0))
    sell_to_detail = round(float(RETAIL_RATE) - float(discount_cust), 4)

    conn = db.get_connection()
    try:
        totals = db.query_ws3_flats_profit_totals(
            conn,
            start_date,
            end_date,
            parent_number=parent_number,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            sell_to_rate=sell_to_detail,
            profit_account_ids=profit_account_ids,
        )
        rate_summary = db.query_ws3_flats_profit_rate_type_summary(
            conn,
            start_date,
            end_date,
            parent_number=parent_number,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            sell_to_rate=sell_to_detail,
            profit_account_ids=profit_account_ids,
        )
        detail = db.query_ws3_flats_profit_detail(
            conn,
            start_date,
            end_date,
            parent_number=parent_number,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            sell_to_rate=sell_to_detail,
            profit_account_ids=profit_account_ids,
        )
        parcel_raw = db.query_parcel_profit_totals(
            conn,
            start_date,
            end_date,
            parent_number=parent_number,
            customer_number=customer_number,
            show_parents=show_parents,
            show_main=show_main,
            profit_account_ids=profit_account_ids,
        )

        wb = Workbook()
        wb.remove(wb.active)

        hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        hdr_fill = PatternFill("solid", start_color="1F4E79")
        subhdr_font = Font(name="Calibri", bold=True, size=11)
        body_font = Font(name="Calibri", size=11)
        thin = Side(style="thin", color="B4B4B4")
        grid = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = "$#,##0.00"
        rate4_fmt = "0.0000"
        int_fmt = "#,##0"

        def write_table(
            tws,
            headers: list[str],
            rows: list[list[Any]],
            *,
            money_cols: set[int] = set(),
            int_cols: set[int] = set(),
            rate4_cols: set[int] = set(),
        ) -> None:
            for col, h in enumerate(headers, start=1):
                c = tws.cell(1, col, h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = grid
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, v in enumerate(row, start=1):
                    cell = tws.cell(r_idx, c_idx, v)
                    cell.font = body_font
                    cell.border = grid
                    if c_idx <= 3:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if c_idx in money_cols and v not in (None, ""):
                        cell.number_format = money_fmt
                    elif c_idx in rate4_cols and v not in (None, ""):
                        cell.number_format = rate4_fmt
                    elif c_idx in int_cols and v not in (None, ""):
                        cell.number_format = int_fmt
            tws.freeze_panes = "A2"
            tws.row_dimensions[1].height = 24

        # --- Parcel Profit first (EFD invoice layout); Summary formulas reference this sheet.
        ws_parcel = wb.create_sheet("Parcel Profit")
        fill_efd_parcel_invoice_worksheet(
            ws_parcel,
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            set_sheet_title=None,
            profit_account_ids=profit_account_ids,
            price_to_efd_adder=fee_efd,
        )

        ws = wb.create_sheet("Summary", 0)
        ws["A1"] = "Profit Report"
        ws["A1"].font = Font(name="Calibri", bold=True, size=16)
        ws["A3"] = "Start date"
        ws["B3"] = start_date
        ws["A4"] = "End date"
        ws["B4"] = end_date
        ws["D3"] = "Run days"
        ws["E3"] = totals.get("run_days", 0)
        ws["D4"] = "Total pieces"
        ws["E4"] = totals.get("total_pieces", 0)
        for r in (3, 4):
            ws[f"A{r}"].font = subhdr_font
            ws[f"B{r}"].font = body_font
        for addr in ("D3", "D4"):
            ws[addr].font = subhdr_font
        ws["E3"].number_format = int_fmt
        ws["E4"].number_format = int_fmt

        ws["A6"] = "Retail rate"
        ws["B6"] = float(RETAIL_RATE)
        ws["A7"] = "Flats discount to Customer"
        ws["B7"] = discount_cust
        ws["A8"] = "Flats discount to EFD"
        ws["B8"] = discount_efd
        ws["A9"] = "Sell-to rate"
        ws["B9"] = "=B6-B7"
        ws["A10"] = "Parcel fee to EFD ($/pc)"
        ws["B10"] = fee_efd
        for r in range(6, 11):
            ws[f"A{r}"].font = subhdr_font
            ws[f"B{r}"].font = body_font
        ws["B6"].number_format = rate4_fmt
        ws["B7"].number_format = rate4_fmt
        ws["B8"].number_format = rate4_fmt
        ws["B9"].number_format = rate4_fmt
        ws["B10"].number_format = rate4_fmt

        ws["A12"] = "Flats EFD profit by sort level (WS3 rate_type)"
        ws["A12"].font = Font(name="Calibri", bold=True, size=12)

        flats_hdr = 13
        sum_headers = [
            "Rate Type",
            "Pieces",
            "Avg USPS cost / pc",
            "Price To customer",
            "Price to EFD",
            "EFD profit",
            "Lineage Profit",
        ]
        for col, h in enumerate(sum_headers, start=1):
            c = ws.cell(flats_hdr, col, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = grid

        n_rates = len(rate_summary)
        data_start = flats_hdr + 1
        sp_type = db.WS3_FLATS_SINGLE_PIECE_RATE_TYPE

        if n_rates == 0:
            ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=7)
            c = ws.cell(
                data_start,
                1,
                "No WS3 flats profit rows found for this date range/account scope. Import the WS3 Customer Mail Detail file for these dates (Scan Now), then re-export.",
            )
            c.font = Font(name="Calibri", italic=True, size=11)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            sum_row = data_start + 1
            ws.cell(sum_row, 6, 0).number_format = money_fmt
            ws.cell(sum_row, 7, 0).number_format = money_fmt
            ws.cell(sum_row, 6).font = body_font
            ws.cell(sum_row, 7).font = body_font
        else:
            for i, row in enumerate(rate_summary):
                rr = data_start + i
                rt = str(row.get("rate_type") or "")
                ws.cell(rr, 1, rt).font = body_font
                ws.cell(rr, 1).border = grid
                ws.cell(rr, 1).alignment = Alignment(horizontal="left", vertical="center")
                pcs = int(row.get("total_pieces") or 0)
                bcell = ws.cell(rr, 2, pcs)
                bcell.font = body_font
                bcell.border = grid
                bcell.number_format = int_fmt
                bcell.alignment = Alignment(horizontal="right", vertical="center")
                avg_usps = row.get("avg_usps_cost_per_piece")
                cc = ws.cell(rr, 3, avg_usps)
                cc.font = body_font
                cc.border = grid
                cc.number_format = rate4_fmt
                cc.alignment = Alignment(horizontal="right", vertical="center")
                if rt == sp_type:
                    ws.cell(rr, 4, "=$B$6").font = body_font
                    ws.cell(rr, 5, f"=D{rr}").font = body_font
                else:
                    ws.cell(rr, 4, "=$B$6-$B$7").font = body_font
                    ws.cell(rr, 5, "=$B$6-$B$8").font = body_font
                for cidx in (4, 5):
                    ws.cell(rr, cidx).border = grid
                    ws.cell(rr, cidx).number_format = rate4_fmt
                    ws.cell(rr, cidx).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(rr, 6, f"=(D{rr}-E{rr})*B{rr}").font = body_font
                ws.cell(rr, 6).border = grid
                ws.cell(rr, 6).number_format = money_fmt
                ws.cell(rr, 7, f"=(E{rr}-C{rr})*B{rr}").font = body_font
                ws.cell(rr, 7).border = grid
                ws.cell(rr, 7).number_format = money_fmt
            sum_row = data_start + n_rates
            ws.cell(sum_row, 6, f"=SUM(F{data_start}:F{sum_row - 1})").font = body_font
            ws.cell(sum_row, 6).border = grid
            ws.cell(sum_row, 6).number_format = money_fmt
            ws.cell(sum_row, 7, f"=SUM(G{data_start}:G{sum_row - 1})").font = body_font
            ws.cell(sum_row, 7).border = grid
            ws.cell(sum_row, 7).number_format = money_fmt

        parcel_title_row = sum_row + 2
        ws.cell(parcel_title_row, 1, "Parcel Profit").font = Font(name="Calibri", bold=True, size=12)
        ph = parcel_title_row + 1
        for col, h in enumerate(["Line", "Description", "Value"], start=1):
            c = ws.cell(ph, col, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = grid

        pcount = int(parcel_raw.get("parcel_count") or 0)
        fully = float(parcel_raw.get("total_fully_paid_postage") or 0.0)
        billing = float(parcel_raw.get("total_billing_amount") or 0.0)

        r1 = ph + 1
        ws.cell(r1, 1, 1).font = body_font
        ws.cell(r1, 2, "Parcel count").font = body_font
        ws.cell(r1, 3, pcount).font = body_font
        ws.cell(r1, 3).number_format = int_fmt
        for c in (1, 2, 3):
            ws.cell(r1, c).border = grid

        r4 = ph + 2
        ws.cell(r4, 1, 4).font = body_font
        ws.cell(r4, 2, "Fully Paid Postage total").font = body_font
        ws.cell(r4, 3, round(fully, 2)).font = body_font
        ws.cell(r4, 3).number_format = money_fmt
        for c in (1, 2, 3):
            ws.cell(r4, c).border = grid

        r5 = ph + 3
        ws.cell(r5, 1, 5).font = body_font
        ws.cell(r5, 2, "Billing Amount total").font = body_font
        ws.cell(r5, 3, round(billing, 2)).font = body_font
        ws.cell(r5, 3).number_format = money_fmt
        for c in (1, 2, 3):
            ws.cell(r5, c).border = grid

        r6 = ph + 4
        ws.cell(r6, 1, 6).font = body_font
        ws.cell(r6, 2, "Parcel fee to EFD (total)").font = body_font
        ws.cell(r6, 3, f"=$B$10*C{r1}").font = body_font
        ws.cell(r6, 3).number_format = money_fmt
        for c in (1, 2, 3):
            ws.cell(r6, c).border = grid

        r8 = ph + 5
        ws.cell(r8, 1, 8).font = body_font
        ws.cell(r8, 2, "EFD Profit").font = body_font
        efd_cell = f"='Parcel Profit'!B{_EFD_PARCEL_INVOICE_SUMMARY_EFD_PROFIT_ROW}"
        ws.cell(r8, 3, efd_cell).font = body_font
        ws.cell(r8, 3).number_format = money_fmt
        for c in (1, 2, 3):
            ws.cell(r8, c).border = grid

        comb_r = r8 + 2
        ws.cell(comb_r, 1, "Combined profit summary").font = Font(name="Calibri", bold=True, size=12)
        ws.cell(comb_r + 1, 2, "EFD").font = subhdr_font
        ws.cell(comb_r + 1, 3, "Lineage").font = subhdr_font
        ws.cell(comb_r + 2, 1, "Flats total profit").font = subhdr_font
        ws.cell(comb_r + 2, 2, f"=F{sum_row}").font = body_font
        ws.cell(comb_r + 2, 2).number_format = money_fmt
        ws.cell(comb_r + 2, 3, f"=G{sum_row}").font = body_font
        ws.cell(comb_r + 2, 3).number_format = money_fmt
        ws.cell(comb_r + 3, 1, "Parcel ").font = subhdr_font
        ws.cell(comb_r + 3, 2, f"=C{r8}").font = body_font
        ws.cell(comb_r + 3, 2).number_format = money_fmt
        ws.cell(comb_r + 3, 3, f"=C{r6}").font = body_font
        ws.cell(comb_r + 3, 3).number_format = money_fmt
        ws.cell(comb_r + 4, 1, "Total").font = subhdr_font
        ws.cell(comb_r + 4, 2, f"=SUM(B{comb_r + 2}:B{comb_r + 3})").font = body_font
        ws.cell(comb_r + 4, 2).number_format = money_fmt
        ws.cell(comb_r + 4, 3, f"=SUM(C{comb_r + 2}:C{comb_r + 3})").font = body_font
        ws.cell(comb_r + 4, 3).number_format = money_fmt

        for col_letter, w in (("A", 34), ("B", 16), ("C", 14), ("D", 12), ("E", 14), ("F", 14), ("G", 16)):
            ws.column_dimensions[col_letter].width = w

        ws2 = wb.create_sheet("Flats Detail", 1)
        det_headers = [
            "Mail Date",
            "Mail ID",
            "Profile",
            "Parent Account",
            "Customer Code",
            "Customer Name",
            "Rate Type",
            "Pieces",
            "Rejected",
            "Postage Claimed",
            "USPS cost / pc",
            "Sell-to / pc",
            "Profit / pc",
            "Total profit",
        ]
        det_rows: list[list[Any]] = []
        for r in detail:
            det_rows.append(
                [
                    r.get("mail_date", ""),
                    r.get("mail_id", ""),
                    r.get("profile_name", ""),
                    f"{r.get('parent_customer_name','')} ({r.get('parent_customer_number','')})".strip(),
                    r.get("customer_code", ""),
                    r.get("customer_name", ""),
                    r.get("rate_type", ""),
                    int(r.get("num_pieces") or 0),
                    int(r.get("pcs_rejected") or 0),
                    r.get("postage_claimed"),
                    r.get("usps_cost_per_piece"),
                    r.get("sell_to_rate"),
                    r.get("profit_per_piece"),
                    r.get("total_profit"),
                ]
            )
        write_table(
            ws2,
            det_headers,
            det_rows,
            int_cols={8, 9},
            money_cols={10, 14},
            rate4_cols={11, 12, 13},
        )
        if not det_rows:
            ws2["A3"] = "No WS3 detail rows found for this date range/account scope."
            ws2["A3"].font = Font(name="Calibri", italic=True, size=11)
        widths = [12, 14, 18, 28, 14, 26, 16, 10, 10, 14, 14, 12, 12, 14]
        for i, w in enumerate(widths, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        fd, tmp = tempfile.mkstemp(
            suffix=f"_Profit_{start_date}_{end_date}.xlsx",
            prefix="profit_",
        )
        os.close(fd)
        out = Path(tmp)
        wb.save(out)
        return out
    finally:
        conn.close()


def _redact_postage_invoice_privacy(
    ws,
    *,
    hide_costs: bool,
    hide_savings: bool,
    last_data_row: int,
    totals_row: int | None,
) -> None:
    """Remove currency from Customer/Retail/Savings columns per dashboard hide flags."""
    # Weight grid L–N: rows 14–31; footer totals row 32; piece-summary row 33 (no money on 33).
    if hide_costs:
        for r in range(14, 32):
            ws.cell(r, 12, None)
            ws.cell(r, 13, None)
            ws.cell(r, 14, None)
        for c in (12, 13, 14):
            ws.cell(32, c, None)
        if last_data_row >= 35:
            for r in range(35, last_data_row + 1):
                ws.cell(r, 5, None)
                ws.cell(r, 6, None)
        if totals_row is not None:
            ws.cell(totals_row, 5, None)
            ws.cell(totals_row, 6, None)
        return

    if hide_savings:
        for r in range(14, 32):
            ws.cell(r, 14, None)
        ws.cell(32, 14, None)
        if last_data_row >= 35:
            for r in range(35, last_data_row + 1):
                ws.cell(r, 6, None)
        if totals_row is not None:
            ws.cell(totals_row, 6, None)


def _write_invoice_sheet(
    ws,
    sheet_idx: int,
    file_date: datetime,
    start_date: str,
    end_date: str,
    parent_number: int,
    parent_name: str,
    contact: dict[str, str],
    rates: dict[int, float],
    weight_data: dict[int, dict[str, Any]],
    children: list,
    discount: float,
    other_pieces: int = 0,
    other_cost: float = 0.0,
    range_total_pieces: int = 0,
    range_total_retail: float = 0.0,
    hide_costs: bool = False,
    hide_savings: bool = False,
) -> None:
    """
    Layout: rows 12–13 (old balance block) removed — weight table starts at row 13 header,
    1–13 oz at 14–26, Letter 27, Foreign 30, totals row 31, piece summary 32, cost centers from 33.
    Weight grid column J references G (retail). Column N savings = I*(G−H) per row; totals-row N = SUM(N14:N30).
    Cost centers: C pieces, D rejects, E charges; F savings = (C−D)*($G$14−$H$14) (Letter row F = 0).
    """
    BOLD = Font(bold=True)
    CURR = "$#,##0.00"
    INT_FMT = "#,##0"

    R_TBL_HDR = 13
    R_OZ_FIRST = 14
    R_LETTER = 27
    R_FOREIGN = 30
    R_TOTALS = 31
    R_PIECE_SUM = 32
    R_CC_HDR = 33
    R_CC_FIRST = 34

    ws["A1"] = "INVOICE # "
    ws["A1"].font = BOLD
    ws["C1"] = sheet_idx
    ws["C1"].font = BOLD
    ws["M2"] = "Week ending"
    ws["M2"].font = BOLD

    ws["A3"] = "Bill to: "
    ws["A3"].font = BOLD
    ws["C3"] = parent_name
    ws["C3"].font = BOLD
    ws["L3"] = "Project Date:"
    ws["L3"].font = BOLD
    ws["M3"] = "=F12"
    ws["M3"].font = BOLD

    ws["A4"] = "Attn:"
    ws["A4"].font = BOLD
    ws["C4"] = contact.get("contact_name", "")
    ws["C4"].font = BOLD
    ws["L4"] = "Customer ID#"
    ws["L4"].font = BOLD
    ws["M4"] = contact.get("customer_id", f"            {parent_number}")
    ws["M4"].font = BOLD

    ws["C5"] = contact.get("address1", "")
    ws["C6"] = contact.get("city_state_zip", "")
    ws["A8"] = None
    ws["C8"] = None
    ws["A9"] = None
    ws["C9"] = None
    ws["A10"] = None
    ws["C10"] = None

    if start_date == end_date:
        period_note = start_date
    else:
        period_note = f"{start_date} to {end_date}"
    ws.cell(11, 9, f"Account Summary for {parent_name} ({period_note})").font = BOLD

    ws["A12"] = None
    c_f12 = ws.cell(12, 6, file_date)
    c_f12.number_format = "MM/DD/YYYY"

    ws["A13"] = None
    ws["A15"] = None
    ws["B15"] = None
    ws["A16"] = None
    ws["B16"] = None
    ws["A17"] = None
    ws["B17"] = None

    for col, val in [
        (6, "Weight"),
        (7, "1st Class"),
        (8, "FC discount"),
        (9, "Total #'s"),
        (10, "IMB rejects"),
        (11, "Total #'s"),
        (12, "Customer Cost"),
        (13, "Retail Cost"),
        (14, "Savings"),
    ]:
        c = ws.cell(R_TBL_HDR, col, val)
        c.font = BOLD

    for oz in range(1, 14):
        r = R_TBL_HDR + oz
        retail = _retail_rate(rates, oz)
        efd = round(max(0.0, retail - discount), 4)
        wd = weight_data.get(oz, {})
        efd_pieces = int(wd.get("efd_pieces", wd.get("pieces", 0)) or 0)
        imb_pieces = int(wd.get("reject_pieces", 0) or 0)
        ws3_pieces = int(wd.get("ws3_pieces", 0) or 0)
        i_discount = max(0, efd_pieces - ws3_pieces)
        reject_pieces = imb_pieces + ws3_pieces

        ws.cell(r, 6, f"{oz} oz").font = BOLD
        ws.cell(r, 7, retail).number_format = CURR
        ws.cell(r, 7).font = BOLD
        ws.cell(r, 8, efd).number_format = CURR
        ws.cell(r, 8).font = BOLD
        ws.cell(r, 9, i_discount).number_format = INT_FMT
        ws.cell(r, 9).font = BOLD
        cj = ws.cell(r, 10, f"=G{r}")
        cj.number_format = CURR
        cj.font = BOLD
        ws.cell(r, 11, reject_pieces).number_format = INT_FMT
        ws.cell(r, 11).font = BOLD
        ws.cell(r, 12, f"=H{r}*I{r}+J{r}*K{r}").number_format = CURR
        ws.cell(r, 12).font = BOLD
        ws.cell(r, 13, f"=G{r}*(I{r}+K{r})").number_format = CURR
        ws.cell(r, 13).font = BOLD
        ws.cell(r, 14, f"=I{r}*(G{r}-H{r})").number_format = CURR
        ws.cell(r, 14).font = BOLD

    # Mail not in the 1–13 oz flat grid: non-flat classes and flat mail at 0 oz / 13+ oz (postage only; no discount savings).
    ws.cell(R_LETTER, 6, "Letter").font = BOLD
    ws.cell(R_LETTER, 7, 0).number_format = CURR
    ws.cell(R_LETTER, 7).font = BOLD
    ws.cell(R_LETTER, 8, 0).number_format = CURR
    ws.cell(R_LETTER, 8).font = BOLD
    ws.cell(R_LETTER, 9, int(other_pieces)).number_format = INT_FMT
    ws.cell(R_LETTER, 9).font = BOLD
    clj = ws.cell(R_LETTER, 10, f"=G{R_LETTER}")
    clj.number_format = CURR
    clj.font = BOLD
    ws.cell(R_LETTER, 11, 0).number_format = INT_FMT
    ws.cell(R_LETTER, 11).font = BOLD
    oc = round(float(other_cost), 2)
    ws.cell(R_LETTER, 12, oc).number_format = CURR
    ws.cell(R_LETTER, 12).font = BOLD
    ws.cell(R_LETTER, 13, 0).number_format = CURR
    ws.cell(R_LETTER, 13).font = BOLD
    ws.cell(R_LETTER, 14, 0).number_format = CURR
    ws.cell(R_LETTER, 14).font = BOLD

    for r in (28, 29):
        ws.cell(r, 6, "").font = BOLD
        for col in (7, 8, 12, 13, 14):
            ws.cell(r, col, 0).number_format = CURR
            ws.cell(r, col).font = BOLD
        cfj = ws.cell(r, 10, f"=G{r}")
        cfj.number_format = CURR
        cfj.font = BOLD
        ws.cell(r, 9, 0).number_format = INT_FMT
        ws.cell(r, 9).font = BOLD
        ws.cell(r, 11, 0).number_format = INT_FMT
        ws.cell(r, 11).font = BOLD

    ws.cell(R_FOREIGN, 6, "Foreign").font = BOLD
    for col in (7, 8):
        ws.cell(R_FOREIGN, col, 0).number_format = CURR
        ws.cell(R_FOREIGN, col).font = BOLD
    cf = ws.cell(R_FOREIGN, 10, f"=G{R_FOREIGN}")
    cf.number_format = CURR
    cf.font = BOLD
    ws.cell(R_FOREIGN, 9, 0).font = BOLD
    ws.cell(R_FOREIGN, 9).number_format = INT_FMT
    ws.cell(R_FOREIGN, 11, 0).number_format = INT_FMT
    ws.cell(R_FOREIGN, 11).font = BOLD
    ws.cell(R_FOREIGN, 12, 0).number_format = CURR
    ws.cell(R_FOREIGN, 12).font = BOLD
    ws.cell(R_FOREIGN, 13, 0).number_format = CURR
    ws.cell(R_FOREIGN, 13).font = BOLD
    ws.cell(R_FOREIGN, 14, 0).number_format = CURR
    ws.cell(R_FOREIGN, 14).font = BOLD

    lo, hi = R_OZ_FIRST, R_FOREIGN
    ws.cell(R_TOTALS, 9, f"=SUM(I{lo}:I{hi})").font = BOLD
    ws.cell(R_TOTALS, 11, f"=SUM(K{lo}:K{hi})").font = BOLD
    ws.cell(R_TOTALS, 12, f"=SUM(L{lo}:L{hi})").number_format = CURR
    ws.cell(R_TOTALS, 12).font = BOLD
    tr = round(float(range_total_retail or 0.0), 2)
    ws.cell(R_TOTALS, 13, tr).number_format = CURR
    ws.cell(R_TOTALS, 13).font = BOLD
    ws.cell(R_TOTALS, 14, f"=SUM(N{lo}:N{hi})").number_format = CURR
    ws.cell(R_TOTALS, 14).font = BOLD

    ws.cell(R_PIECE_SUM, 8, "Total Pieces").font = BOLD
    ws.cell(R_PIECE_SUM, 9, range_total_pieces).number_format = INT_FMT
    ws.cell(R_PIECE_SUM, 9).font = BOLD

    for col, val in [
        (1, "Cost Centers "),
        (2, "CUSTOMER NAME"),
        (3, "# Pieces "),
        (4, "Rejects"),
        (5, "Charges "),
        (6, "Savings "),
    ]:
        ws.cell(R_CC_HDR, col, val).font = BOLD

    child_list = list(children)
    if int(other_pieces or 0) > 0 or abs(float(other_cost or 0.0)) > 1e-9:
        child_list.append(
            {
                "customer_number": "",
                "customer_name": "Letter",
                "pieces": int(other_pieces or 0),
                "rejects": 0,
                "cost": round(float(other_cost or 0.0), 2),
                "savings": 0.0,
            }
        )
    for i, child in enumerate(child_list):
        row = R_CC_FIRST + i
        pieces = int(child["pieces"] or 0)
        rejects = int(child.get("rejects", 0) or 0)
        cost = round(float(child["cost"] or 0.0), 2)
        cn = child.get("customer_number")
        is_letter_cc = (
            _row_value(child, "customer_name") == "Letter"
            and cn in (None, "", "—")
        )
        ws.cell(row, 1, cn if cn not in (None, "") else "—").font = BOLD
        ws.cell(row, 2, _row_value(child, "customer_name")).font = BOLD
        ws.cell(row, 3, pieces).number_format = INT_FMT
        ws.cell(row, 3).font = BOLD
        ws.cell(row, 4, rejects).number_format = INT_FMT
        ws.cell(row, 4).font = BOLD
        ws.cell(row, 5, cost).number_format = CURR
        ws.cell(row, 5).font = BOLD
        if is_letter_cc:
            ws.cell(row, 6, 0).number_format = CURR
        else:
            ws.cell(
                row,
                6,
                f"=(C{row}-D{row})*($G${R_OZ_FIRST}-$H${R_OZ_FIRST})",
            ).number_format = CURR
        ws.cell(row, 6).font = BOLD

    last_data_row = (
        R_CC_FIRST + len(child_list) - 1 if child_list else R_CC_HDR - 1
    )

    for r in range(R_CC_FIRST, last_data_row + 1):
        for cell in ws[r]:
            if cell.value is not None:
                _apply_bold_preserve_font(cell)

    t = last_data_row + 1
    totals_row: int | None = None
    if last_data_row >= R_CC_FIRST:
        totals_row = t
        ws.cell(t, 1, "Totals").font = BOLD
        c = ws.cell(t, 3, f"=SUM(C{R_CC_FIRST}:C{last_data_row})")
        c.number_format = INT_FMT
        c.font = BOLD
        c = ws.cell(t, 4, f"=SUM(D{R_CC_FIRST}:D{last_data_row})")
        c.number_format = INT_FMT
        c.font = BOLD
        c = ws.cell(t, 5, f"=SUM(E{R_CC_FIRST}:E{last_data_row})")
        c.number_format = CURR
        c.font = BOLD
        c = ws.cell(t, 6, f"=SUM(F{R_CC_FIRST}:F{last_data_row})")
        c.number_format = CURR
        c.font = BOLD

    _redact_postage_invoice_privacy(
        ws,
        hide_costs=hide_costs,
        hide_savings=hide_savings,
        last_data_row=last_data_row,
        totals_row=totals_row,
    )

    for r in range(1, 14):
        for cell in ws[r]:
            if cell.value is not None:
                _apply_bold_preserve_font(cell)
    for r in range(14, R_CC_HDR + 1):
        for cell in ws[r]:
            if cell.value is not None:
                _apply_bold_preserve_font(cell)

    ws.freeze_panes = "G2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 12


def parcel_report_scope_label(parent_number: int | None, customer_number: int | None) -> str:
    """Filename segment for parcel export: parent, optional customer, or ALL."""
    parts: list[str] = []
    if parent_number is not None:
        parts.append(str(parent_number))
    if customer_number is not None:
        parts.append(f"c{customer_number}")
    return "_".join(parts) if parts else "ALL"


def parcel_report_download_name(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
) -> str:
    scope = parcel_report_scope_label(parent_number, customer_number)
    return f"Parcel_Report_{scope}_{start_date}_{end_date}.xlsx"


def parcel_counts_download_name(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
) -> str:
    scope = parcel_report_scope_label(parent_number, customer_number)
    return f"Parcel_Counts_{scope}_{start_date}_{end_date}.xlsx"


_FILENAME_BAD_CHARS_RE = re.compile(r'[\\/:*?"<>|]+')


def _safe_filename_piece(s: str) -> str:
    # Keep spaces/parentheses/underscores for readability, but remove characters that break downloads.
    out = _FILENAME_BAD_CHARS_RE.sub("-", str(s))
    out = re.sub(r"\s+", " ", out).strip()
    return out or "Report"


def _short_mdy(date_str: str) -> str:
    """M-D-YYYY with no leading zeros; falls back to raw if parsing fails."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return f"{d.month}-{d.day}-{d.year}"
    except ValueError:
        return str(date_str)


def efd_report_date_range_label(start_date: str, end_date: str) -> str:
    """Date range suffix for EFD bundle downloads (e.g. 6-8 to 6-12 or 6-8-2026)."""
    try:
        s = datetime.strptime(start_date, "%Y-%m-%d")
        e = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        return f"{_short_mdy(start_date)} to {_short_mdy(end_date)}"
    if start_date == end_date:
        return f"{s.month}-{s.day}-{s.year}"
    if s.year == e.year:
        return f"{s.month}-{s.day} to {e.month}-{e.day}"
    return f"{_short_mdy(start_date)} to {_short_mdy(end_date)}"


def efd_account_report_download_name(
    *,
    title_name: str | None,
    parent_number: int,
    report_label: str,
    start_date: str,
    end_date: str,
    ext: str,
) -> str:
    """
    EFD weekly bundle per-account download name.

    Example: "KC Presort LLC -EFD (3906) Parcel invoice 6-8 to 6-12.xlsx"
    """
    raw_title = (title_name or report_label).strip()
    display_title = raw_title if "-EFD" in raw_title else f"{raw_title} -EFD"
    base = _safe_filename_piece(display_title)
    date_range = _safe_filename_piece(efd_report_date_range_label(start_date, end_date))
    ext_clean = str(ext).lstrip(".")
    return f"{base} ({int(parent_number)}) {report_label} {date_range}.{ext_clean}"


def postage_invoice_download_name(
    *,
    title_name: str | None,
    parent_number: int,
    end_date: str,
) -> str:
    """
    Friendly Postage Invoice download name (standalone Export Postage Invoice).

    Example: "Blue Cross Blue Shield (3901) Postage invoice 6-12-2026.xlsx"
    """
    cust_name = _safe_filename_piece((title_name or f"Account {parent_number}").strip())
    try:
        dt = datetime.strptime(end_date, "%Y-%m-%d")
        end_label = _safe_filename_piece(f"{dt.month}-{dt.day}-{dt.year}")
    except ValueError:
        end_label = _safe_filename_piece(_short_mdy(end_date))
    return f"{cust_name} ({int(parent_number)}) Postage invoice {end_label}.xlsx"


def parcel_invoice_download_name(*, title_name: str | None, parent_number: int | None, end_date: str) -> str:
    """
    Friendly Parcel Invoice download name.

    Example: "Security Benefit_Zinnia -EFD (3900) Parcel invoice 4-10-2026.xlsx"
    """
    raw_title = (title_name or "Parcel Invoice").strip()
    display_title = raw_title if raw_title.rstrip().endswith("-EFD") else f"{raw_title} -EFD"
    base = _safe_filename_piece(display_title)
    end_short = _safe_filename_piece(_short_mdy(end_date))
    if parent_number is None:
        return f"{base} Parcel invoice {end_short}.xlsx"
    return f"{base} ({int(parent_number)}) Parcel invoice {end_short}.xlsx"


def aggregate_parcel_count_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Roll up parcel grid rows to one line per date × parent × child (same as dashboard CSV)."""
    m: dict[tuple[Any, ...], dict[str, Any]] = {}
    for r in rows:
        key = (
            r.get("date"),
            r.get("parent_name"),
            r.get("parent_number"),
            r.get("child_name"),
            r.get("child_number"),
        )
        if key not in m:
            m[key] = {
                "date": key[0],
                "parent_name": key[1],
                "parent_number": key[2],
                "child_name": key[3],
                "child_number": key[4],
                **{f"lb_{i}": 0 for i in range(1, 11)},
                "lb_10plus": 0,
                "total_qty": 0,
                "total_billed": 0.0,
                "total_retail": 0.0,
            }
        a = m[key]
        for i in range(1, 11):
            a[f"lb_{i}"] += int(r.get(f"lb_{i}") or 0)
        a["lb_10plus"] += int(r.get("lb_10plus") or 0)
        a["total_qty"] += int(r.get("total_qty") or 0)
        a["total_billed"] += float(r.get("total_billed") or 0)
        a["total_retail"] += float(r.get("total_retail") or 0)

    def sort_key(x: dict[str, Any]) -> tuple[Any, ...]:
        def _n(v: Any) -> float:
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        return (
            str(x.get("date") or ""),
            str(x.get("parent_name") or "").lower(),
            _n(x.get("parent_number")),
            str(x.get("child_name") or "").lower(),
            _n(x.get("child_number")),
        )

    return sorted(m.values(), key=sort_key)


def export_parcel_counts_report_xlsx(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
    show_parents: bool,
    show_main: bool,
    consolidate: bool,
    remove_zeros: bool,
    hide_costs: bool,
    hide_customer_numbers: bool = True,
) -> Path:
    """Formatted workbook: PARCELS (COUNTS) with weight buckets; last column = Retail Cost (when costs on).

    Uses ``.xlsx`` (Office Open XML) for currency formatting; Excel opens as native format.
    """
    conn = db.get_connection()
    try:
        data = db.query_parcels(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            consolidate,
            remove_zeros,
            hide_costs,
        )
    finally:
        conn.close()

    agg = aggregate_parcel_count_rows(data.get("rows") or [])

    wb = Workbook()
    ws = wb.active
    ws.title = "PARCELS (COUNTS)"[:31]

    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    body_font = Font(name="Calibri", size=11)
    thin = Side(style="thin", color="B4B4B4")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = "$#,##0.00"
    int_fmt = "#,##0"

    headers_text = ["Date", "Parent Name", "Child Name"]
    if not hide_customer_numbers:
        headers_text.append("Child Number")
    headers_text.extend([*[f"{i} lb" for i in range(1, 11)], "10+ lb", "Total Qty"])
    if not hide_costs:
        headers_text.append("Retail Cost")

    name_col_count = 3 + (0 if hide_customer_numbers else 1)
    first_lb_col = name_col_count + 1

    for col, h in enumerate(headers_text, start=1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = grid

    for ri, row in enumerate(agg, start=2):
        values: list[Any] = [
            row.get("date"),
            row.get("parent_name"),
            row.get("child_name"),
        ]
        if not hide_customer_numbers:
            values.append(row.get("child_number"))
        values.extend(
            [
                *[int(row.get(f"lb_{i}") or 0) for i in range(1, 11)],
                int(row.get("lb_10plus") or 0),
                int(row.get("total_qty") or 0),
            ]
        )
        if not hide_costs:
            # Retail-only costing: write one retail-cost column.
            values.append(round(float(row.get("total_retail") or 0), 2))
        ncols = len(values)
        for col, v in enumerate(values, start=1):
            cell = ws.cell(ri, col, v)
            cell.font = body_font
            cell.border = grid
            if col <= name_col_count:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if not hide_costs and col == ncols:
                cell.number_format = money_fmt
            elif col >= first_lb_col:
                cell.number_format = int_fmt

    ws.freeze_panes = f"{get_column_letter(first_lb_col)}2"
    ws.row_dimensions[1].height = 28

    col_widths = [12.0, 24.0, 30.0]
    if not hide_customer_numbers:
        col_widths.append(10.0)
    col_widths.extend([9.0] * 10 + [9.0, 11.0])
    if not hide_costs:
        col_widths.append(14.0)
    for i, w in enumerate(col_widths, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, w)

    fd, tmp = tempfile.mkstemp(suffix="_parcel_counts.xlsx", prefix="parcel_counts_")
    os.close(fd)
    out = Path(tmp)
    wb.save(out)
    return out


# Min column widths for stacked over-10lb (A–G) + parcel invoice (A–E) blocks; merged with zone-summary widths.
_AF_HM_MIN_WIDTHS: dict[int, float] = {
    1: 30.0,
    2: 30.0,
    3: 12.0,
    4: 14.0,
    5: 14.0,
    6: 14.0,
    7: 14.0,
    8: 14.0,
}


def write_parcel_af_hm_sections(ws, section_start: int, sections: dict[str, Any]) -> None:
    """Over-10lb block (customer name in column A), optional totals row, then per-customer invoice + total."""
    sec_hdr = Font(name="Arial", bold=True, size=10)
    sec_fill = PatternFill("solid", start_color="E7E6E6")
    tot_font = Font(name="Arial", bold=True, size=10)
    cur_fmt = "$#,##0.00"
    int_fmt = "0"
    data_font = Font(name="Arial", size=10)

    hrows = sections.get("heavy_rows") or []
    crows = sections.get("customers") or []

    hdr_af = [
        "Customer Name",
        "Count",
        "lbs.",
        "Zone",
        "Retail",
        "Discount",
        "Savings",
    ]
    hdr_hm = [
        "Customer #",
        "Customer Name",
        "Items (customer)",
        "Retail",
        "Discount",
        "Savings",
    ]

    cur_row = section_start
    if hrows:
        for col, h in enumerate(hdr_af, 1):
            c = ws.cell(cur_row, col, h)
            c.font = sec_hdr
            c.fill = sec_fill
            c.alignment = Alignment(horizontal="center")
        for i, hr in enumerate(hrows):
            r = cur_row + 1 + i
            ws.cell(r, 1, hr.get("customer_name") or "(no name)")
            ws.cell(r, 2, hr["count"])
            ws.cell(r, 3, hr["lbs"])
            ws.cell(r, 4, hr["zone"])
            ws.cell(r, 5, hr["base"]).number_format = cur_fmt
            ws.cell(r, 6, hr["efd"]).number_format = cur_fmt
            ws.cell(r, 7, hr["savings"]).number_format = cur_fmt
            for col in range(1, 8):
                cell = ws.cell(r, col)
                cell.font = data_font
                if col == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            for col in (2, 3, 4):
                ws.cell(r, col).number_format = int_fmt
        first_h = cur_row + 1
        last_h = cur_row + len(hrows)
        tot_h = cur_row + 1 + len(hrows)
        ht = ws.cell(tot_h, 1, "Total")
        ht.font = tot_font
        ht.fill = sec_fill
        ht.alignment = Alignment(horizontal="left", vertical="center")
        hb = ws.cell(tot_h, 2, f"=SUM(B{first_h}:B{last_h})")
        hb.font = tot_font
        hb.fill = sec_fill
        hb.number_format = int_fmt
        hb.alignment = Alignment(horizontal="right", vertical="center")
        for col in (3, 4, 5, 6, 7):
            ws.cell(tot_h, col).fill = sec_fill
        # Discount total (col F) = SUMPRODUCT(count, discount_per_piece).
        hf = ws.cell(
            tot_h,
            6,
            f"=SUMPRODUCT(B{first_h}:B{last_h},F{first_h}:F{last_h})",
        )
        hf.font = tot_font
        hf.fill = sec_fill
        hf.number_format = cur_fmt
        hf.alignment = Alignment(horizontal="right", vertical="center")
        # Savings total (col G) = SUM of row savings.
        hg = ws.cell(tot_h, 7, f"=SUM(G{first_h}:G{last_h})")
        hg.font = tot_font
        hg.fill = sec_fill
        hg.number_format = cur_fmt
        hg.alignment = Alignment(horizontal="right", vertical="center")
        cur_row = tot_h + 2

    for col, h in enumerate(hdr_hm, 1):
        c = ws.cell(cur_row, col, h)
        c.font = sec_hdr
        c.fill = sec_fill
        c.alignment = Alignment(horizontal="center")

    for i, cr in enumerate(crows):
        r = cur_row + 1 + i
        ws.cell(r, 1, cr["customer_number"])
        ws.cell(r, 2, cr["name"])
        ws.cell(r, 3, cr["qty"])
        ws.cell(r, 4, cr["cost"]).number_format = cur_fmt
        ws.cell(r, 5, cr.get("discount") or 0.0).number_format = cur_fmt
        ws.cell(r, 6, cr["savings"]).number_format = cur_fmt
        for col in range(1, 7):
            cell = ws.cell(r, col)
            cell.font = data_font
            if col in (1, 2):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(r, 3).number_format = int_fmt

    tot_r = cur_row + 1 + len(crows)
    t = ws.cell(tot_r, 1, "Total")
    t.font = tot_font
    t.fill = sec_fill
    if crows:
        d1 = cur_row + 1
        d2 = cur_row + len(crows)
        c3 = ws.cell(tot_r, 3, f"=SUM(C{d1}:C{d2})")
        c3.font = tot_font
        c3.number_format = int_fmt
        c3.fill = sec_fill
        c4 = ws.cell(tot_r, 4, f"=SUM(D{d1}:D{d2})")
        c4.font = tot_font
        c4.number_format = cur_fmt
        c4.fill = sec_fill
        c5 = ws.cell(tot_r, 5, f"=SUM(E{d1}:E{d2})")
        c5.font = tot_font
        c5.number_format = cur_fmt
        c5.fill = sec_fill
        c6 = ws.cell(tot_r, 6, f"=SUM(F{d1}:F{d2})")
        c6.font = tot_font
        c6.number_format = cur_fmt
        c6.fill = sec_fill
    else:
        c3z = ws.cell(tot_r, 3, 0)
        c3z.font = tot_font
        c3z.number_format = int_fmt
        c3z.fill = sec_fill
        z4 = ws.cell(tot_r, 4, 0.0)
        z4.font = tot_font
        z4.number_format = cur_fmt
        z4.fill = sec_fill
        z5 = ws.cell(tot_r, 5, 0.0)
        z5.font = tot_font
        z5.number_format = cur_fmt
        z5.fill = sec_fill
        z6 = ws.cell(tot_r, 6, 0.0)
        z6.font = tot_font
        z6.number_format = cur_fmt
        z6.fill = sec_fill


def fill_parcel_report_worksheet(
    ws,
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
) -> None:
    """Write parcel piece line-item table and a count total row only (no over-10lb / invoice blocks)."""
    conn = db.get_connection()
    try:
        filtered = db.query_parcel_report_rows(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
        )
    finally:
        conn.close()

    # Last column is IMPB (billing CSV column CE / index 82); postage currency cols removed per spec.
    headers = [
        "Customer #",
        "Customer Name",
        "Parent Name",
        "Piece ID",
        "Time Stamp",
        "Mail Class",
        "Zone",
        "Weight (oz)",
        "Weight (lbs)",
        "Count",
        "Department",
        "Handling Type",
        "IMPB",
    ]
    header_fill = PatternFill("solid", start_color="BDD7EE")
    header_font = Font(name="Arial", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    data_font = Font(name="Arial", size=10)
    for i, row in enumerate(filtered, 2):
        ws.cell(i, 1, row["custom_account_code"])
        ws.cell(i, 2, row["account_name"])
        ws.cell(i, 3, row["parent_name"])
        ws.cell(i, 4, row["piece_id"])
        ws.cell(i, 5, row["time_stamp"])
        ws.cell(i, 6, row["usps_mail_class"])
        ws.cell(i, 7, row["zone"])
        ws.cell(i, 8, row["weight_oz"])
        ws.cell(i, 9, f"=H{i}/16")
        ws.cell(i, 10, 1)
        ws.cell(i, 11, row["department_name"])
        ws.cell(i, 12, row["handling_type"])
        ws.cell(i, 13, row["impb"] if row["impb"] is not None else "")
        ws.cell(i, 9).number_format = "0.000"
        for col in range(1, 14):
            ws.cell(i, col).font = data_font

    n = max(len(filtered) + 1, 1)
    tot = n + 1
    bold = Font(name="Arial", bold=True, size=11)
    top_border = Border(top=Side(style="thin"))
    ws.cell(tot, 1, "TOTALS").font = bold
    if len(filtered) == 0:
        c = ws.cell(tot, 10, 0)
        c.font = bold
        c.border = top_border
    else:
        c = ws.cell(tot, 10, f"=SUM(J2:J{n})")
        c.font = bold
        c.border = top_border

    widths = [12, 30, 30, 28, 18, 25, 6, 12, 12, 8, 25, 20, 42]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"


def export_parcel_report(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{start_date} to {end_date}"[:31]
    fill_parcel_report_worksheet(
        ws,
        start_date,
        end_date,
        parent_number,
        customer_number,
        show_parents,
        show_main,
    )
    scope = parcel_report_scope_label(parent_number, customer_number)
    out = Path(
        tempfile.mkstemp(
            suffix=f"_Parcel_{scope}_{start_date}_{end_date}.xlsx",
            prefix="parcel_",
        )[1]
    )
    wb.save(out)
    return out


def export_parcel_billing_csv(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
    show_parents: bool,
    show_main: bool,
) -> Path:
    """
    Consolidated parcel export: full raw billing_records payload as CSV.

    Output columns:
      - parent_name
      - parent_number
      - all billing_records columns (in table order)
    """
    conn = db.get_connection()
    try:
        # Keep table order stable without hard-coding ~95 headers.
        pragma = conn.execute("PRAGMA table_info(billing_records)").fetchall()
        billing_cols: list[str] = []
        for r in pragma:
            # Works for sqlite3.Row and tuple rows.
            try:
                name = r["name"]
            except (TypeError, KeyError, IndexError):
                name = r[1] if isinstance(r, (list, tuple)) and len(r) > 1 else None
            if name:
                billing_cols.append(str(name))
        if not billing_cols:
            raise ValueError("No billing_records columns found (table missing?)")

        # GA rates once; Priority Mail is chosen per billing row date (latest tariff on or before that day).
        ga_rates = db.get_ground_advantage_retail_rates(conn)
        pm_by_day: dict[str, tuple] = {}

        def _pm_for_billing_row(ts_raw: Any) -> tuple:
            od = db.iso_date_from_billing_timestamp(ts_raw)
            key = od if od is not None else end_date
            if key not in pm_by_day:
                pm_by_day[key] = db.get_priority_mail_retail_rates(conn, as_of_date=key)
            return pm_by_day[key]

        rows = db.query_parcel_billing_rows_full(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
        )
        # `compute_retail_cost_for_piece` takes a conn, but with cached rate tables
        # it will not query per-row.
        conn_for_cost = conn
    except Exception:
        conn.close()
        raise

    fd, tmp = tempfile.mkstemp(
        suffix=f"_Parcel_Billing_{parcel_report_scope_label(parent_number, customer_number)}_{start_date}_{end_date}.csv",
        prefix="parcel_billing_",
    )
    os.close(fd)
    out = Path(tmp)

    retail_col = "retail_cost"
    if retail_col in billing_cols:
        # Extremely unlikely, but don't duplicate a column name.
        retail_col = "retail_cost_lookup"

    # Insert retail column immediately after Billing Amount (source col X).
    cols_with_retail = list(billing_cols)
    try:
        i_bill = cols_with_retail.index("billing_amount")
        cols_with_retail.insert(i_bill + 1, retail_col)
    except ValueError:
        # Fallback: append at end if schema differs.
        cols_with_retail.append(retail_col)

    headers = ["parent_name", "parent_number", *cols_with_retail]
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            parent_name = r["parent_name"] if "parent_name" in r.keys() else None
            parent_number = r["parent_number"] if "parent_number" in r.keys() else None

            weight_oz = r["weight_oz"] if "weight_oz" in r.keys() else None
            zone_raw = r["zone"] if "zone" in r.keys() else None
            fallback_retail = (
                r["fully_paid_postage"] if "fully_paid_postage" in r.keys() else None
            )
            ts_raw = r["time_stamp"] if "time_stamp" in r.keys() else None
            priced = db.compute_retail_cost_for_piece(
                conn_for_cost,
                weight_oz=weight_oz,
                zone_raw=zone_raw,
                fallback_retail=fallback_retail,
                ga_rates=ga_rates,
                pm_rates=_pm_for_billing_row(ts_raw),
            )
            retail_val = priced.get("retail")
            if retail_val is not None:
                try:
                    retail_val = round(float(retail_val), 2)
                except (TypeError, ValueError):
                    retail_val = None

            row_by_col: dict[str, Any] = {c: (r[c] if c in r.keys() else None) for c in billing_cols}
            row_by_col[retail_col] = retail_val
            w.writerow(
                [
                    parent_name,
                    parent_number,
                    *[row_by_col.get(c) for c in cols_with_retail],
                ]
            )

    conn.close()
    return out


# EFD Parcel Invoice XLSX: fixed column subset so billing_amount = Excel column X (24),
# Price to EFD = Y (25), EFD Revenue = Z (26) = retail_cost (AA) − Price to EFD (Y), retail_cost = AA (27).
_EFD_PARCEL_INVOICE_PREFIX_KEYS: tuple[str, ...] = (
    "parent_name",
    "parent_number",
    "id",
    "billing_import_id",
    "custom_account_code",
    "account_name",
    "piece_id",
    "machine_serial",
    "time_stamp",
    "weight_oz",
    "handling_type",
    "usps_mail_class",
    "usps_mail_prep_type",
    "routing_category",
    "routing_string",
    "bundle_qualification",
    "bundle_zip",
    "account_id",
    "customer_barcode_symbology",
    "customer_barcode",
    "department_id",
    "department_name",
    "manifest_id",
)
_EFD_PARCEL_INVOICE_TAIL_KEYS: tuple[str, ...] = (
    "imb_tracking_code",
    "impb",
    "efn",
    "surcharge_postage",
    "fss",
    "tub_number",
    "postal_discounts",
    "hr_address",
    "hr_city",
    "hr_state",
    "hr_zip",
)
# Row 6 = total parcel quantities; rows 7–8 blank; header row 9; data row 10+.
_EFD_PARCEL_INVOICE_HEADER_ROW = 9
_EFD_PARCEL_INVOICE_DATA_START = 10
_EFD_PARCEL_INVOICE_MONEY_FMT = "$#,##0.00"
_EFD_PARCEL_INVOICE_QTY_FMT = "#,##0"
# Summary block on EFD Parcel Invoice sheet: EFD Profit total is always B5.
_EFD_PARCEL_INVOICE_SUMMARY_EFD_PROFIT_ROW = 5
_EFD_PARCEL_INVOICE_SUMMARY_EFD_PROFIT_COL = 2


def _efd_price_to_efd_adder_literal(parcel_reseller: float) -> str:
    """Non-negative adder formatted for Excel formulas (billing + adder → Price to EFD)."""
    x = max(0.0, float(parcel_reseller or 0.0))
    x = round(x, 6)
    if abs(x - round(x)) < 1e-12:
        return str(int(round(x)))
    return f"{x:.6f}".rstrip("0").rstrip(".")


def fill_efd_parcel_invoice_worksheet(
    ws,
    conn: sqlite3.Connection,
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
    show_parents: bool,
    show_main: bool,
    *,
    set_sheet_title: str | None = "EFD Parcel Invoice",
    profit_account_ids: list[int] | None = None,
    price_to_efd_adder: float = 1.25,
) -> tuple[str, int, int, int]:
    """
    Write EFD Parcel Invoice layout to an existing worksheet (summary rows 1–6, header row 9, data from row 10).

    Returns ``(title, header_row, data_start_row, last_data_row)`` for cross-sheet formulas.
    Does not ``save`` or ``close`` the connection.
    ``price_to_efd_adder`` is the per-piece dollar adder in column Y (Price to EFD = billing + adder).
    """
    adder_lit = _efd_price_to_efd_adder_literal(price_to_efd_adder)
    ga_rates = db.get_ground_advantage_retail_rates(conn)
    pm_by_day: dict[str, tuple] = {}

    def _pm_for_billing_row(ts_raw: Any) -> tuple:
        od = db.iso_date_from_billing_timestamp(ts_raw)
        key = od if od is not None else end_date
        if key not in pm_by_day:
            pm_by_day[key] = db.get_priority_mail_retail_rates(conn, as_of_date=key)
        return pm_by_day[key]

    rows = db.query_parcel_billing_rows_full(
        conn,
        start_date,
        end_date,
        parent_number,
        customer_number,
        show_parents,
        show_main,
        profit_account_ids,
    )
    title = db.parcel_summary_title_name(conn, parent_number, customer_number)
    conn_for_cost = conn

    hdr_row = _EFD_PARCEL_INVOICE_HEADER_ROW
    data_start = _EFD_PARCEL_INVOICE_DATA_START
    col_x = 24
    col_y = 25
    col_z = 26
    col_aa = 27
    lx = get_column_letter(col_x)
    ly = get_column_letter(col_y)
    laa = get_column_letter(col_aa)

    if set_sheet_title is not None:
        ws.title = set_sheet_title

    font_body = Font(name="Calibri", size=11)
    font_hdr = Font(name="Calibri", bold=True, size=11)

    ws.cell(1, 1, title).font = font_body
    ws.cell(1, 2, _short_mdy(start_date)).font = font_body
    ws.cell(1, 3, _short_mdy(end_date)).font = font_body

    ws.cell(2, 1, "Parcel Cost").font = font_body
    ws.cell(3, 1, "Price to EFD").font = font_body
    ws.cell(4, 1, "Price To customer").font = font_body
    ws.cell(5, 1, "EFD Profit").font = font_body
    ws.cell(6, 1, "Total parcel quantities").font = font_body

    n = len(rows)
    last_data = hdr_row + n
    if n == 0:
        b2 = ws.cell(2, 2, 0)
        b2.font = font_body
        b2.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT
        for rr in (3, 4, 5):
            c = ws.cell(rr, 2, 0)
            c.font = font_body
            c.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT
        b6 = ws.cell(6, 2, 0)
        b6.font = font_body
        b6.number_format = _EFD_PARCEL_INVOICE_QTY_FMT
    else:
        b2 = ws.cell(2, 2, f"=SUM({lx}{data_start}:{lx}{last_data})")
        b2.font = font_body
        b2.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT
        b3 = ws.cell(3, 2, f"=SUM({ly}{data_start}:{ly}{last_data})")
        b3.font = font_body
        b3.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT
        b4 = ws.cell(4, 2, f"=SUM({laa}{data_start}:{laa}{last_data})")
        b4.font = font_body
        b4.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT
        b5 = ws.cell(5, 2, "=B4-B3")
        b5.font = font_body
        b5.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT
        b6 = ws.cell(
            6,
            2,
            f"=COUNTA(A{data_start}:A{last_data})",
        )
        b6.font = font_body
        b6.number_format = _EFD_PARCEL_INVOICE_QTY_FMT

    headers: list[str] = [
        *_EFD_PARCEL_INVOICE_PREFIX_KEYS,
        "billing_amount",
        "Price to EFD",
        "EFD Revenue",
        "retail_cost",
        *_EFD_PARCEL_INVOICE_TAIL_KEYS,
    ]
    for col_1, h in enumerate(headers, 1):
        cell = ws.cell(hdr_row, col_1, h)
        cell.font = font_hdr

    for i, r in enumerate(rows):
        excel_row = data_start + i
        for j, key in enumerate(_EFD_PARCEL_INVOICE_PREFIX_KEYS, 1):
            v = r[key] if key in r.keys() else None
            ws.cell(excel_row, j, v).font = font_body

        bill = r["billing_amount"] if "billing_amount" in r.keys() else None
        cx = ws.cell(excel_row, col_x, bill)
        cx.font = font_body
        if bill is not None:
            try:
                cx.value = round(float(bill), 2)
            except (TypeError, ValueError):
                pass
        cx.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT

        cy = ws.cell(excel_row, col_y, f"={lx}{excel_row}+{adder_lit}")
        cy.font = font_body
        cy.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT

        cz = ws.cell(excel_row, col_z, f"={laa}{excel_row}-{ly}{excel_row}")
        cz.font = font_body
        cz.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT

        weight_oz = r["weight_oz"] if "weight_oz" in r.keys() else None
        zone_raw = r["zone"] if "zone" in r.keys() else None
        fallback_retail = r["fully_paid_postage"] if "fully_paid_postage" in r.keys() else None
        ts_raw = r["time_stamp"] if "time_stamp" in r.keys() else None
        priced = db.compute_retail_cost_for_piece(
            conn_for_cost,
            weight_oz=weight_oz,
            zone_raw=zone_raw,
            fallback_retail=fallback_retail,
            ga_rates=ga_rates,
            pm_rates=_pm_for_billing_row(ts_raw),
        )
        retail_val = priced.get("retail")
        if retail_val is not None:
            try:
                retail_val = round(float(retail_val), 2)
            except (TypeError, ValueError):
                retail_val = None

        caa = ws.cell(excel_row, col_aa, retail_val)
        caa.font = font_body
        caa.number_format = _EFD_PARCEL_INVOICE_MONEY_FMT

        tail_start = col_aa + 1
        for k, key in enumerate(_EFD_PARCEL_INVOICE_TAIL_KEYS):
            v = r[key] if key in r.keys() else None
            ws.cell(excel_row, tail_start + k, v).font = font_body

    for col_1 in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_1)].width = min(
            48.0, max(9.0, len(str(headers[col_1 - 1])) * 1.1)
        )

    last_data_row = last_data if n else hdr_row
    return title, hdr_row, data_start, last_data_row


def efd_parcel_invoice_download_name(
    title_name: str,
    *,
    parent_number: int | None,
    customer_number: int | None,
    end_date: str,
) -> str:
    """Filename: Parent Name (custcode) M-D-YYYY.xlsx"""
    base = _safe_filename_piece((title_name or "Report").strip())
    end_short = _safe_filename_piece(_short_mdy(end_date))
    if parent_number is not None:
        code = int(parent_number)
    elif customer_number is not None:
        code = int(customer_number)
    else:
        code = "ALL"
    return f"{base} ({code}) {end_short}.xlsx"


def export_efd_parcel_invoice_xlsx(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
    show_parents: bool,
    show_main: bool,
    *,
    efd_parcel_fee: float = 1.25,
) -> tuple[Path, str]:
    """
    EFD Parcel Invoice workbook: summary + line items (same rows as consolidated parcel CSV),
    fixed columns so billing_amount is column X and retail_cost is column AA.

    Returns (temp xlsx path, display title used in A1 / filename).
    """
    conn = db.get_connection()
    try:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        fee = max(0.0, float(efd_parcel_fee or 0.0))
        title, _hr, _ds, _ld = fill_efd_parcel_invoice_worksheet(
            ws,
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            set_sheet_title="EFD Parcel Invoice",
            price_to_efd_adder=fee,
        )
        scope = parcel_report_scope_label(parent_number, customer_number)
        fd, tmp = tempfile.mkstemp(
            suffix=f"_EFD_Parcel_{scope}_{start_date}_{end_date}.xlsx",
            prefix="efd_parcel_invoice_",
        )
        os.close(fd)
        out = Path(tmp)
        wb.save(out)
        return out, title
    finally:
        conn.close()


# EFD weekly invoice: fixed parent accounts and summary tab layout (Helper Files/EFD4-27 to 5-1.xlsx).
_EFD_WEEKLY_ACCOUNTS: tuple[tuple[int, str], ...] = (
    (3901, "BCBS"),
    (3899, "GEHA"),
    (3900, "Zinnia"),
)
# Shared EFD parent list for weekly invoice and daily volumes exports.
EFD_REPORT_ACCOUNTS: tuple[tuple[int, str], ...] = _EFD_WEEKLY_ACCOUNTS

_EFD_WEEKLY_MONEY_FMT = "$#,##0.00"
_EFD_WEEKLY_QTY_FMT = "#,##0"
_EFD_WEEKLY_SINGLE_POSTAGE_SHEET = "Postage"
_EFD_WEEKLY_SINGLE_PARCEL_SHEET = "Parcel"
_EFD_WEEKLY_PARENT_TO_LABEL: dict[int, str] = dict(_EFD_WEEKLY_ACCOUNTS)


def efd_report_scope_label(parent_number: int) -> str:
    """Display scope for filenames, e.g. ``BCBS (3901)``."""
    label = efd_weekly_summary_label(parent_number)
    return f"{label} ({int(parent_number)})"


def efd_weekly_summary_label(parent_number: int) -> str:
    """Summary tab label (BCBS, GEHA, Zinnia) for an EFD weekly parent number."""
    label = _EFD_WEEKLY_PARENT_TO_LABEL.get(int(parent_number))
    if label is None:
        allowed = ", ".join(str(p) for p, _ in _EFD_WEEKLY_ACCOUNTS)
        raise ValueError(
            f"parent_number must be one of EFD weekly accounts ({allowed}), got {parent_number}"
        )
    return label


def _efd_weekly_accounts_for_export(
    parent_number: int | None,
) -> list[tuple[int, str]]:
    if parent_number is None:
        return list(_EFD_WEEKLY_ACCOUNTS)
    label = _EFD_WEEKLY_PARENT_TO_LABEL.get(int(parent_number))
    if label is None:
        allowed = ", ".join(str(p) for p, _ in _EFD_WEEKLY_ACCOUNTS)
        raise ValueError(
            f"parent_number must be one of EFD weekly accounts ({allowed}), got {parent_number}"
        )
    return [(int(parent_number), label)]


def _efd_weekly_postage_sheet_name(parent_number: int, *, combined_sheet_names: bool) -> str:
    if combined_sheet_names:
        return f"{parent_number} Postage"[:31]
    return _EFD_WEEKLY_SINGLE_POSTAGE_SHEET


def _efd_weekly_parcel_sheet_name(parent_number: int, *, combined_sheet_names: bool) -> str:
    if combined_sheet_names:
        return f"{parent_number} Parcel"[:31]
    return _EFD_WEEKLY_SINGLE_PARCEL_SHEET


def _excel_sheet_ref(sheet_name: str, cell: str) -> str:
    escaped = str(sheet_name).replace("'", "''")
    return f"='{escaped}'!{cell}"


def efd_weekly_invoice_download_name(start_date: str, end_date: str) -> str:
    """Example: EFD 4-27 to 5-1.xlsx (month-day when range is within one calendar year)."""
    try:
        s = datetime.strptime(start_date, "%Y-%m-%d")
        e = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        return _safe_filename_piece(f"EFD {_short_mdy(start_date)} to {_short_mdy(end_date)}.xlsx")
    if start_date == end_date:
        return _safe_filename_piece(f"EFD {s.month}-{s.day}-{s.year}.xlsx")
    if s.year == e.year:
        return _safe_filename_piece(f"EFD {s.month}-{s.day} to {e.month}-{e.day}.xlsx")
    return _safe_filename_piece(f"EFD {_short_mdy(start_date)} to {_short_mdy(end_date)}.xlsx")


def efd_weekly_account_download_name(
    summary_label: str,
    start_date: str,
    end_date: str,
) -> str:
    """Example: EFD BCBS 4-27 to 5-1.xlsx"""
    label = _safe_filename_piece(summary_label.strip() or "Account")
    try:
        s = datetime.strptime(start_date, "%Y-%m-%d")
        e = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        return _safe_filename_piece(
            f"EFD {label} {_short_mdy(start_date)} to {_short_mdy(end_date)}.xlsx"
        )
    if start_date == end_date:
        return _safe_filename_piece(f"EFD {label} {s.month}-{s.day}-{s.year}.xlsx")
    if s.year == e.year:
        return _safe_filename_piece(
            f"EFD {label} {s.month}-{s.day} to {e.month}-{e.day}.xlsx"
        )
    return _safe_filename_piece(
        f"EFD {label} {_short_mdy(start_date)} to {_short_mdy(end_date)}.xlsx"
    )


def _write_efd_weekly_summary_sheet(
    ws,
    account_refs: list[tuple[str, str, str]],
) -> None:
    """
    Summary grid: each account has Flats, Parcels, and Total rows; grand total at bottom.

    ``account_refs``: list of (summary_label, postage_sheet_name, parcel_sheet_name).
    """
    font_body = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", bold=True, size=11)

    subtotal_cost_rows: list[int] = []
    subtotal_qty_rows: list[int] = []
    row = 1
    for label, postage_sheet, parcel_sheet in account_refs:
        flats_row = row
        ws.cell(row, 1, label).font = font_body
        ws.cell(row, 2, "Flats").font = font_body
        c_flats = ws.cell(row, 3, _excel_sheet_ref(postage_sheet, "L31"))
        c_flats.font = font_body
        c_flats.number_format = _EFD_WEEKLY_MONEY_FMT
        d_flats = ws.cell(row, 4, _excel_sheet_ref(postage_sheet, "I32"))
        d_flats.font = font_body
        d_flats.number_format = _EFD_WEEKLY_QTY_FMT
        row += 1

        parcels_row = row
        ws.cell(row, 1, label).font = font_body
        ws.cell(row, 2, "Parcels").font = font_body
        c_parc = ws.cell(row, 3, _excel_sheet_ref(parcel_sheet, "B3"))
        c_parc.font = font_body
        c_parc.number_format = _EFD_WEEKLY_MONEY_FMT
        d_parc = ws.cell(row, 4, _excel_sheet_ref(parcel_sheet, "B6"))
        d_parc.font = font_body
        d_parc.number_format = _EFD_WEEKLY_QTY_FMT
        row += 1

        ws.cell(row, 1, label).font = font_bold
        ws.cell(row, 2, "Total").font = font_bold
        c_sub = ws.cell(row, 3, f"=SUM(C{flats_row}:C{parcels_row})")
        c_sub.font = font_bold
        c_sub.number_format = _EFD_WEEKLY_MONEY_FMT
        d_sub = ws.cell(row, 4, f"=SUM(D{flats_row}:D{parcels_row})")
        d_sub.font = font_bold
        d_sub.number_format = _EFD_WEEKLY_QTY_FMT
        subtotal_cost_rows.append(row)
        subtotal_qty_rows.append(row)
        row += 1

    if len(account_refs) > 1:
        cost_refs = ",".join(f"C{r}" for r in subtotal_cost_rows)
        qty_refs = ",".join(f"D{r}" for r in subtotal_qty_rows)
        ws.cell(row, 1, "Grand Total").font = font_bold
        ws.cell(row, 2, "Grand Total").font = font_bold
        c_grand = ws.cell(row, 3, f"=SUM({cost_refs})")
        c_grand.font = font_bold
        c_grand.number_format = _EFD_WEEKLY_MONEY_FMT
        d_grand = ws.cell(row, 4, f"=SUM({qty_refs})")
        d_grand.font = font_bold
        d_grand.number_format = _EFD_WEEKLY_QTY_FMT

    ws.column_dimensions["A"].width = 12.0
    ws.column_dimensions["B"].width = 12.0
    ws.column_dimensions["C"].width = 14.0
    ws.column_dimensions["D"].width = 12.0


def _build_efd_weekly_workbook(
    conn: sqlite3.Connection,
    start_date: str,
    end_date: str,
    accounts: list[tuple[int, str]],
    *,
    combined_sheet_names: bool,
    discount: float,
    efd_parcel_fee: float,
    show_parents: bool,
    show_main: bool,
    remove_zeros: bool,
    hide_costs: bool,
    hide_savings: bool,
) -> Workbook:
    fee = max(0.0, float(efd_parcel_fee or 0.0))
    wb = Workbook()
    wb.remove(wb.active)

    account_refs: list[tuple[str, str, str]] = []
    for parent_number, summary_label in accounts:
        postage_sheet = _efd_weekly_postage_sheet_name(
            parent_number, combined_sheet_names=combined_sheet_names
        )
        parcel_sheet = _efd_weekly_parcel_sheet_name(
            parent_number, combined_sheet_names=combined_sheet_names
        )

        ws_postage = wb.create_sheet(postage_sheet)
        fill_postage_invoice_worksheet(
            ws_postage,
            conn,
            parent_number,
            start_date,
            end_date,
            discount=discount,
            show_parents=show_parents,
            show_main=show_main,
            remove_zeros=remove_zeros,
            hide_costs=hide_costs,
            hide_savings=hide_savings,
            set_sheet_title=postage_sheet,
            empty_invoice_if_no_data=True,
        )

        ws_parcel = wb.create_sheet(parcel_sheet)
        fill_efd_parcel_invoice_worksheet(
            ws_parcel,
            conn,
            start_date,
            end_date,
            parent_number,
            None,
            show_parents,
            show_main,
            set_sheet_title=parcel_sheet,
            price_to_efd_adder=fee,
        )

        account_refs.append((summary_label, postage_sheet, parcel_sheet))

    ws_summary = wb.create_sheet("Summary", 0)
    _write_efd_weekly_summary_sheet(ws_summary, account_refs)
    return wb


def export_efd_weekly_invoice_xlsx(
    start_date: str,
    end_date: str,
    *,
    discount: float = 0.10,
    efd_parcel_fee: float = 1.25,
    show_parents: bool = True,
    show_main: bool = True,
    remove_zeros: bool = False,
    hide_costs: bool = False,
    hide_savings: bool = False,
    parent_number: int | None = None,
) -> Path:
    """
    EFD weekly workbook: Summary plus postage and EFD parcel per account.

    ``parent_number`` None → all three EFD parents in one file (combined sheet names).
    ``parent_number`` set → single-account file (Summary, Postage, Parcel).
    """
    accounts = _efd_weekly_accounts_for_export(parent_number)
    combined_sheet_names = parent_number is None
    conn = db.get_connection()
    try:
        wb = _build_efd_weekly_workbook(
            conn,
            start_date,
            end_date,
            accounts,
            combined_sheet_names=combined_sheet_names,
            discount=discount,
            efd_parcel_fee=efd_parcel_fee,
            show_parents=show_parents,
            show_main=show_main,
            remove_zeros=remove_zeros,
            hide_costs=hide_costs,
            hide_savings=hide_savings,
        )
        if parent_number is None:
            suffix = f"_EFD_Weekly_{start_date}_{end_date}.xlsx"
            prefix = "efd_weekly_"
        else:
            suffix = f"_EFD_Weekly_{parent_number}_{start_date}_{end_date}.xlsx"
            prefix = "efd_weekly_acct_"
        fd, tmp = tempfile.mkstemp(suffix=suffix, prefix=prefix)
        os.close(fd)
        out = Path(tmp)
        wb.save(out)
        return out
    finally:
        conn.close()


POSTAGE_REPORTS_DIR = db.ROOT / "PostageReports"
_EFD_WEEKLY_BUNDLE_ACCOUNT_PARENTS: tuple[int, ...] = (3901, 3899, 3900)


def efd_weekly_bundle_folder_name(end_date: str) -> str:
    """Folder name from end date, e.g. ``Weekly EFD 6-12-26``."""
    try:
        d = datetime.strptime(end_date, "%Y-%m-%d")
        yy = d.year % 100
        return f"Weekly EFD {d.month}-{d.day}-{yy}"
    except ValueError:
        return f"Weekly EFD {_short_mdy(end_date)}"


def efd_weekly_bundle_output_dir(end_date: str) -> Path:
    """``PostageReports/Weekly EFD M-D-YY/`` under project root."""
    out_dir = POSTAGE_REPORTS_DIR / efd_weekly_bundle_folder_name(end_date)
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _efd_weekly_customer_name(parent_number: int) -> str:
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT customer_name FROM customers WHERE customer_number = ?",
            (int(parent_number),),
        ).fetchone()
        return (row["customer_name"] if row else f"Account {parent_number}").strip()
    finally:
        conn.close()


def _save_temp_export(temp: Path, out_dir: Path, filename: str) -> None:
    dest = out_dir / filename
    shutil.copy2(temp, dest)
    temp.unlink(missing_ok=True)


def save_efd_weekly_bundle(
    start_date: str,
    end_date: str,
    *,
    discount: float = 0.10,
    postage_discount: float = 0.10,
    efd_parcel_fee: float = 1.25,
    parcel_discount: float = 0.25,
    show_parents: bool = True,
    show_main: bool = True,
    remove_zeros: bool = False,
    hide_costs: bool = False,
    hide_savings: bool = False,
) -> dict[str, Any]:
    """
    Generate all 10 EFD weekly bundle reports and save under
    ``PostageReports/Weekly EFD {end-date}/``.
    """
    out_dir = efd_weekly_bundle_output_dir(end_date)
    folder_relative = f"PostageReports/{efd_weekly_bundle_folder_name(end_date)}"
    saved: list[str] = []
    failed: list[dict[str, str]] = []

    try:
        temp = export_efd_weekly_invoice_xlsx(
            start_date,
            end_date,
            discount=float(discount),
            efd_parcel_fee=float(efd_parcel_fee),
            show_parents=show_parents,
            show_main=show_main,
            remove_zeros=remove_zeros,
            hide_costs=hide_costs,
            hide_savings=hide_savings,
            parent_number=None,
        )
        name = efd_weekly_invoice_download_name(start_date, end_date)
        _save_temp_export(temp, out_dir, name)
        saved.append(name)
    except Exception as e:
        failed.append({"label": "Weekly invoice (combined)", "error": str(e)})

    for pn in _EFD_WEEKLY_BUNDLE_ACCOUNT_PARENTS:
        summary_label = efd_weekly_summary_label(pn)
        try:
            temp = export_efd_weekly_invoice_xlsx(
                start_date,
                end_date,
                discount=float(discount),
                efd_parcel_fee=float(efd_parcel_fee),
                show_parents=show_parents,
                show_main=show_main,
                remove_zeros=remove_zeros,
                hide_costs=hide_costs,
                hide_savings=hide_savings,
                parent_number=pn,
            )
            name = efd_weekly_account_download_name(summary_label, start_date, end_date)
            _save_temp_export(temp, out_dir, name)
            saved.append(name)
        except Exception as e:
            failed.append({"label": f"Weekly invoice ({pn})", "error": str(e)})

    for pn in _EFD_WEEKLY_BUNDLE_ACCOUNT_PARENTS:
        summary_label = efd_weekly_summary_label(pn)
        try:
            temp = export_postage_invoice(
                pn,
                start_date,
                end_date,
                discount=float(postage_discount),
                show_parents=show_parents,
                show_main=show_main,
                remove_zeros=remove_zeros,
                hide_costs=hide_costs,
                hide_savings=hide_savings,
            )
            cust_name = _efd_weekly_customer_name(pn)
            name = postage_invoice_download_name(
                title_name=cust_name,
                parent_number=pn,
                end_date=end_date,
            )
            _save_temp_export(temp, out_dir, name)
            saved.append(name)
        except Exception as e:
            failed.append({"label": f"Postage invoice ({summary_label})", "error": str(e)})

    for pn in _EFD_WEEKLY_BUNDLE_ACCOUNT_PARENTS:
        summary_label = efd_weekly_summary_label(pn)
        try:
            conn = db.get_connection()
            try:
                summary = db.query_parcel_zone_summary(
                    conn,
                    start_date,
                    end_date,
                    parent_number=pn,
                    customer_number=None,
                    show_parents=show_parents,
                    show_main=show_main,
                    hide_costs=False,
                    parcel_discount=float(parcel_discount),
                )
            finally:
                conn.close()
            temp = export_parcel_zone_summary_xlsx(
                summary,
                start_date=start_date,
                end_date=end_date,
                parent_number=pn,
                customer_number=None,
                show_parents=show_parents,
                show_main=show_main,
                parcel_discount=float(parcel_discount),
            )
            name = efd_account_report_download_name(
                title_name=summary.get("title_name"),
                parent_number=pn,
                report_label="Parcel invoice",
                start_date=start_date,
                end_date=end_date,
                ext="xlsx",
            )
            _save_temp_export(temp, out_dir, name)
            saved.append(name)
        except Exception as e:
            failed.append({"label": f"Parcel invoice ({summary_label})", "error": str(e)})

    return {
        "folder": str(out_dir.resolve()),
        "folder_relative": folder_relative,
        "saved": saved,
        "failed": failed,
    }


# Stacked zone block uses columns A–I (Parcel Summary Final example.xlsx).
_PARCEL_SUMMARY_STACK_COL_WIDTHS: list[float] = [
    12.0,
    30.0,
    13.0,
    6.6640625,
    18.0,
    25.0,
    13.0,
    13.0,
    12.0,
]


def export_parcel_zone_summary_xlsx(
    summary: dict[str, Any],
    *,
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None = None,
    show_parents: bool = True,
    show_main: bool = True,
    parcel_discount: float = 0.25,
) -> Path:
    """Zone × weight workbook from `query_parcel_zone_summary` (CSV retail/EFD rates × DB quantities).

    **Sheet 1 — Parcel Invoice:** four zone pairs stacked vertically in columns **A–I**. Per-row
    **Costs** (H) = ``C×D + F×G`` (discount × count for each zone side); **Savings** (I) =
    ``parcel_discount × (D+G)``. Footer totals match the same rules; then **over-10lb** and
    **per-customer invoice** blocks below.

    **Sheet 2 — Parcel Report:** line-item detail only (same as standalone parcel report export);
    tab title is the date range.
    """
    blocks = summary["blocks"]
    num_blocks = len(blocks)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parcel Invoice"

    fill_prior = PatternFill("solid", start_color="D9E2F3")
    fill_efd = PatternFill("solid", start_color="D6F5F5")
    fill_cost = PatternFill("solid", start_color="F8CBAD")
    fill_save = PatternFill("solid", start_color="C6EFCE")
    fill_title = PatternFill("solid", start_color="2F5597")
    font_hdr = Font(name="Arial", bold=True, size=10)
    font_title = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    font_tot_w = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    font_date = Font(name="Calibri", size=11)
    font_body = Font(name="Calibri", size=11)
    font_tot_label = Font(name="Arial", bold=True, size=11)
    side = Side(style="thin", color="000000")
    grid = Border(left=side, right=side, top=side, bottom=side)
    cur_fmt = "$#,##0.00"
    int_fmt = "#,##0"

    raw_title = (summary.get("title_name") or "Parcel Invoice").strip()
    display_title = (
        raw_title if raw_title.rstrip().endswith("-EFD") else f"{raw_title} -EFD"
    )

    a1 = ws.cell(1, 1, summary.get("report_date") or "")
    a1.font = font_date
    ws.merge_cells("B1:I1")
    tcell = ws.cell(1, 2, display_title)
    tcell.font = font_title
    tcell.fill = fill_title
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 18.0

    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="right", vertical="center")
    center_vert = Alignment(vertical="center")

    current_row = 2
    for bi, block in enumerate(blocks):
        ws.row_dimensions[current_row].height = 29.0
        za, zb = block["zone_a"], block["zone_b"]
        hdrs = [
            "Weight",
            f"Retail Zone {za}",
            f"Discount Zone {za}",
            "Count",
            f"Retail Zone {zb}",
            f"Discount Zone {zb}",
            "Count",
            "Costs",
            "Savings",
        ]
        for col_1, h in enumerate(hdrs, 1):
            cell = ws.cell(current_row, col_1, h)
            cell.font = font_hdr
            cell.border = grid
            cell.alignment = hdr_align
            rel = col_1 - 1
            if rel in (1, 4):
                cell.fill = fill_prior
            elif rel in (2, 5):
                cell.fill = fill_efd
            elif rel == 7:
                cell.fill = fill_cost
            elif rel == 8:
                cell.fill = fill_save

        for ri in range(10):
            r = current_row + 1 + ri
            ws.row_dimensions[r].height = 16.0
            rowrec = block["rows"][ri]
            a = rowrec["zone_a"]
            b = rowrec["zone_b"]
            pd = float(parcel_discount or 0.0)
            vals: list[Any] = [
                rowrec["weight_label"],
                a.get("priority"),
                a.get("efd"),
                a.get("count"),
                b.get("priority"),
                b.get("efd"),
                b.get("count"),
                rowrec.get("costs"),
                rowrec.get("savings"),
            ]
            for col_1, v in enumerate(vals, 1):
                rel = col_1 - 1
                hide_row_costs = summary.get("hide_costs")
                if not hide_row_costs and rel == 7:
                    cell = ws.cell(r, col_1, f"=C{r}*D{r}+F{r}*G{r}")
                elif not hide_row_costs and rel == 8:
                    cell = ws.cell(r, col_1, f"={pd}*(D{r}+G{r})")
                else:
                    cell = ws.cell(r, col_1, v if v is not None else "")
                cell.font = font_body
                cell.border = grid
                if rel == 0:
                    cell.alignment = center_vert
                else:
                    cell.alignment = data_align
                if rel in (1, 4):
                    cell.fill = fill_prior
                elif rel in (2, 5):
                    cell.fill = fill_efd
                elif rel == 7:
                    cell.fill = fill_cost
                elif rel == 8:
                    cell.fill = fill_save
                if rel in (1, 2, 4, 5) and v is not None and v != "":
                    cell.number_format = cur_fmt
                if rel in (3, 6) and v is not None and v != "":
                    cell.number_format = int_fmt
                if rel in (7, 8) and not hide_row_costs:
                    cell.number_format = cur_fmt

        current_row += 11
        if bi < num_blocks - 1:
            current_row += 2

    tot_row = current_row + 1
    tp = summary.get("total_pieces", 0)
    m = ws.cell(tot_row, 2, f"Total Pieces: {tp:,}")
    m.alignment = Alignment(horizontal="center", vertical="center")
    m.font = font_tot_label
    m.border = grid

    tc = summary.get("total_cost")
    ts = summary.get("total_savings")
    cell_cost = ws.cell(tot_row, 8, tc if tc is not None else "")
    cell_cost.font = font_tot_w
    cell_cost.fill = PatternFill("solid", start_color="C00000")
    cell_cost.number_format = cur_fmt
    cell_cost.border = grid
    cell_cost.alignment = Alignment(horizontal="right", vertical="center")
    cell_save = ws.cell(tot_row, 9, ts if ts is not None else "")
    cell_save.font = font_tot_w
    cell_save.fill = PatternFill("solid", start_color="000000")
    cell_save.number_format = cur_fmt
    cell_save.border = grid
    cell_save.alignment = Alignment(horizontal="right", vertical="center")

    conn = db.get_connection()
    try:
        sections = db.compute_parcel_report_af_hm_sections(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            parcel_discount=float(parcel_discount or 0.0),
        )
    finally:
        conn.close()
    section_start = tot_row + 3
    write_parcel_af_hm_sections(ws, section_start, sections)

    for col_1, w in enumerate(_PARCEL_SUMMARY_STACK_COL_WIDTHS, 1):
        letter = get_column_letter(col_1)
        cur = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(cur, w)
    for col in range(1, 10):
        letter = get_column_letter(col)
        w = ws.column_dimensions[letter].width or 0
        if col in _AF_HM_MIN_WIDTHS:
            ws.column_dimensions[letter].width = max(w, _AF_HM_MIN_WIDTHS[col])

    ws_report = wb.create_sheet(title=f"{start_date} to {end_date}"[:31])
    fill_parcel_report_worksheet(
        ws_report,
        start_date,
        end_date,
        parent_number,
        customer_number,
        show_parents,
        show_main,
    )

    out = Path(
        tempfile.mkstemp(
            suffix="_parcel_zone_summary.xlsx",
            prefix="parcel_",
        )[1]
    )
    wb.save(out)
    return out


def export_ground_advantage_zone_pricing_csv() -> Path:
    """
    Export Ground Advantage retail zone pricing from `ground_advantage_retail`
    (matrix rows by weight × zone, plus fee/note lines). One row per database row.
    If the table is empty, copies `Postage Price/USPS Ground Advantage Retail.csv` when present.
    """
    conn = db.get_connection()
    try:
        cur = conn.execute(
            """
            SELECT effective_date, source_file_name, row_type, weight_unit,
                   weight_max, zone, label, price
            FROM ground_advantage_retail
            ORDER BY sort_group, sort_order
            """
        )
        rows = [tuple(r) for r in cur.fetchall()]
    finally:
        conn.close()

    fallback_src = db.ROOT / "Postage Price" / "USPS Ground Advantage Retail.csv"
    if not rows and fallback_src.is_file():
        out = Path(
            tempfile.mkstemp(
                suffix="_Ground_Advantage_Zone_Pricing.csv",
                prefix="ga_zones_",
            )[1]
        )
        out.write_bytes(fallback_src.read_bytes())
        return out

    out = Path(
        tempfile.mkstemp(
            suffix="_Ground_Advantage_Zone_Pricing.csv",
            prefix="ga_zones_",
        )[1]
    )
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "effective_date",
                "source_file_name",
                "row_type",
                "weight_unit",
                "weight_max",
                "zone",
                "label",
                "price",
            ]
        )
        for r in rows:
            ed, sfn, rt, wu, wm, z, lab, price = r
            w.writerow(
                [
                    ed or "",
                    sfn or "",
                    rt or "",
                    wu or "",
                    "" if wm is None else wm,
                    "" if z is None else z,
                    lab or "",
                    "" if price is None else price,
                ]
            )
    return out


def ground_advantage_zone_pricing_download_name() -> str:
    """Filename for Ground Advantage zone pricing CSV (includes effective date when known)."""
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT MAX(effective_date) FROM ground_advantage_retail WHERE effective_date IS NOT NULL"
        ).fetchone()
        eff = (row[0] or "").strip() if row else ""
    finally:
        conn.close()
    if eff:
        safe = eff.replace("/", "-").replace(" ", "_")
        return f"Ground_Advantage_Zone_Pricing_{safe}.csv"
    return "Ground_Advantage_Zone_Pricing.csv"
