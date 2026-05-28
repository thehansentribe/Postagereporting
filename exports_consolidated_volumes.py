"""Consolidated volumes Excel export (flats + parcels + over-10lb + zone summary)."""

from __future__ import annotations

import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import db
from exports import _retail_rate, aggregate_parcel_count_rows


_FILENAME_BAD_CHARS_RE = re.compile(r'[\\/:*?"<>|]+')


def _safe_filename_piece(s: str) -> str:
    # Keep spaces/parentheses for readability, but remove characters that break downloads.
    out = _FILENAME_BAD_CHARS_RE.sub("-", s)
    out = re.sub(r"\s+", " ", out).strip()
    return out or "Report"


def _short_mdy(date_str: str) -> str:
    """M-D-YYYY with no leading zeros; falls back to raw if parsing fails."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return f"{d.month}-{d.day}-{d.year}"
    except ValueError:
        return str(date_str)


def consolidated_volumes_download_name(account_scope_label: str, end_date: str) -> str:
    scope = _safe_filename_piece(account_scope_label or "All Accounts")
    end_short = _safe_filename_piece(_short_mdy(end_date))
    return f"{scope} {end_short}.xlsx"


def total_presort_reject_pieces_from_postage_rows(rows: list[dict[str, Any]]) -> int:
    """WS3 presort rejects appear as synthetic postage rows (mail class Presort rejects)."""
    return sum(
        int(r.get("total_qty") or 0)
        for r in rows
        if r.get("mail_class") == db.WS3_REJECT_MAIL_CLASS
    )


def total_flats_retail_invoice_basis(
    rows: list[dict[str, Any]],
    retail_rates_by_oz: dict[int, float],
) -> float:
    """
    Sum of first-class retail for invoice flat classes, weights 1–13 oz only (same basis as
    postage invoice column M per oz row). Excludes synthetic Presort rejects rows.
    """
    flat_set = db.POSTAGE_INVOICE_FLAT_MAIL_CLASSES
    total = 0.0
    for r in rows:
        if r.get("mail_class") == db.WS3_REJECT_MAIL_CLASS:
            continue
        if r.get("mail_class") not in flat_set:
            continue
        for oz in range(1, 14):
            k = f"oz_{oz}"
            pc = int(r.get(k) or 0)
            if pc:
                total += float(pc) * float(_retail_rate(retail_rates_by_oz, oz))
    return round(total, 2)


def _find_block_side_for_zone(blocks: list[dict[str, Any]], z: int) -> tuple[dict[str, Any], str] | None:
    for b in blocks:
        if b.get("zone_a") == z:
            return (b, "a")
        if b.get("zone_b") == z:
            return (b, "b")
    return None


def _zone_cell_priority_count(blocks: list[dict[str, Any]], z: int, ri: int) -> tuple[Any, Any, int]:
    hit = _find_block_side_for_zone(blocks, z)
    if not hit:
        return None, None, 0
    block, side = hit
    rw = block["rows"][ri]
    cell = rw["zone_a"] if side == "a" else rw["zone_b"]
    pr = cell.get("priority")
    disc = cell.get("efd")
    cnt = int(cell.get("count") or 0)
    return pr, disc, cnt


def _parcel_consolidated_headers(hide_customer_numbers: bool) -> list[str]:
    """Column order must match the PARCELS (COUNTS) sheet (retail is last column)."""
    return [
        "Date",
        "Parent Name",
        *(["Parent Number"] if not hide_customer_numbers else []),
        "Child Name",
        *(["Child Number"] if not hide_customer_numbers else []),
        *[f"{i} lb" for i in range(1, 11)],
        "10+ lb",
        "Total Qty",
        "Retail Cost",
    ]


def _flats_consolidated_column_plan(hide_customer_numbers: bool) -> tuple[list[str], list[str]]:
    oz_keys = [f"oz_{i}" for i in range(13)] + ["oz_13", "oz_13plus"]
    headers = [
        "Date",
        "Parent Name",
        *(["Parent Number"] if not hide_customer_numbers else []),
        "Child Name",
        *(["Child Number"] if not hide_customer_numbers else []),
        "Class",
        *[f"{i} oz" for i in range(13)],
        "13 oz",
        "13+ oz",
        "Total Qty",
        "Retail cost",
    ]
    keys = [
        "date",
        "parent_name",
        *(["parent_number"] if not hide_customer_numbers else []),
        "child_name",
        *(["child_number"] if not hide_customer_numbers else []),
        "mail_class",
        *oz_keys,
        "total_qty",
        "retail_cost",
    ]
    return headers, keys


def flats_summary_retail_formula(
    *,
    flats_sheet_quoted: str,
    f_headers: list[str],
    retail_col_letter: str,
    last_data_row: int,
) -> str:
    """
    Excel formula for Summary **Flats Total retail Postage Cost**: sum the FLATS retail column
    minus rows whose Class is synthetic presort rejects or allocated rejects (matches
    ``query_postage`` total_retail_cost semantics).
    """
    class_idx = f_headers.index("Class") + 1
    class_col = get_column_letter(class_idx)

    def _xl_str(s: str) -> str:
        return '"' + str(s).replace('"', '""') + '"'

    pr = db.WS3_REJECT_MAIL_CLASS
    ar = db.WS3_REJECT_ALLOCATED_MAIL_CLASS
    rng_retail = (
        f"{flats_sheet_quoted}!{retail_col_letter}2:{retail_col_letter}{last_data_row}"
    )
    rng_class = f"{flats_sheet_quoted}!{class_col}2:{class_col}{last_data_row}"
    return (
        f"=SUM({rng_retail})"
        f"-SUMIF({rng_class},{_xl_str(pr)},{rng_retail})"
        f"-SUMIF({rng_class},{_xl_str(ar)},{rng_retail})"
    )


def _write_zone_summary_stacked_sheet(ws: Any, zone_data: dict[str, Any], *, start_row: int = 1) -> int:
    hide = zone_data.get("hide_costs") is True
    blocks = zone_data.get("blocks") or []
    if not blocks:
        c = ws.cell(start_row, 1, "No zone summary data.")
        c.font = Font(name="Calibri", size=11)
        return start_row + 1

    row_indices = [i for i in range(10) if i != 8]
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="2E5090")
    sec_font = Font(name="Calibri", bold=True, size=12)
    body = Font(name="Calibri", size=11)
    thin = Side(style="thin", color="B4B4B4")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = "$#,##0.00"
    int_fmt = "#,##0"

    r = start_row
    for z in range(1, 10):
        if _find_block_side_for_zone(blocks, z) is None:
            continue
        if r > start_row:
            r += 1
        title = ws.cell(r, 1, f"Zone {z}")
        title.font = sec_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        hdrs = ["Weight", f"Retail Z{z}", f"Discount Z{z}", f"Count Z{z}", "Retail total", "Savings"]
        for c, h in enumerate(hdrs, start=1):
            cell = ws.cell(r, c, h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = grid
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1
        for ri in row_indices:
            wl = blocks[0]["rows"][ri]["weight_label"]
            pri, disc, cnt = _zone_cell_priority_count(blocks, z, ri)
            retail_total: float | None = None
            savings_total: float | None = None
            if not hide and pri is not None:
                retail_total = round(float(pri) * cnt, 2)
                if disc is not None:
                    savings_total = round((float(pri) - float(disc)) * cnt, 2)
            c1 = ws.cell(r, 1, wl)
            c1.font = body
            c1.border = grid
            c1.alignment = Alignment(horizontal="left", vertical="center")
            pcell = ws.cell(r, 2)
            if not hide and pri is not None:
                pcell.value = float(pri)
                pcell.number_format = money_fmt
            pcell.font = body
            pcell.border = grid
            pcell.alignment = Alignment(horizontal="right", vertical="center")
            dcell = ws.cell(r, 3)
            if not hide and disc is not None:
                dcell.value = float(disc)
                dcell.number_format = money_fmt
            dcell.font = body
            dcell.border = grid
            dcell.alignment = Alignment(horizontal="right", vertical="center")
            ccell = ws.cell(r, 4, cnt)
            ccell.font = body
            ccell.border = grid
            ccell.number_format = int_fmt
            ccell.alignment = Alignment(horizontal="right", vertical="center")
            lcell = ws.cell(r, 5)
            if retail_total is not None:
                lcell.value = retail_total
                lcell.number_format = money_fmt
            lcell.font = body
            lcell.border = grid
            lcell.alignment = Alignment(horizontal="right", vertical="center")
            scell = ws.cell(r, 6)
            if savings_total is not None:
                scell.value = savings_total
                scell.number_format = money_fmt
            scell.font = body
            scell.border = grid
            scell.alignment = Alignment(horizontal="right", vertical="center")
            r += 1

    for col, w in enumerate([14.0, 14.0, 14.0, 12.0, 14.0, 14.0], start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, w)
    return r


def export_consolidated_volumes_xlsx(
    start_date: str,
    end_date: str,
    parent_number: int | None,
    customer_number: int | None,
    show_parents: bool,
    show_main: bool,
    consolidate: bool,
    remove_zeros: bool,
    hide_costs_summary: bool,
    hide_customer_numbers: bool,
    account_scope_label: str,
    parcel_discount: float = 0.25,
) -> Path:
    conn = db.get_connection()
    try:
        postage = db.query_postage(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            consolidate,
            remove_zeros,
            hide_costs=False,
        )
        parcels = db.query_parcels(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            consolidate,
            remove_zeros,
            hide_costs=False,
        )
        heavy = db.query_parcel_over_10lb_lines(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
        )
        zone = db.query_parcel_zone_summary(
            conn,
            start_date,
            end_date,
            parent_number,
            customer_number,
            show_parents,
            show_main,
            hide_costs=False,
            parcel_discount=float(parcel_discount or 0.0),
        )
    finally:
        conn.close()

    flat_rows = postage.get("rows") or []
    rejected_flats = total_presort_reject_pieces_from_postage_rows(flat_rows)
    parcel_agg = aggregate_parcel_count_rows(parcels.get("rows") or [])
    zone_pieces = int(zone.get("total_pieces") or 0)

    if not flat_rows and not parcel_agg and not heavy and zone_pieces == 0:
        raise ValueError("No data for the selected date range and filters.")

    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    body_font = Font(name="Calibri", size=11)
    label_font = Font(name="Calibri", bold=True, size=11)
    thin = Side(style="thin", color="B4B4B4")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = "$#,##0.00"
    int_fmt = "#,##0"

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:B1")
    t = ws_sum.cell(1, 1, "Consolidated volumes report")
    t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    def kv_row(row: int, label: str, value: Any, *, money: bool = False, is_int: bool = False) -> None:
        a = ws_sum.cell(row, 1, label)
        a.font = label_font
        a.alignment = Alignment(horizontal="left", vertical="center")
        b = ws_sum.cell(row, 2, value)
        b.font = body_font
        b.alignment = Alignment(horizontal="left", vertical="center")
        if money and value is not None:
            b.number_format = money_fmt
        if is_int:
            b.number_format = int_fmt

    r = 3
    kv_row(r, "Start date", start_date)
    r += 1
    kv_row(r, "End date", end_date)
    r += 1
    kv_row(r, "Account scope", account_scope_label)
    r += 1
    kv_row(r, "Total flats pieces (postage)", int(postage.get("total_pieces") or 0), is_int=True)
    r += 1
    kv_row(r, "Total parcel pieces", int(parcels.get("total_pieces") or 0), is_int=True)
    r += 1
    combined = int(postage.get("total_pieces") or 0) + int(parcels.get("total_pieces") or 0)
    kv_row(r, "Combined volume", combined, is_int=True)
    r += 1
    kv_row(r, "Total flats rejected items", rejected_flats, is_int=True)
    r += 1

    f_headers, _ = _flats_consolidated_column_plan(hide_customer_numbers)
    p_headers = _parcel_consolidated_headers(hide_customer_numbers)
    flats_retail_col = get_column_letter(len(f_headers))
    parcel_retail_col = get_column_letter(len(p_headers))
    flats_sheet_q = "'FLATS (POSTAGE)'"
    parcel_sheet_q = "'PARCELS (COUNTS)'"

    if not hide_costs_summary:
        row_flats_retail = r
        a1 = ws_sum.cell(r, 1, "Flats Total retail Postage Cost")
        a1.font = label_font
        a1.alignment = Alignment(horizontal="left", vertical="center")
        b1 = ws_sum.cell(r, 2)
        b1.font = body_font
        b1.alignment = Alignment(horizontal="left", vertical="center")
        if flat_rows:
            last_fr = 1 + len(flat_rows)
            b1.value = flats_summary_retail_formula(
                flats_sheet_quoted=flats_sheet_q,
                f_headers=list(f_headers),
                retail_col_letter=flats_retail_col,
                last_data_row=last_fr,
            )
            b1.number_format = money_fmt
        else:
            flat_num = (
                round(float(postage["total_retail_cost"]), 2)
                if postage.get("total_retail_cost") is not None
                else 0.0
            )
            b1.value = flat_num
            b1.number_format = money_fmt
        r += 1

        row_parcel_retail = r
        a2 = ws_sum.cell(r, 1, "Parcels Total Retail Postage cost")
        a2.font = label_font
        a2.alignment = Alignment(horizontal="left", vertical="center")
        b2 = ws_sum.cell(r, 2)
        b2.font = body_font
        b2.alignment = Alignment(horizontal="left", vertical="center")
        if parcel_agg:
            last_pr = 1 + len(parcel_agg)
            b2.value = f"=SUM({parcel_sheet_q}!{parcel_retail_col}2:{parcel_retail_col}{last_pr})"
            b2.number_format = money_fmt
        else:
            b2.value = round(float(parcels.get("total_retail") or 0.0), 2)
            b2.number_format = money_fmt
        r += 1

        a_tot = ws_sum.cell(r, 1, "Total Retail Cost")
        a_tot.font = label_font
        a_tot.alignment = Alignment(horizontal="left", vertical="center")
        b_tot = ws_sum.cell(r, 2, f"=B{row_flats_retail}+B{row_parcel_retail}")
        b_tot.font = body_font
        b_tot.alignment = Alignment(horizontal="left", vertical="center")
        b_tot.number_format = money_fmt
        r += 1

    ws_sum.column_dimensions["A"].width = 40.0
    ws_sum.column_dimensions["B"].width = 36.0

    if flat_rows:
        ws_f = wb.create_sheet("FLATS (POSTAGE)")
        f_headers, f_keys = _flats_consolidated_column_plan(hide_customer_numbers)
        for col, h in enumerate(f_headers, start=1):
            c = ws_f.cell(1, col, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = grid
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, row in enumerate(flat_rows, start=2):
            for col, k in enumerate(f_keys, start=1):
                v = row.get(k)
                cell = ws_f.cell(ri, col)
                cell.font = body_font
                cell.border = grid
                if k == "retail_cost":
                    if v is not None:
                        cell.value = round(float(v), 2)
                        cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif k.startswith("oz_") or k == "total_qty":
                    cell.value = int(v or 0)
                    cell.number_format = int_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.value = v
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_f.freeze_panes = "G2" if not hide_customer_numbers else "E2"
        ws_f.row_dimensions[1].height = 26
        for i in range(1, len(f_headers) + 1):
            letter = get_column_letter(i)
            if i == 1:
                base = 14.0
            elif i in (2, 4):
                base = 22.0
            elif not hide_customer_numbers and i in (3, 5):
                base = 14.0
            elif i <= (6 if not hide_customer_numbers else 4):
                base = 14.0
            else:
                base = 10.0
            ws_f.column_dimensions[letter].width = max(ws_f.column_dimensions[letter].width or 0, base)

    if parcel_agg:
        ws_p = wb.create_sheet("PARCELS (COUNTS)")
        p_headers = _parcel_consolidated_headers(hide_customer_numbers)
        for col, h in enumerate(p_headers, start=1):
            c = ws_p.cell(1, col, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = grid
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, prow in enumerate(parcel_agg, start=2):
            vals: list[Any] = [
                prow.get("date"),
                prow.get("parent_name"),
                *([prow.get("parent_number")] if not hide_customer_numbers else []),
                prow.get("child_name"),
                *([prow.get("child_number")] if not hide_customer_numbers else []),
                *[int(prow.get(f"lb_{i}") or 0) for i in range(1, 11)],
                int(prow.get("lb_10plus") or 0),
                int(prow.get("total_qty") or 0),
                round(float(prow.get("total_retail") or 0), 2),
            ]
            ncols = len(vals)
            for col, v in enumerate(vals, start=1):
                cell = ws_p.cell(ri, col, v)
                cell.font = body_font
                cell.border = grid
                if col <= (5 if not hide_customer_numbers else 3):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if col == ncols:
                    cell.number_format = money_fmt
                elif col >= (6 if not hide_customer_numbers else 4):
                    cell.number_format = int_fmt
        ws_p.freeze_panes = "F2" if not hide_customer_numbers else "D2"
        ws_p.row_dimensions[1].height = 28
        col_widths = [12.0, 24.0]
        if not hide_customer_numbers:
            col_widths.append(14.0)
        col_widths.append(30.0)
        if not hide_customer_numbers:
            col_widths.append(14.0)
        col_widths.extend([9.0] * 10 + [9.0, 11.0, 14.0])
        for i, w in enumerate(col_widths, start=1):
            letter = get_column_letter(i)
            ws_p.column_dimensions[letter].width = max(ws_p.column_dimensions[letter].width or 0, w)

    if heavy:
        ws_h = wb.create_sheet("Over 10 lb")
        hz = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        hz_fill = PatternFill("solid", start_color="375623")
        sorted_h = sorted(
            heavy,
            key=lambda x: (
                int(x.get("zone") or 0),
                str(x.get("child_name") or "").lower(),
                int(x.get("lbs") or 0),
            ),
        )
        r = 1
        prev_z: int | None = None
        for row in sorted_h:
            z = row.get("zone")
            z_int = int(z) if z is not None else 0
            if z_int != prev_z:
                if prev_z is not None:
                    r += 1
                ws_h.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                zt = ws_h.cell(r, 1, f"Zone {z}")
                zt.font = hz
                zt.fill = hz_fill
                zt.alignment = Alignment(horizontal="left", vertical="center")
                r += 1
                for c, h in enumerate(["Child name", "lbs", "Zone", "Cost (base)"], start=1):
                    hc = ws_h.cell(r, c, h)
                    hc.font = hdr_font
                    hc.fill = hdr_fill
                    hc.border = grid
                    hc.alignment = Alignment(horizontal="center", vertical="center")
                r += 1
                prev_z = z_int
            c1 = ws_h.cell(r, 1, row.get("child_name") or "")
            c1.font = body_font
            c1.border = grid
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c2 = ws_h.cell(r, 2, int(row.get("lbs") or 0))
            c2.font = body_font
            c2.border = grid
            c2.number_format = int_fmt
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c3 = ws_h.cell(r, 3, z)
            c3.font = body_font
            c3.border = grid
            c3.alignment = Alignment(horizontal="right", vertical="center")
            bc = ws_h.cell(r, 4, round(float(row.get("base") or 0), 2))
            bc.number_format = money_fmt
            bc.font = body_font
            bc.border = grid
            bc.alignment = Alignment(horizontal="right", vertical="center")
            r += 1
        for col, w in enumerate([32.0, 10.0, 10.0, 14.0], start=1):
            letter = get_column_letter(col)
            ws_h.column_dimensions[letter].width = max(ws_h.column_dimensions[letter].width or 0, w)

    if zone_pieces > 0 and (zone.get("blocks") or []):
        ws_z = wb.create_sheet("Zone summary")
        ws_z.merge_cells("A1:D1")
        zt = ws_z.cell(1, 1, "Parcel zone summary (Priority × count by zone)")
        zt.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        zt.fill = PatternFill("solid", start_color="2F5597")
        zt.alignment = Alignment(horizontal="center", vertical="center")
        ws_z.row_dimensions[1].height = 24
        _write_zone_summary_stacked_sheet(ws_z, zone, start_row=3)
        ws_z.freeze_panes = "A3"

    fd, tmp = tempfile.mkstemp(suffix="_consolidated_volumes.xlsx", prefix="volumes_")
    os.close(fd)
    out = Path(tmp)
    wb.save(out)
    return out
