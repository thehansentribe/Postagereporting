"""Tests for Profit Report (Flats) JSON endpoint.

These tests are skipped when Flask isn't installed in the test environment.
"""

from __future__ import annotations

import importlib
import tempfile
from pathlib import Path

import pytest

import db as dbmod
import watcher as watchermod

pytest.importorskip("flask")


def _client(monkeypatch, db_path: Path):
    """
    Build a Flask test client against an isolated temp DB.

    Notes:
    - app.py runs db.init_db() at import time; we set DB_PATH before reloading.
    - app.py also calls watcher.ensure_dirs() at import time; we patch it to no-op.
    - requests call a before_request hook that starts the watcher thread; patch that to no-op.
    """
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    monkeypatch.setattr(watchermod, "ensure_dirs", lambda: None)
    import app as appmod

    appmod = importlib.reload(appmod)
    monkeypatch.setattr(appmod, "_ensure_watcher", lambda: None)
    appmod.app.config.update(TESTING=True)
    return appmod.app.test_client()


def test_api_export_profit_report_rejects_negative_discount_efd(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_api_efd.db"
        client = _client(monkeypatch, p)
        r = client.get(
            "/api/export/profit-report-xlsx?start_date=2026-04-01&end_date=2026-04-07&discount_efd=-0.01"
        )
        assert r.status_code == 400
        j = r.get_json()
        assert "discount_efd" in j.get("error", "").lower()


def test_api_profit_flats_no_data_returns_404_with_export_message(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_empty.db"
        client = _client(monkeypatch, p)
        r = client.get("/api/profit/flats?start_date=2026-04-01&end_date=2026-04-07")
        assert r.status_code == 404
        j = r.get_json()
        assert j["empty"] is True
        assert "No WS3 flats profit rows found for this date range/account scope." in j["error"]
        # Meta still returned for UI to render context.
        assert j["meta"]["start_date"] == "2026-04-01"
        assert j["meta"]["end_date"] == "2026-04-07"
        assert j["meta"]["parcel_fee"] == pytest.approx(1.25, rel=1e-9)
        assert j["meta"].get("profit_accounts") in (None, [])


def test_api_profit_flats_success_returns_totals_and_rows(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_ok.db"
        client = _client(monkeypatch, p)

        conn = dbmod.get_connection()
        try:
            # Parent account used for WS3 profile scope.
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) VALUES (100, 'Parent Co', NULL, NULL)"
            )
            conn.execute(
                "INSERT INTO ws3_netsort_customers (customer_code, customer_name) VALUES ('301079', 'Acme Dept')"
            )
            conn.execute(
                "INSERT INTO ws3_mail_runs (mail_date, mail_id, run_datetime, source_file_name) VALUES ('2026-04-03', 'M1', '2026-04-03 12:00:00', 't.xlsx')"
            )
            run_id = conn.execute(
                "SELECT run_id FROM ws3_mail_runs WHERE mail_date='2026-04-03' AND mail_id='M1'"
            ).fetchone()["run_id"]
            conn.execute(
                "INSERT INTO ws3_profiles (profile_name, parent_customer_number, reject_fee) VALUES ('Profile 1', 100, NULL)"
            )
            profile_id = conn.execute(
                "SELECT id FROM ws3_profiles WHERE profile_name='Profile 1'"
            ).fetchone()["id"]
            # Minimal WS3 detail row with required fields (num_pieces>0, usps_cost_per_piece non-null).
            conn.execute(
                """
                INSERT INTO ws3_mail_detail (
                    run_id, profile_id, customer_code, rate_type,
                    postage_claimed, postage_applied, num_pieces, pcs_accepted, pcs_rejected,
                    cost_per_piece, usps_cost_per_piece
                ) VALUES (?, ?, '301079', 'ADC Auto', 10.00, 9.70, 10, 10, 0, 1.00, 1.0000)
                """,
                (run_id, profile_id),
            )
            conn.commit()
        finally:
            conn.close()

        r = client.get(
            "/api/profit/flats?start_date=2026-04-01&end_date=2026-04-07&parent_number=100&discount=0.10"
        )
        assert r.status_code == 200
        j = r.get_json()
        assert "meta" in j
        assert "totals" in j
        assert "rate_summary" in j
        assert "detail" in j
        assert j["meta"]["sell_to_rate"] == pytest.approx(1.53, rel=1e-9)
        assert j["totals"]["total_pieces"] == 10
        assert len(j["rate_summary"]) >= 1
        assert len(j["detail"]) >= 1


def test_api_profit_flats_single_piece_pass_through_zero_profit(monkeypatch):
    """Single Piece uses USPS cost as sell-to (no negative margin vs discounted flat rate)."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_single_piece.db"
        client = _client(monkeypatch, p)

        conn = dbmod.get_connection()
        try:
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) VALUES (100, 'Parent Co', NULL, NULL)"
            )
            conn.execute(
                "INSERT INTO ws3_netsort_customers (customer_code, customer_name) VALUES ('301079', 'Acme Dept')"
            )
            conn.execute(
                "INSERT INTO ws3_mail_runs (mail_date, mail_id, run_datetime, source_file_name) VALUES ('2026-04-03', 'M1', '2026-04-03 12:00:00', 't.xlsx')"
            )
            run_id = conn.execute(
                "SELECT run_id FROM ws3_mail_runs WHERE mail_date='2026-04-03' AND mail_id='M1'"
            ).fetchone()["run_id"]
            conn.execute(
                "INSERT INTO ws3_profiles (profile_name, parent_customer_number, reject_fee) VALUES ('Profile 1', 100, NULL)"
            )
            profile_id = conn.execute(
                "SELECT id FROM ws3_profiles WHERE profile_name='Profile 1'"
            ).fetchone()["id"]
            conn.execute(
                """
                INSERT INTO ws3_mail_detail (
                    run_id, profile_id, customer_code, rate_type,
                    postage_claimed, postage_applied, num_pieces, pcs_accepted, pcs_rejected,
                    cost_per_piece, usps_cost_per_piece
                ) VALUES (?, ?, '301079', 'ADC Auto', 123.00, 120.00, 100, 100, 0, 1.23, 1.2300)
                """,
                (run_id, profile_id),
            )
            conn.execute(
                """
                INSERT INTO ws3_mail_detail (
                    run_id, profile_id, customer_code, rate_type,
                    postage_claimed, postage_applied, num_pieces, pcs_accepted, pcs_rejected,
                    cost_per_piece, usps_cost_per_piece
                ) VALUES (?, ?, '301079', 'Single Piece', 29.34, 29.34, 18, 18, 0, 1.63, 1.6300)
                """,
                (run_id, profile_id),
            )
            conn.commit()
        finally:
            conn.close()

        r = client.get(
            "/api/profit/flats?start_date=2026-04-01&end_date=2026-04-07&parent_number=100&discount=0.10"
        )
        assert r.status_code == 200
        j = r.get_json()
        sell_to = j["meta"]["sell_to_rate"]
        assert sell_to == pytest.approx(1.53, rel=1e-9)

        sp = next(x for x in j["rate_summary"] if x["rate_type"] == "Single Piece")
        assert sp["total_pieces"] == 18
        assert sp["total_profit"] == pytest.approx(0.0, abs=0.01)
        assert sp["avg_profit_per_piece"] == pytest.approx(0.0, abs=0.0001)
        assert sp["sell_to_rate"] == pytest.approx(sp["avg_usps_cost_per_piece"], rel=1e-6)

        adc = next(x for x in j["rate_summary"] if x["rate_type"] == "ADC Auto")
        assert adc["sell_to_rate"] == pytest.approx(sell_to, rel=1e-6)
        expected_adc_profit = round((1.53 - 1.23) * 100, 2)
        assert adc["total_profit"] == pytest.approx(expected_adc_profit, rel=1e-6)

        assert j["totals"]["total_profit"] == pytest.approx(expected_adc_profit, rel=1e-6)


def test_api_profit_flats_post_json_matches_get_parent_scope(monkeypatch):
    """POST JSON carries the same scope and discount as GET query params."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_flats_post.db"
        client = _client(monkeypatch, p)

        conn = dbmod.get_connection()
        try:
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) VALUES (100, 'Parent Co', NULL, NULL)"
            )
            conn.execute(
                "INSERT INTO ws3_netsort_customers (customer_code, customer_name) VALUES ('301079', 'Acme Dept')"
            )
            conn.execute(
                "INSERT INTO ws3_mail_runs (mail_date, mail_id, run_datetime, source_file_name) VALUES ('2026-04-03', 'M1', '2026-04-03 12:00:00', 't.xlsx')"
            )
            run_id = conn.execute(
                "SELECT run_id FROM ws3_mail_runs WHERE mail_date='2026-04-03' AND mail_id='M1'"
            ).fetchone()["run_id"]
            conn.execute(
                "INSERT INTO ws3_profiles (profile_name, parent_customer_number, reject_fee) VALUES ('Profile 1', 100, NULL)"
            )
            profile_id = conn.execute(
                "SELECT id FROM ws3_profiles WHERE profile_name='Profile 1'"
            ).fetchone()["id"]
            conn.execute(
                """
                INSERT INTO ws3_mail_detail (
                    run_id, profile_id, customer_code, rate_type,
                    postage_claimed, postage_applied, num_pieces, pcs_accepted, pcs_rejected,
                    cost_per_piece, usps_cost_per_piece
                ) VALUES (?, ?, '301079', 'ADC Auto', 10.00, 9.70, 10, 10, 0, 1.00, 1.0000)
                """,
                (run_id, profile_id),
            )
            conn.commit()
        finally:
            conn.close()

        r_get = client.get(
            "/api/profit/flats?start_date=2026-04-01&end_date=2026-04-07&parent_number=100&discount=0.10"
        )
        assert r_get.status_code == 200
        j_get = r_get.get_json()

        r_post = client.post(
            "/api/profit/flats",
            json={
                "start_date": "2026-04-01",
                "end_date": "2026-04-07",
                "parent_number": 100,
                "discount": 0.10,
                "show_parents": True,
                "show_main": True,
            },
            content_type="application/json",
        )
        assert r_post.status_code == 200
        j_post = r_post.get_json()

        assert j_post["meta"]["sell_to_rate"] == j_get["meta"]["sell_to_rate"]
        assert j_post["totals"] == j_get["totals"]
        assert j_post["rate_summary"] == j_get["rate_summary"]
        assert j_post["detail"] == j_get["detail"]


def test_api_export_profit_report_uses_query_parcel_fee_fallback_for_summary_b10(monkeypatch):
    """When ``efd_parcel_fee`` is omitted, ``parcel_fee`` fills Summary B10 (bookmark compat)."""
    from io import BytesIO

    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_export_api_b10.db"
        client = _client(monkeypatch, p)
        r = client.get(
            "/api/export/profit-report-xlsx?"
            "start_date=2026-04-01&end_date=2026-04-07&parcel_fee=0.66"
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.data), data_only=False)
        try:
            assert wb["Summary"]["B10"].value == pytest.approx(0.66, rel=1e-9)
            assert wb["Summary"]["A10"].value == "Parcel fee to EFD ($/pc)"
        finally:
            wb.close()


def test_api_export_profit_report_efd_parcel_fee_overrides_parcel_fee_for_b10(monkeypatch):
    from io import BytesIO

    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_export_api_b10_split.db"
        client = _client(monkeypatch, p)
        r = client.get(
            "/api/export/profit-report-xlsx?"
            "start_date=2026-04-01&end_date=2026-04-07"
            "&parcel_fee=0.40&efd_parcel_fee=2.15"
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.data), data_only=False)
        try:
            assert wb["Summary"]["B10"].value == pytest.approx(2.15, rel=1e-9)
        finally:
            wb.close()


def test_api_export_efd_parcel_invoice_legacy_parcel_fee_query(monkeypatch):
    """``parcel_fee`` alone on the EFD export URL still sets column Y when ``efd_parcel_fee`` omitted."""
    from io import BytesIO

    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_api_legacy_parcel_fee.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()

        conn = dbmod.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (500, 'EfdFb')"
        )
        conn.execute(
            "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'x.csv', 1)"
        )
        imp_id = conn.execute("SELECT id FROM billing_imports WHERE billing_id='b1'").fetchone()["id"]
        conn.execute(
            """
            INSERT INTO billing_records (
                billing_import_id, custom_account_code, time_stamp, weight_oz,
                billing_amount, zone, imb_tracking_code, impb
            )
            VALUES (?, 500, '4/1/2026 15:34', 16.0, 4.00, '5', 'tr', 'ip')
            """,
            (imp_id,),
        )
        conn.commit()
        conn.close()

        monkeypatch.setattr(
            dbmod,
            "compute_retail_cost_for_piece",
            lambda *a, **k: {"retail": 9.0, "zone": 5},
        )

        monkeypatch.setattr(watchermod, "ensure_dirs", lambda: None)
        import app as appmod

        appmod = importlib.reload(appmod)
        monkeypatch.setattr(appmod, "_ensure_watcher", lambda: None)
        appmod.app.config.update(TESTING=True)
        client = appmod.app.test_client()

        r = client.get(
            "/api/export/efd-parcel-invoice-xlsx?"
            "start_date=2026-04-01&end_date=2026-04-30&parent_number=500"
            "&parcel_fee=0.91"
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.data), data_only=False)
        try:
            ws = wb.active
            assert ws.cell(10, 25).value == "=X10+0.91"
        finally:
            wb.close()



def _seed_ws3_flats(conn, runs):
    """
    Seed minimal WS3 data. ``runs`` maps mail_date -> list of
    (rate_type, postage_claimed, num_pieces, usps_cost_per_piece).
    """
    conn.execute(
        "INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) VALUES (100, 'Parent Co', NULL, NULL)"
    )
    conn.execute(
        "INSERT INTO ws3_netsort_customers (customer_code, customer_name) VALUES ('301079', 'Acme Dept')"
    )
    conn.execute(
        "INSERT INTO ws3_profiles (profile_name, parent_customer_number, reject_fee) VALUES ('Profile 1', 100, NULL)"
    )
    profile_id = conn.execute(
        "SELECT id FROM ws3_profiles WHERE profile_name='Profile 1'"
    ).fetchone()["id"]
    for i, (mail_date, rows) in enumerate(sorted(runs.items()), start=1):
        conn.execute(
            "INSERT INTO ws3_mail_runs (mail_date, mail_id, run_datetime, source_file_name) VALUES (?, ?, ?, 't.xlsx')",
            (mail_date, f"M{i}", f"{mail_date} 12:00:00"),
        )
        run_id = conn.execute(
            "SELECT run_id FROM ws3_mail_runs WHERE mail_date=? AND mail_id=?",
            (mail_date, f"M{i}"),
        ).fetchone()["run_id"]
        for rate_type, claimed, pieces, cost_pp in rows:
            conn.execute(
                """
                INSERT INTO ws3_mail_detail (
                    run_id, profile_id, customer_code, rate_type,
                    postage_claimed, postage_applied, num_pieces, pcs_accepted, pcs_rejected,
                    cost_per_piece, usps_cost_per_piece
                ) VALUES (?, ?, '301079', ?, ?, ?, ?, ?, 0, ?, ?)
                """,
                (run_id, profile_id, rate_type, claimed, claimed, pieces, pieces, cost_pp, cost_pp),
            )


def test_api_profit_flats_dated_tariff_splits_summary_at_rate_case(monkeypatch):
    """Retail per row follows the flats tariff in effect on each mail date."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_tariff.db"
        client = _client(monkeypatch, p)
        conn = dbmod.get_connection()
        try:
            conn.execute(
                "INSERT INTO flat_rate_costs (weight_not_over_oz, rate_retail, effective_date) VALUES (1.0, 1.63, '1900-01-01')"
            )
            conn.execute(
                "INSERT INTO flat_rate_costs (weight_not_over_oz, rate_retail, effective_date) VALUES (1.0, 1.69, '2026-07-12')"
            )
            _seed_ws3_flats(
                conn,
                {
                    "2026-07-10": [("ADC Auto", 123.0, 100, 1.23)],
                    "2026-07-13": [("ADC Auto", 123.0, 100, 1.23)],
                },
            )
            conn.commit()
        finally:
            conn.close()

        r = client.get(
            "/api/profit/flats?start_date=2026-07-01&end_date=2026-07-16&parent_number=100&discount=0.10&discount_efd=0.23"
        )
        assert r.status_code == 200
        j = r.get_json()

        # Meta retail is as-of the report end date (post rate case).
        assert j["meta"]["retail_rate"] == pytest.approx(1.69)
        assert j["meta"]["tariff_effective_date"] == "2026-07-12"

        # Detail rows carry each mail date's tariff.
        by_date = {d["mail_date"]: d for d in j["detail"]}
        assert by_date["2026-07-10"]["retail_rate"] == pytest.approx(1.63)
        assert by_date["2026-07-13"]["retail_rate"] == pytest.approx(1.69)
        assert by_date["2026-07-10"]["price_to_customer"] == pytest.approx(1.53)
        assert by_date["2026-07-13"]["price_to_customer"] == pytest.approx(1.59)

        # Summary splits ADC Auto into one row per tariff retail.
        adc_rows = [x for x in j["rate_summary"] if x["rate_type"] == "ADC Auto"]
        assert sorted(x["retail_rate"] for x in adc_rows) == [
            pytest.approx(1.63),
            pytest.approx(1.69),
        ]
        for x in adc_rows:
            assert x["sell_to_rate"] == pytest.approx(x["retail_rate"] - 0.10)
            assert x["avg_price_to_efd"] == pytest.approx(x["retail_rate"] - 0.23)

        # Combined totals: (1.53-1.23)*100 + (1.59-1.23)*100.
        assert j["totals"]["total_profit"] == pytest.approx(30.0 + 36.0)


def test_api_profit_flats_two_layer_split_and_single_piece(monkeypatch):
    """Supplier profit = (retail-0.23)-cost; EFD profit = 0.13/pc; Single Piece both zero."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_two_layer.db"
        client = _client(monkeypatch, p)
        conn = dbmod.get_connection()
        try:
            _seed_ws3_flats(
                conn,
                {
                    "2026-04-03": [
                        ("ADC Auto", 123.0, 100, 1.23),
                        ("Single Piece", 29.34, 18, 1.63),
                    ]
                },
            )
            conn.commit()
        finally:
            conn.close()

        r = client.get(
            "/api/profit/flats?start_date=2026-04-01&end_date=2026-04-07&parent_number=100&discount=0.10&discount_efd=0.23"
        )
        assert r.status_code == 200
        j = r.get_json()

        adc = next(x for x in j["rate_summary"] if x["rate_type"] == "ADC Auto")
        # Fallback retail 1.63 on an empty tariff table keeps legacy numbers.
        assert adc["supplier_profit_total"] == pytest.approx(round((1.40 - 1.23) * 100, 2))
        assert adc["efd_profit_total"] == pytest.approx(round(0.13 * 100, 2))
        assert adc["total_profit"] == pytest.approx(
            adc["supplier_profit_total"] + adc["efd_profit_total"]
        )

        sp = next(x for x in j["rate_summary"] if x["rate_type"] == "Single Piece")
        assert sp["supplier_profit_total"] == pytest.approx(0.0, abs=0.01)
        assert sp["efd_profit_total"] == pytest.approx(0.0, abs=0.01)

        assert j["totals"]["total_supplier_profit"] == pytest.approx(17.0)
        assert j["totals"]["total_efd_profit"] == pytest.approx(13.0)
        assert j["totals"]["total_profit"] == pytest.approx(30.0)

        det_sp = next(x for x in j["detail"] if x["rate_type"] == "Single Piece")
        assert det_sp["price_to_customer"] == pytest.approx(det_sp["usps_cost_per_piece"])
        assert det_sp["price_to_efd"] == pytest.approx(det_sp["usps_cost_per_piece"])
        assert det_sp["supplier_profit_total"] == pytest.approx(0.0, abs=0.01)
        assert det_sp["efd_profit_total"] == pytest.approx(0.0, abs=0.01)
