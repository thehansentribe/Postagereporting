"""Tests for importer helpers."""

import sqlite3
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

import importer
import watcher


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("0986", 986),
        ("0001", 1),
        ("88", 88),
        ("0", 0),
        ("", 0),
        (123, 123),
    ],
)
def test_strip_zeros(raw, expected):
    assert importer.strip_zeros(raw) == expected


def test_strip_zeros_invalid():
    assert importer.strip_zeros("abc") is None
    assert importer.strip_zeros(None) is None


def _raw_export_row_14(customer: str, code: Any, parent_company: Any = None) -> list[Any]:
    r: list[Any] = [None] * 14
    r[0], r[1], r[13] = customer, code, parent_company
    return r


def test_import_customers_raw_export_resolves_parent_by_name(monkeypatch, tmp_path):
    import db as dbmod

    db_path = tmp_path / "c.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()
    p = tmp_path / "raw.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(_raw_export_row_14("Customer", "Code", "Parent Company"))
    ws.append(_raw_export_row_14("ParentCo", "100", None))
    ws.append(_raw_export_row_14("Child One", "200", "ParentCo"))
    wb.save(p)
    wb.close()

    out = importer.import_customers_raw_export_xlsx(str(p), db_path)
    assert out["rows_imported"] == 2
    assert not out["warnings"]

    conn = sqlite3.connect(str(db_path))
    try:
        row = conn.execute(
            "SELECT customer_name, parent_number, parent_name FROM customers WHERE customer_number = 200"
        ).fetchone()
        assert row[0] == "Child One"
        assert row[1] == 100
        assert row[2] == "ParentCo"
    finally:
        conn.close()


def test_import_customers_raw_export_unresolved_parent_warning(monkeypatch, tmp_path):
    import db as dbmod

    db_path = tmp_path / "c.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()
    p = tmp_path / "raw.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(_raw_export_row_14("Customer", "Code", "Parent Company"))
    ws.append(_raw_export_row_14("Kid", "300", "NoSuchParent"))
    wb.save(p)
    wb.close()

    out = importer.import_customers_raw_export_xlsx(str(p), db_path)
    assert out["rows_imported"] == 1
    assert any("unknown in column A" in w for w in out["warnings"])

    conn = sqlite3.connect(str(db_path))
    try:
        row = conn.execute(
            "SELECT parent_number, parent_name FROM customers WHERE customer_number = 300"
        ).fetchone()
        assert row[0] is None
        assert row[1] == "NoSuchParent"
    finally:
        conn.close()


def test_import_customers_raw_export_duplicate_name_first_code_wins(monkeypatch, tmp_path):
    import db as dbmod

    db_path = tmp_path / "c.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()
    p = tmp_path / "raw.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(_raw_export_row_14("Customer", "Code", "Parent Company"))
    ws.append(_raw_export_row_14("DupName", "10", None))
    ws.append(_raw_export_row_14("DupName", "20", None))
    ws.append(_raw_export_row_14("Kid", "30", "DupName"))
    wb.save(p)
    wb.close()

    out = importer.import_customers_raw_export_xlsx(str(p), db_path)
    assert out["rows_imported"] == 3
    assert any("Duplicate customer name" in w for w in out["warnings"])

    conn = sqlite3.connect(str(db_path))
    try:
        row = conn.execute(
            "SELECT parent_number FROM customers WHERE customer_number = 30"
        ).fetchone()
        assert row[0] == 10
    finally:
        conn.close()


def test_resolve_xlsx_path_ooxml_named_xls_copied_to_xlsx(tmp_path):
    """A real OOXML workbook saved with a .xls name must resolve to a readable .xlsx."""
    src_xlsx = tmp_path / "real.xlsx"
    wb = Workbook()
    wb.active["A1"] = "hello"
    wb.save(src_xlsx)
    wb.close()

    mislabeled = tmp_path / "WS3_FCFL_CustomerMailDetail(5).xls"
    mislabeled.write_bytes(src_xlsx.read_bytes())

    resolved = importer.resolve_xlsx_path(str(mislabeled), out_dir=str(tmp_path))
    assert resolved.lower().endswith(".xlsx")
    assert Path(resolved).is_file()
    # openpyxl rejects .xls extensions, so a readable result proves the copy happened.
    from openpyxl import load_workbook

    wb2 = load_workbook(resolved, read_only=True, data_only=True)
    try:
        assert wb2.active["A1"].value == "hello"
    finally:
        wb2.close()


def test_resolve_xlsx_path_native_xlsx_returned_unchanged(tmp_path):
    src_xlsx = tmp_path / "real.xlsx"
    wb = Workbook()
    wb.save(src_xlsx)
    wb.close()
    assert importer.resolve_xlsx_path(str(src_xlsx)) == str(src_xlsx)


def test_resolve_xlsx_path_ole2_routes_to_conversion(tmp_path, monkeypatch):
    """Genuine legacy BIFF (OLE2 magic) must be sent through LibreOffice conversion."""
    legacy = tmp_path / "old.xls"
    legacy.write_bytes(importer._OLE2_MAGIC + b"\x00" * 16)

    calls: list[str] = []

    def fake_convert(path, out_dir=None):
        calls.append(path)
        return str(tmp_path / "converted.xlsx")

    monkeypatch.setattr(importer, "convert_xls_to_xlsx", fake_convert)
    out = importer.resolve_xlsx_path(str(legacy))
    assert out == str(tmp_path / "converted.xlsx")
    assert calls == [str(legacy)]


def test_resolve_xlsx_path_unknown_xls_falls_back_to_conversion(tmp_path, monkeypatch):
    """Unknown content with a .xls name preserves prior behavior (convert)."""
    weird = tmp_path / "mystery.xls"
    weird.write_bytes(b"not a known magic")

    monkeypatch.setattr(
        importer, "convert_xls_to_xlsx", lambda path, out_dir=None: "sentinel.xlsx"
    )
    assert importer.resolve_xlsx_path(str(weird)) == "sentinel.xlsx"


@pytest.mark.parametrize(
    "filename,expected",
    [
        ("BM_3_20_26_report.csv", "2026-03-20"),
        ("BM_11_5_26.xls", "2026-11-05"),
        ("BM 3.19.26.xls", "2026-03-19"),
        ("BM 3.19.26.csv", "2026-03-19"),
        ("BM_3_20_2026_report.csv", "2026-03-20"),
    ],
)
def test_parse_bm_date(filename, expected):
    assert importer.parse_bm_date(filename) == expected


def test_parse_bm_date_bad():
    with pytest.raises(ValueError, match="Cannot parse"):
        importer.parse_bm_date("not_a_bm_file.csv")


def test_parse_bm_raw_csv_skips_class_header_in_column_g(tmp_path):
    """Repeated sheet headers put the word 'Class' in column G; do not use as mail_class."""
    p = tmp_path / "BM_1_2_26.csv"
    lines = [
        "Pitney Bowes,,,,,,,,,,,,,,,,,,",
        "1234,,,,,,,,,,,,,,,,,,",
        ",,,,,,1CA5DFlt,,,,,,,,,,,,,",
        ",,,,,,Class,,,,,,,,,,,,,",
        ",,,,,,,,,,,,2.0,,,1,,1.0,,,,",
    ]
    p.write_text("\n".join(lines), encoding="utf-8-sig")
    rows = importer.parse_bm_raw_csv(str(p))
    assert len(rows) == 1
    assert rows[0]["mail_class"] == "1CA5DFlt"


def test_parse_bm_raw_csv_synthetic(tmp_path):
    """Minimal raw BM-style CSV: headers, account row, class row, one data row."""
    p = tmp_path / "BM_1_2_26.csv"
    lines = [
        "Pitney Bowes,,,,,,,,,,,,,,,,,,",
        "1234,,,,,,,,,,,,,,,,,,",
        ",,,,,,1CA,,,,,,,,,,,,,",
        ",,,,,,,,,,,,3.0,,,10,,25.50,,,,",
    ]
    p.write_text("\n".join(lines), encoding="utf-8-sig")
    rows = importer.parse_bm_raw_csv(str(p))
    assert len(rows) == 1
    assert rows[0] == {
        "account_code": "1234",
        "mail_class": "1CA",
        "weight_oz": 3.0,
        "pieces": 10,
        "total_cost": 25.5,
    }


def test_parse_bm_raw_csv_fixture_bm_3_19_26():
    root = Path(__file__).resolve().parent.parent
    samples = root / "projectfiles" / "samples"
    fixture = samples / "BM 3.19.26.csv"
    if not fixture.is_file():
        fixture = root / "BM 3.19.26.csv"
    report = samples / "BM_3_19_26_report.csv"
    if not report.is_file():
        report = root / "BM_3_19_26_report.csv"
    if not fixture.is_file():
        pytest.skip("BM 3.19.26.csv not found (checked projectfiles/samples and root)")
    rows = importer.parse_bm_raw_csv(str(fixture))
    assert len(rows) == 3312
    assert rows[0]["account_code"] == "8393"
    assert rows[0]["mail_class"] == "NOCLASS"
    assert rows[0]["weight_oz"] == 0.0
    # Comma-formatted weight / pieces from raw export
    oc = next(r for r in rows if r["mail_class"] == "OtherCls" and r["weight_oz"] == 1120.0)
    assert oc["pieces"] == 0
    # First non-zero cost in file (1ClFlat 2 oz)
    cl = next(r for r in rows if r["mail_class"] == "1ClFlat" and r["weight_oz"] == 2.0)
    assert cl["pieces"] == 1
    assert cl["total_cost"] == 1.9
    if report.is_file():
        import csv

        def norm_cost(x: float) -> str:
            return f"{float(x):.3f}"

        with open(report, encoding="utf-8", newline="") as f:
            expected = [
                (
                    row["Account Code"],
                    row["Class"],
                    row["Weight  (oz.)"],
                    row["Pieces"],
                    norm_cost(row["Total Cost"]),
                )
                for row in csv.DictReader(f)
            ]
        got = [
            (
                r["account_code"],
                r["mail_class"],
                str(int(r["weight_oz"]) if r["weight_oz"] == int(r["weight_oz"]) else r["weight_oz"]),
                str(r["pieces"]),
                norm_cost(r["total_cost"]),
            )
            for r in rows
        ]
        assert got == expected


def test_import_bm_csv_diverts_othercls_1120_to_presort_rejects(monkeypatch, tmp_path):
    import db as dbmod

    # Init an empty DB for import to write into.
    db_path = tmp_path / "imp.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    conn = dbmod.get_connection()
    try:
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (1234, 'Co')"
        )
        conn.commit()
    finally:
        conn.close()

    # Minimal BM report CSV with a single uplift row.
    p = tmp_path / "BM_4_23_26_report.csv"
    p.write_text(
        "\n".join(
            [
                "Account Code,Class,Weight  (oz.),Pieces,Total Cost",
                "1234,OtherCls,1120.0,7,0.66",
            ]
        ),
        encoding="utf-8",
    )

    out = importer.import_bm_csv(str(p), db_path, file_date_override="2026-04-23")
    assert out["rows_imported"] == 0
    assert out["diverted_presort_reject_pieces"] == 7

    conn = dbmod.get_connection()
    try:
        n_postage = conn.execute("SELECT COUNT(*) FROM postage_data").fetchone()[0]
        assert int(n_postage) == 0
        n_rej = conn.execute(
            "SELECT COALESCE(SUM(reject_count),0) FROM postage_presort_rejects"
        ).fetchone()[0]
        assert int(n_rej) == 7
    finally:
        conn.close()


def test_import_bm_csv_rejects_different_file_same_report_date(monkeypatch, tmp_path):
    """A second BM file for the same report date must not double-count."""
    import db as dbmod

    db_path = tmp_path / "imp.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    conn = dbmod.get_connection()
    try:
        conn.execute("INSERT INTO customers (customer_number, customer_name) VALUES (88, 'Co')")
        conn.execute(
            "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('other_report.csv', '2026-05-01', 1)"
        )
        conn.commit()
    finally:
        conn.close()

    p = tmp_path / "BM_5_1_26_report.csv"
    p.write_text(
        "Account Code,Class,Weight  (oz.),Pieces,Total Cost\n88,1CA5DFlt,2.0,1,1.5\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="already imported"):
        importer.import_bm_csv(str(p), db_path, file_date_override="2026-05-01")

    conn = dbmod.get_connection()
    try:
        n = conn.execute(
            "SELECT COUNT(*) FROM postage_imports WHERE file_name = ?", (p.name,)
        ).fetchone()[0]
        assert int(n) == 0
    finally:
        conn.close()


def test_import_bm_csv_replace_same_filename_allowed(monkeypatch, tmp_path):
    """Re-importing the same BM report file replaces the prior import for that filename."""
    import db as dbmod

    db_path = tmp_path / "imp.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    conn = dbmod.get_connection()
    try:
        conn.execute("INSERT INTO customers (customer_number, customer_name) VALUES (88, 'Co')")
        conn.commit()
    finally:
        conn.close()

    p = tmp_path / "BM_5_1_26_report.csv"
    body = "Account Code,Class,Weight  (oz.),Pieces,Total Cost\n88,1CA5DFlt,2.0,1,1.5\n"
    p.write_text(body, encoding="utf-8")
    out1 = importer.import_bm_csv(str(p), db_path, file_date_override="2026-05-01")
    assert out1["rows_imported"] == 1
    out2 = importer.import_bm_csv(str(p), db_path, file_date_override="2026-05-01")
    assert out2["rows_imported"] == 1

    conn = dbmod.get_connection()
    try:
        n = conn.execute(
            "SELECT COUNT(*) FROM postage_imports WHERE file_name = ?", (p.name,)
        ).fetchone()[0]
        assert int(n) == 1
        n_data = conn.execute(
            "SELECT COUNT(*) FROM postage_data WHERE file_date = '2026-05-01'"
        ).fetchone()[0]
        assert int(n_data) == 1
    finally:
        conn.close()


def test_parse_bm_raw_csv_commas_in_numbers(tmp_path):
    """Thousands separators in weight and pieces must parse."""
    p = tmp_path / "BM_1_2_26.csv"
    lines = [
        "Pitney Bowes,,,,,,,,,,,,,,,,,,",
        "1234,,,,,,,,,,,,,,,,,,",
        ",,,,,,ZClass,,,,,,,,,,,,,",
        ',,,,,,,,,,,,"1,120.5",,,"5,786",,12.340,,,,',
    ]
    p.write_text("\n".join(lines), encoding="utf-8-sig")
    rows = importer.parse_bm_raw_csv(str(p))
    assert len(rows) == 1
    assert rows[0]["weight_oz"] == 1120.5
    assert rows[0]["pieces"] == 5786
    assert rows[0]["total_cost"] == 12.34


def test_watcher_bm_report_vs_raw_detection():
    assert watcher._is_bm_report_csv("BM 3.19.26_report.csv")
    assert watcher._is_bm_report_csv("BM_3_19_26_report.csv")
    assert not watcher._is_bm_raw_export("BM 3.19.26_report.csv")
    assert watcher._is_bm_raw_export("BM 3.19.26.csv")
    assert watcher._is_bm_raw_export(
        "DM Weight Break by Account-Carrier-Class 04242026 - 053056.7513322.xlsx"
    )


def test_write_report_csv_renames_csv_source(tmp_path):
    src = tmp_path / "BM_1_2_26.csv"
    src.write_text("", encoding="utf-8")
    out = importer.write_report_csv(
        [
            {
                "account_code": "1",
                "mail_class": "X",
                "weight_oz": 1.0,
                "pieces": 1,
                "total_cost": 0.1,
            }
        ],
        str(src),
        str(tmp_path),
    )
    assert Path(out).name == "BM_1_2_26_report.csv"


def test_read_bm_report_date_from_xlsx_happy_path(tmp_path):
    p = tmp_path / "DM Weight Break by Account-Carrier-Class 04242026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["P3"].value = "04/24/2026"
    ws["S3"].value = "04/24/2026"
    wb.save(p)

    assert importer.read_bm_report_date_from_xlsx(str(p)) == "2026-04-24"


def test_read_bm_report_date_from_xlsx_mismatch_raises(tmp_path):
    p = tmp_path / "DM Weight Break by Account-Carrier-Class 04242026.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["P3"].value = "04/24/2026"
    ws["S3"].value = "04/25/2026"
    wb.save(p)

    with pytest.raises(ValueError, match=r"BM report date mismatch"):
        importer.read_bm_report_date_from_xlsx(str(p))


@pytest.mark.parametrize(
    "filename,expected",
    [
        ("Priority mail zones 4-27-26.xlsx", "2026-04-27"),
        ("zones 12-1-25.xlsx", "2025-12-01"),
    ],
)
def test_parse_priority_mail_filename_effective_date(filename, expected):
    assert importer.parse_priority_mail_filename_effective_date(filename) == expected


def test_parse_priority_mail_filename_effective_date_missing():
    assert importer.parse_priority_mail_filename_effective_date("zones.xlsx") is None


def test_parse_compact_priority_mail_zone_matrix_alt_zones_only_header_row():
    """Layout matching 'Priority mail zones 4-27-26.xlsx': zones 1–9 in row 1 A–I, lb in column A."""
    raw = [
        [1, 2, 3, 4, 5, 6, 7, 8, 9],
        [
            1,
            11,
            11.5,
            11.65,
            11.9,
            13.05,
            14.5,
            15.6,
            16.95,
            34.25,
        ],
        [2, 11.95, 12.65, 13.45, 14.35, 16.85, 18.15, 19.45, 21.55, 42.4],
        [3, 12.45, 13.3, 14, 15.15, 18.05, 20.95, 23.85, 27.1, 53.45],
    ]
    rows = importer._parse_compact_priority_mail_zone_matrix_xlsx(
        raw, "zones.xlsx", effective_date="2026-04-27"
    )
    assert rows is not None and len(rows) == 27
    w1_z1 = next(
        r
        for r in rows
        if r["weight_max"] == 1.0 and r["zone"] == 1 and r["effective_date"] == "2026-04-27"
    )
    assert w1_z1["price"] == 11
    assert w1_z1["row_type"] == "matrix"


def test_parse_compact_priority_mail_zone_matrix_alt_header_zones_permutes_columns():
    """Headers may list zones out of USPS order as long as 1–9 appear once."""
    raw = [[9, 8, 7, 6, 5, 4, 3, 2, 1], [1.0] + list(range(10, 19))]
    rows = importer._parse_compact_priority_mail_zone_matrix_xlsx(raw, "p.xlsx", effective_date=None)
    assert rows is not None and len(rows) == 9
    by_zone = {r["zone"]: r["price"] for r in rows}
    assert by_zone == {z: float(19 - z) for z in range(1, 10)}


def test_import_priority_mail_retail_replaces_same_effective_slice_only(monkeypatch, tmp_path):
    import sqlite3

    import db as db_actual

    db_path = tmp_path / "prio.db"
    monkeypatch.setattr(db_actual, "DB_PATH", db_path)
    db_actual.init_db()
    conn = db_actual.get_connection()
    conn.executemany(
        """
        INSERT INTO priority_mail_retail (
            effective_date, row_type, zone, weight_unit, weight_max, price, sort_group, sort_order
        ) VALUES (?, 'matrix', ?, 'lb', ?, ?, 1, ?)
        """,
        [
            ("2020-01-01", 3, 1.0, 1.0, 0),
            ("2026-06-01", 4, 1.0, 2.0, 0),
        ],
    )
    conn.commit()
    conn.close()

    xlsx = tmp_path / "mat.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["(lbs.)", 1, 2, 3, 4, 5, 6, 7, 8, 9])
    ws.append([1, 11, 11, 11, 11, 11, 11, 11, 11, 11])
    wb.save(xlsx)

    out = importer.import_priority_mail_retail(str(xlsx), db_path, effective_date="2026-06-01")
    assert out["effective_date"] == "2026-06-01"
    assert out["rows_imported"] == 9

    conn = sqlite3.connect(str(db_path))
    try:
        rows = conn.execute(
            "SELECT zone, price, effective_date FROM priority_mail_retail ORDER BY effective_date, zone"
        ).fetchall()
        by_eff: dict[str, list[tuple]] = {}
        for z, price, eff in rows:
            by_eff.setdefault(eff, []).append((z, price))
        assert "2020-01-01" in by_eff
        assert any(z == 3 and price == 1.0 for z, price in by_eff["2020-01-01"])
        june = by_eff.get("2026-06-01", [])
        assert len(june) == 9
        assert all(z in range(1, 10) and price == 11.0 for z, price in june)
    finally:
        conn.close()


FCM_RETAIL_SAMPLE = """First-Class Mail and EDDM - Retail,,,,Final,,6/15/2026,,
,,,,,,
First-Class Mail,,,,,,
LETTERS,,FLATS,,,,
Weight Not Over (ounces),,Weight Not Over (ounces),,,,
1,0.82,1,1.69,,,
2,1.11,2,1.98,,,
"""

FCM_COMM_FLATS_SAMPLE = """First-Class Mail - Commercial - Flats,,,,,Final,,6/15/2026,,
Weight Not Over (ounces),5-Digit,3-Digit,Mixed,Presorted,
1,1.025,1.264,1.585,1.590,
2,1.315,1.554,1.875,1.880,
"""

PM_RETAIL_SAMPLE = """Priority Mail - Retail,,Final,,6/15/2026
Flat Rate Envelopes,12.90
Weight Not Over (Lbs),Zones
,Zone 1,Zone 2
1,11.00,11.50
2,11.95,12.65
"""


def test_parse_notice123_fcm_retail_flats(tmp_path):
    p = tmp_path / "FCM & EDDM - Retail.csv"
    p.write_text(FCM_RETAIL_SAMPLE, encoding="utf-8")
    out = importer.parse_notice123_fcm_retail_flats(str(p))
    assert out[1.0] == 1.69
    assert out[2.0] == 1.98


def test_parse_notice123_fcm_comm_flats_presort(tmp_path):
    p = tmp_path / "FCM - Comm Flats.csv"
    p.write_text(FCM_COMM_FLATS_SAMPLE, encoding="utf-8")
    rows = importer.parse_notice123_fcm_comm_flats_presort(str(p))
    assert len(rows) == 2
    assert rows[0]["rate_5digit"] == 1.025
    assert rows[0]["rate_machinable_pres"] == 1.59


def test_import_notice123_flat_rates_and_future_effective_date(monkeypatch, tmp_path):
    import sqlite3

    import db as db_actual

    db_path = tmp_path / "flats.db"
    monkeypatch.setattr(db_actual, "DB_PATH", db_path)
    db_actual.init_db()

    retail = tmp_path / "retail.csv"
    presort = tmp_path / "presort.csv"
    retail.write_text(FCM_RETAIL_SAMPLE, encoding="utf-8")
    presort.write_text(FCM_COMM_FLATS_SAMPLE, encoding="utf-8")

    importer.import_notice123_flat_rates(
        str(retail), str(presort), db_path, effective_date="2026-07-01"
    )

    conn = db_actual.get_connection()
    try:
        before = db_actual.get_flat_rate_costs(conn, as_of_date="2026-06-15")
        assert before["rows"] == []
        active = db_actual.get_flat_rate_costs(conn, as_of_date="2026-07-15")
        assert len(active["rows"]) == 2
        assert active["tariff_effective_date"] == "2026-07-01"
        assert active["rows"][0]["rate_retail"] == 1.69
        assert active["rows"][0]["rate_5digit"] == 1.025
    finally:
        conn.close()


def test_import_notice123_rate_case_zip(monkeypatch, tmp_path):
    import sqlite3
    import zipfile

    import db as db_actual

    db_path = tmp_path / "notice.db"
    extract_dir = tmp_path / "Notice123"
    monkeypatch.setattr(db_actual, "DB_PATH", db_path)
    monkeypatch.setattr(importer, "NOTICE123_DIR", extract_dir)
    db_actual.init_db()

    inner = "July 2026 Price Change - Notice 123"
    folder = tmp_path / "zip_src" / inner
    folder.mkdir(parents=True)
    (folder / "PM Retail.csv").write_text(PM_RETAIL_SAMPLE, encoding="utf-8")
    (folder / "FCM & EDDM - Retail.csv").write_text(FCM_RETAIL_SAMPLE, encoding="utf-8")
    (folder / "FCM - Comm Flats.csv").write_text(FCM_COMM_FLATS_SAMPLE, encoding="utf-8")

    zip_path = tmp_path / "notice.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        for f in folder.rglob("*"):
            if f.is_file():
                zf.write(f, f.relative_to(tmp_path / "zip_src"))

    out = importer.import_notice123_rate_case(
        zip_path, db_path, effective_date="2026-07-01", dest_dir=extract_dir
    )
    assert out["effective_date"] == "2026-07-01"
    assert out["priority_mail"]["rows_imported"] > 0
    assert out["flats"]["rows_imported"] == 2

    conn = sqlite3.connect(str(db_path))
    try:
        pm_count = conn.execute(
            "SELECT COUNT(*) FROM priority_mail_retail WHERE effective_date = ?",
            ("2026-07-01",),
        ).fetchone()[0]
        flat_count = conn.execute(
            "SELECT COUNT(*) FROM flat_rate_costs WHERE effective_date = ?",
            ("2026-07-01",),
        ).fetchone()[0]
        assert pm_count > 0
        assert flat_count == 2
    finally:
        conn.close()


def test_find_notice123_csv_prefers_nested_notice_folder(tmp_path):
    root = tmp_path / "tree"
    root.mkdir()
    (root / "PM Retail.csv").write_text("stale", encoding="utf-8")
    nested = root / "July 2026 Price Change - Notice 123"
    nested.mkdir()
    (nested / "PM Retail.csv").write_text("nested", encoding="utf-8")

    chosen, count = importer._find_notice123_csv(root, "PM Retail.csv")
    assert count == 2
    assert chosen == nested / "PM Retail.csv"


def test_import_notice123_rate_case_replaces_stale_root_csvs(monkeypatch, tmp_path):
    import sqlite3
    import zipfile

    import db as db_actual

    db_path = tmp_path / "notice.db"
    notice_dir = tmp_path / "Notice123"
    notice_dir.mkdir()
    (notice_dir / "PM Retail.csv").write_text("stale-root", encoding="utf-8")

    monkeypatch.setattr(db_actual, "DB_PATH", db_path)
    monkeypatch.setattr(importer, "NOTICE123_DIR", notice_dir)
    db_actual.init_db()

    inner = "July 2026 Price Change - Notice 123"
    folder = tmp_path / "zip_src" / inner
    folder.mkdir(parents=True)
    (folder / "PM Retail.csv").write_text(PM_RETAIL_SAMPLE, encoding="utf-8")
    (folder / "FCM & EDDM - Retail.csv").write_text(FCM_RETAIL_SAMPLE, encoding="utf-8")
    (folder / "FCM - Comm Flats.csv").write_text(FCM_COMM_FLATS_SAMPLE, encoding="utf-8")

    zip_path = tmp_path / "notice.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        for f in folder.rglob("*"):
            if f.is_file():
                zf.write(f, f.relative_to(tmp_path / "zip_src"))

    out = importer.import_notice123_rate_case(
        zip_path, db_path, effective_date="2026-07-01", dest_dir=notice_dir
    )
    assert out["priority_mail"]["rows_imported"] > 0
    assert out["flats"]["rows_imported"] == 2
    assert not (notice_dir / "PM Retail.csv").exists()
    assert (notice_dir / inner / "PM Retail.csv").exists()

    conn = sqlite3.connect(str(db_path))
    try:
        pm_count = conn.execute(
            "SELECT COUNT(*) FROM priority_mail_retail WHERE effective_date = ?",
            ("2026-07-01",),
        ).fetchone()[0]
        assert pm_count > 0
    finally:
        conn.close()
