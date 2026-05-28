"""Tests for profit report Excel export (example layout + EFD parcel sheet)."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

import db as dbmod
import exports

pytest.importorskip("flask")


def test_export_profit_report_xlsx_summary_formulas_and_parcel_efd_sheet(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_export.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()

        conn = dbmod.get_connection()
        try:
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) "
                "VALUES (100, 'Parent Co', NULL, NULL)"
            )
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) "
                "VALUES (101, 'Child Co', 100, 'Parent Co')"
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('bx', 'b.csv', 1)"
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class,
                    final_postage, fully_paid_postage, billing_amount,
                    zone
                ) VALUES (1, 101, 'Child Co', '4/15/2026 10:00', 16.0, 'CLS', 2.00, 3.00, 5.00, '1')
                """
            )
            conn.commit()
        finally:
            conn.close()

        out = exports.export_profit_report_xlsx(
            "2026-04-01",
            "2026-04-30",
            parent_number=100,
            customer_number=None,
            show_parents=True,
            show_main=True,
            flats_discount=0.10,
            flats_discount_efd=0.23,
            efd_parcel_fee=0.5,
        )
        assert out.is_file()
        wb = load_workbook(out, data_only=False)

        assert wb.sheetnames == ["Summary", "Flats Detail", "Parcel Profit"]

        summary = wb["Summary"]
        assert summary["A1"].value == "Profit Report"
        assert summary["B6"].value == 1.63
        assert summary["B7"].value == pytest.approx(0.1, rel=1e-9)
        assert summary["B8"].value == pytest.approx(0.23, rel=1e-9)
        assert summary["B9"].value == "=B6-B7"
        assert summary["A10"].value == "Parcel fee to EFD ($/pc)"
        assert summary["B10"].value == pytest.approx(0.5, rel=1e-9)

        pp = wb["Parcel Profit"]
        assert pp.cell(exports._EFD_PARCEL_INVOICE_HEADER_ROW, 24).value == "billing_amount"
        assert pp.cell(10, 25).value == "=X10+0.5"
        assert pp.cell(5, 2).value == "=B4-B3"

        r8 = None
        for r in range(1, summary.max_row + 1):
            if summary.cell(r, 2).value == "EFD Profit":
                r8 = r
                break
        assert r8 is not None
        assert "Parcel Profit" in str(summary.cell(r8, 3).value)
        assert "B5" in str(summary.cell(r8, 3).value)

        wb.close()
        out.unlink(missing_ok=True)


def test_parcel_profit_from_raw_unit():
    raw = {
        "parcel_count": 2,
        "total_final_postage": 2.0,
        "total_fully_paid_postage": 3.0,
        "total_billing_amount": 7.0,
    }
    out = dbmod.parcel_profit_from_raw(raw)
    assert out["computed"]["postage_fee"] == pytest.approx(2.5, rel=1e-9)
    assert out["computed"]["lineage_revenue"] == pytest.approx(7.5, rel=1e-9)
    assert out["computed"]["efd_profit"] == pytest.approx(-6.5, rel=1e-9)
    assert len(out["lines"]) == 7


def test_parcel_profit_from_raw_custom_fee_per_piece():
    raw = {
        "parcel_count": 3,
        "total_final_postage": 1.0,
        "total_fully_paid_postage": 1.0,
        "total_billing_amount": 10.0,
    }
    out = dbmod.parcel_profit_from_raw(raw, parcel_fee_per_piece=2.0)
    assert out["computed"]["postage_fee"] == pytest.approx(6.0, rel=1e-9)
    line6 = next(x for x in out["lines"] if x["line_no"] == 6)
    assert "2.00" in line6["label"]


def test_export_profit_report_custom_efd_parcel_fee_b10(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "profit_export_fee.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()

        out = exports.export_profit_report_xlsx(
            "2026-04-01",
            "2026-04-30",
            parent_number=None,
            customer_number=None,
            show_parents=True,
            show_main=True,
            flats_discount=0.10,
            flats_discount_efd=0.23,
            efd_parcel_fee=1.33,
        )
        wb = load_workbook(out, data_only=False)
        summary = wb["Summary"]
        assert summary["A10"].value == "Parcel fee to EFD ($/pc)"
        assert summary["B10"].value == pytest.approx(1.33, rel=1e-9)
        parcel_title_row = None
        for r in range(1, summary.max_row + 1):
            if summary.cell(r, 1).value == "Parcel Profit":
                parcel_title_row = r
                break
        assert parcel_title_row is not None
        ph = parcel_title_row + 1
        r6 = ph + 4
        assert summary.cell(r6, 3).value == f"=$B$10*C{ph + 1}"
        wb.close()
        out.unlink(missing_ok=True)
