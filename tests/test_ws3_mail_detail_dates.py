"""WS3 mail date from BI8 (Mail ID) per Helper Files/fix_netsort_date.md."""

from __future__ import annotations

import tempfile
import uuid
from pathlib import Path

import pytest
from openpyxl import Workbook

import db as dbmod
import ws3_mail_detail


def test_parse_mail_id_date_ok():
    assert ws3_mail_detail.parse_mail_id_date("040626_F") == "2026-04-06"
    assert ws3_mail_detail.parse_mail_id_date("041726_F") == "2026-04-17"


@pytest.mark.parametrize(
    "raw",
    [
        "",
        "   ",
        "bad",
        "12345_F",
        "1234567_F",
        "99-99-99_F",
    ],
)
def test_parse_mail_id_date_invalid(raw):
    assert ws3_mail_detail.parse_mail_id_date(raw) is None


def test_read_bi8_round_trip(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws["BI8"] = "040626_F"
    p = tmp_path / "mini.xlsx"
    wb.save(p)
    wb.close()

    assert ws3_mail_detail.read_bi8_mail_id(str(p)) == "040626_F"
    assert ws3_mail_detail.read_mail_date_from_bi8(str(p)) == "2026-04-06"


def test_process_ws3_uses_bi8_mail_date(monkeypatch, tmp_path: Path):
    """Minimal workbook: only BI8 set; parse_ws3_xlsx yields empty detail (no crash)."""
    wb = Workbook()
    ws = wb.active
    ws["BI8"] = "040626_F"
    xlsx = tmp_path / "ws3_min.xlsx"
    wb.save(xlsx)
    wb.close()

    with tempfile.TemporaryDirectory() as td:
        dbp = Path(td) / "t.db"
        monkeypatch.setattr(dbmod, "DB_PATH", dbp)
        dbmod.init_db()

        fname = f"WS3_FCFL_CustomerMailDetail_test_{uuid.uuid4().hex}.xlsx"
        dest = Path(td) / fname
        dest.write_bytes(xlsx.read_bytes())

        out = ws3_mail_detail.process_ws3_mail_detail_file(str(dest), dbp)
        assert out.get("skipped") is False

        conn = dbmod.get_connection()
        try:
            row = conn.execute(
                "SELECT mail_date, mail_id FROM ws3_mail_runs ORDER BY run_id DESC LIMIT 1"
            ).fetchone()
            assert row is not None
            assert row["mail_date"] == "2026-04-06"
        finally:
            conn.close()


def test_integration_sample_file_if_present(monkeypatch, tmp_path: Path):
    """
    If the sample XLS from the plan exists in the workspace, import it once into a temp DB.

    Skips when the file is not checked in (common in gitignored Helper Files).
    """
    sample = (
        Path(__file__).resolve().parent.parent
        / "Helper Files"
        / "WS3_FCFL_CustomerMailDetail 4-6-26.xls"
    )
    if not sample.is_file():
        pytest.skip(f"Sample file not present: {sample}")

    with tempfile.TemporaryDirectory() as td:
        dbp = Path(td) / "t.db"
        monkeypatch.setattr(dbmod, "DB_PATH", dbp)
        dbmod.init_db()

        fname = f"WS3_FCFL_CustomerMailDetail_4-6-26_copy_{uuid.uuid4().hex}.xls"
        dest = Path(td) / fname
        dest.write_bytes(sample.read_bytes())

        out = ws3_mail_detail.process_ws3_mail_detail_file(str(dest), dbp)
        assert out.get("skipped") is False
        assert out.get("mail_date") == "2026-04-06"

        conn = dbmod.get_connection()
        try:
            row = conn.execute(
                "SELECT mail_date FROM ws3_mail_runs WHERE mail_date = ?",
                ("2026-04-06",),
            ).fetchone()
            assert row is not None
        finally:
            conn.close()
