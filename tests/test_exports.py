"""Tests for export helpers (parcel roll-up, etc.)."""

from __future__ import annotations

import csv
import tempfile
from pathlib import Path

import pytest

import db
import exports
import exports_consolidated_volumes
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _parcel_row(
    *,
    date: str = "2026-04-03",
    parent_name: str = "P",
    parent_number: int = 1,
    child_name: str = "C",
    child_number: int = 10,
    lb_1: int = 0,
    lb_2: int = 0,
    total_qty: int = 0,
    total_billed: float = 0.0,
    total_retail: float = 0.0,
) -> dict:
    return {
        "date": date,
        "parent_name": parent_name,
        "parent_number": parent_number,
        "child_name": child_name,
        "child_number": child_number,
        "lb_1": lb_1,
        "lb_2": lb_2,
        "lb_3": 0,
        "lb_4": 0,
        "lb_5": 0,
        "lb_6": 0,
        "lb_7": 0,
        "lb_8": 0,
        "lb_9": 0,
        "lb_10": 0,
        "lb_10plus": 0,
        "total_qty": total_qty,
        "total_billed": total_billed,
        "total_retail": total_retail,
    }


def test_aggregate_parcel_count_rows_sums_costs_across_split_rows() -> None:
    """API returns one row per mail class × zone; roll-up must sum billed and retail."""
    rows = [
        _parcel_row(lb_1=1, total_qty=1, total_billed=5.0, total_retail=10.0),
        _parcel_row(lb_2=1, total_qty=1, total_billed=3.0, total_retail=7.0),
    ]
    agg = exports.aggregate_parcel_count_rows(rows)
    assert len(agg) == 1
    a = agg[0]
    assert a["lb_1"] == 1
    assert a["lb_2"] == 1
    assert a["total_qty"] == 2
    assert abs(a["total_billed"] - 8.0) < 1e-9
    assert abs(a["total_retail"] - 17.0) < 1e-9


def test_parcel_counts_download_name() -> None:
    assert "Parcel_Counts" in exports.parcel_counts_download_name("a", "b", None, None)
    assert exports.parcel_counts_download_name("2026-01-01", "2026-01-31", 12, None).endswith(
        ".xlsx"
    )


def test_export_parcel_counts_report_xlsx_child_number_column(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "parcel_counts.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()

        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Acme')"
        )
        conn.execute(
            "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'x.csv', 1)"
        )
        import_id = conn.execute(
            "SELECT id FROM billing_imports WHERE billing_id = 'b1'"
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO billing_records (
                billing_import_id, custom_account_code, time_stamp, weight_oz,
                usps_mail_class, billing_amount, fully_paid_postage, zone
            )
            VALUES (?, 200, '4/3/2026 10:00', 16.0, 'PRIORITY', 5.0, 6.0, '2')
            """,
            (import_id,),
        )
        conn.commit()
        conn.close()

        monkeypatch.setattr(
            db,
            "compute_retail_cost_for_piece",
            lambda *a, **k: {"retail": 6.0, "zone": 2},
        )

        out_hidden = exports.export_parcel_counts_report_xlsx(
            "2026-04-01",
            "2026-04-30",
            None,
            None,
            show_parents=True,
            show_main=True,
            consolidate=False,
            remove_zeros=False,
            hide_costs=False,
            hide_customer_numbers=True,
        )
        wb_h = load_workbook(out_hidden)
        ws_h = wb_h.active
        assert ws_h.cell(1, 4).value == "1 lb"
        assert "Child Number" not in [ws_h.cell(1, c).value for c in range(1, 6)]
        out_hidden.unlink(missing_ok=True)

        out_show = exports.export_parcel_counts_report_xlsx(
            "2026-04-01",
            "2026-04-30",
            None,
            None,
            show_parents=True,
            show_main=True,
            consolidate=False,
            remove_zeros=False,
            hide_costs=False,
            hide_customer_numbers=False,
        )
        wb_s = load_workbook(out_show)
        ws_s = wb_s.active
        assert ws_s.cell(1, 4).value == "Child Number"
        assert ws_s.cell(1, 5).value == "1 lb"
        assert ws_s.cell(2, 4).value == 200
        out_show.unlink(missing_ok=True)


def test_parcel_invoice_download_name_uses_title_customer_and_end_date() -> None:
    n = exports.parcel_invoice_download_name(
        title_name="Security Benefit_Zinnia",
        parent_number=3900,
        end_date="2026-04-10",
    )
    assert n == "Security Benefit_Zinnia -EFD (3900) Parcel invoice 4-10-2026.xlsx"


def test_consolidated_volumes_download_name() -> None:
    n = exports_consolidated_volumes.consolidated_volumes_download_name("KC Presort (3906)", "2026-04-07")
    assert n.startswith("KC Presort (3906) 4-7-2026")
    assert n.endswith(".xlsx")


def test_allocate_integer_proportional_exact_split() -> None:
    assert exports.allocate_integer_proportional(10, [1.0, 2.0, 3.0, 4.0]) == [1, 2, 3, 4]
    assert sum(exports.allocate_integer_proportional(10, [1.0, 2.0, 3.0, 4.0])) == 10


def test_allocate_integer_proportional_largest_remainder() -> None:
    out = exports.allocate_integer_proportional(7, [1.0, 1.0, 1.0])
    assert sum(out) == 7
    assert max(out) - min(out) <= 1


def test_allocate_integer_proportional_zero_total() -> None:
    assert exports.allocate_integer_proportional(0, [3.0, 5.0]) == [0, 0]


def test_allocate_integer_proportional_zero_weights_splits_evenly() -> None:
    out = exports.allocate_integer_proportional(5, [0.0, 0.0])
    assert out == [3, 2]
    assert sum(out) == 5


def test_export_flats_data_grid_xlsx_headers_and_piece_count(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "flats_grid.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()
        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Acme')"
        )
        conn.execute(
            "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('x.csv', '2026-03-01', 1)"
        )
        conn.execute(
            """
            INSERT INTO postage_data (
                import_id, file_date, account_code, mail_class,
                weight_oz, pieces, total_cost, unmatched_account
            ) VALUES (1, '2026-03-01', 200, '1ClFlat', 2.0, 5, 10.0, 0)
            """
        )
        conn.commit()
        conn.close()

        out = exports.export_flats_data_grid_xlsx(
            "2026-03-01",
            "2026-03-01",
            hide_costs=False,
            sort_key="date",
            sort_dir=1,
        )
        assert out.is_file()
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "Flats"
        assert ws.cell(1, 1).value == "Date"
        assert ws.cell(1, 4).value == "Class"
        assert ws.cell(1, 20).value == "Total Qty"
        assert ws.cell(1, 21).value == "Total Cost"
        # Col 7 = oz_2 (date, parent, child, class, oz_0..oz_1, oz_2)
        assert ws.cell(2, 7).value == 5
        assert ws.cell(2, 20).value == 5
        assert ws.cell(2, 21).value == 10.0
        out.unlink(missing_ok=True)

        out_show = exports.export_flats_data_grid_xlsx(
            "2026-03-01",
            "2026-03-01",
            hide_costs=False,
            hide_customer_numbers=False,
            sort_key="date",
            sort_dir=1,
        )
        wb2 = load_workbook(out_show)
        ws2 = wb2.active
        assert ws2.cell(1, 4).value == "Child Number"
        assert ws2.cell(1, 5).value == "Class"
        assert ws2.cell(2, 4).value == 200
        assert ws2.cell(2, 8).value == 5
        out_show.unlink(missing_ok=True)


def test_export_flats_data_grid_csv_headers_and_hide_costs(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "flats_grid_csv.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()
        conn = db.get_connection()
        conn.execute("INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Acme')")
        conn.execute(
            "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('x.csv', '2026-03-01', 1)"
        )
        conn.execute(
            """
            INSERT INTO postage_data (
                import_id, file_date, account_code, mail_class,
                weight_oz, pieces, total_cost, unmatched_account
            ) VALUES (1, '2026-03-01', 200, '1ClFlat', 2.0, 5, 10.0, 0)
            """
        )
        conn.commit()
        conn.close()

        out = exports.export_flats_data_grid_csv(
            "2026-03-01",
            "2026-03-01",
            hide_costs=False,
            sort_key="date",
            sort_dir=1,
        )
        assert out.is_file()
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            row = next(reader)
        assert header[0:4] == ["Date", "Parent Name", "Child Name", "Class"]
        assert "Child Number" not in header
        assert "Total Qty" in header
        assert header[-1] == "Total Cost"
        assert len(row) == len(header)
        out.unlink(missing_ok=True)

        out_show = exports.export_flats_data_grid_csv(
            "2026-03-01",
            "2026-03-01",
            hide_costs=False,
            hide_customer_numbers=False,
            sort_key="date",
            sort_dir=1,
        )
        with open(out_show, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header_show = next(reader)
            row_show = next(reader)
        assert header_show[0:5] == [
            "Date",
            "Parent Name",
            "Child Name",
            "Child Number",
            "Class",
        ]
        assert row_show[3] == "200"
        assert len(row_show) == len(header_show)
        out_show.unlink(missing_ok=True)

        out2 = exports.export_flats_data_grid_csv(
            "2026-03-01",
            "2026-03-01",
            hide_costs=True,
            sort_key="date",
            sort_dir=1,
        )
        with open(out2, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header2 = next(reader)
            row2 = next(reader)
        assert header2[0:4] == ["Date", "Parent Name", "Child Name", "Class"]
        assert "Total Qty" in header2
        assert "Total Cost" not in header2
        assert len(row2) == len(header2)
        out2.unlink(missing_ok=True)


def test_total_flats_retail_invoice_basis() -> None:
    rates = {2: 1.9, 3: 2.17}
    row_fl = {"mail_class": "1CA5DFlt", "oz_2": 10, "oz_3": 2}
    row_pr = {"mail_class": db.WS3_REJECT_MAIL_CLASS, "oz_2": 99}
    row_parcel = {"mail_class": "Priority", "oz_2": 5}
    total = exports_consolidated_volumes.total_flats_retail_invoice_basis(
        [row_fl, row_pr, row_parcel], rates
    )
    assert total == pytest.approx(10 * 1.9 + 2 * 2.17, rel=1e-9)


def test_total_flats_retail_invoice_basis() -> None:
    rates = {2: 1.9, 3: 2.17}
    row_fl = {"mail_class": "1CA5DFlt", "oz_2": 10, "oz_3": 2}
    row_pr = {"mail_class": db.WS3_REJECT_MAIL_CLASS, "oz_2": 99}
    row_parcel = {"mail_class": "Priority", "oz_2": 5}
    total = exports_consolidated_volumes.total_flats_retail_invoice_basis(
        [row_fl, row_pr, row_parcel], rates
    )
    assert total == pytest.approx(10 * 1.9 + 2 * 2.17, rel=1e-9)


def test_total_presort_reject_pieces_from_postage_rows() -> None:
    rows = [
        {"mail_class": "1ClFlat", "total_qty": 100},
        {"mail_class": db.WS3_REJECT_MAIL_CLASS, "total_qty": 5},
        {"mail_class": db.WS3_REJECT_MAIL_CLASS, "total_qty": 3},
    ]
    assert exports_consolidated_volumes.total_presort_reject_pieces_from_postage_rows(rows) == 8
    assert exports_consolidated_volumes.total_presort_reject_pieces_from_postage_rows([]) == 0


def test_flats_summary_retail_formula_excludes_reject_classes() -> None:
    """Summary retail uses SUM minus SUMIF for Presort rejects and Rejects; Class column from headers."""
    sheet = "'FLATS (POSTAGE)'"
    for remove_nums in (True, False):
        f_headers, _ = exports_consolidated_volumes._flats_consolidated_column_plan(remove_nums)
        class_col = f_headers.index("Class") + 1

        assert get_column_letter(class_col) == ("D" if remove_nums else "F")

        retail_col = get_column_letter(len(f_headers))
        last_row = 5
        f = exports_consolidated_volumes.flats_summary_retail_formula(
            flats_sheet_quoted=sheet,
            f_headers=f_headers,
            retail_col_letter=retail_col,
            last_data_row=last_row,
        )
        assert f.startswith("=SUM(")
        assert '"Presort rejects"' in f
        assert '"Rejects"' in f
        cc = get_column_letter(class_col)
        assert f"{sheet}!{cc}2:{cc}{last_row}" in f
        assert f"{sheet}!{retail_col}2:{retail_col}{last_row}" in f
        assert f.count("SUMIF(") == 2


def test_export_parcel_billing_csv_writes_full_header(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "billing.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()

        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Acme')"
        )
        conn.execute(
            "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'x.csv', 1)"
        )
        import_id = conn.execute(
            "SELECT id FROM billing_imports WHERE billing_id = 'b1'"
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO billing_records (billing_import_id, custom_account_code, time_stamp, weight_oz)
            VALUES (?, ?, ?, ?)
            """,
            (import_id, 200, "4/1/2026 15:34", 16.0),
        )
        conn.commit()
        conn.close()

        out = exports.export_parcel_billing_csv(
            "2026-04-01",
            "2026-04-01",
            200,
            None,
            show_parents=True,
            show_main=True,
        )
        assert out.is_file()
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            row = next(reader)
        assert header[0:2] == ["parent_name", "parent_number"]
        assert "custom_account_code" in header
        assert "time_stamp" in header
        assert "weight_oz" in header
        assert "retail_cost" in header
        assert header.index("retail_cost") == header.index("billing_amount") + 1
        assert len(row) == len(header)
        out.unlink(missing_ok=True)


def test_efd_parcel_invoice_download_name() -> None:
    n = exports.efd_parcel_invoice_download_name(
        "Blue Cross -EFD",
        parent_number=3901,
        customer_number=None,
        end_date="2026-04-13",
    )
    assert n == "Blue Cross -EFD (3901) 4-13-2026.xlsx"
    n2 = exports.efd_parcel_invoice_download_name(
        "Child Co",
        parent_number=None,
        customer_number=2256,
        end_date="2026-04-01",
    )
    assert "(2256)" in n2 and n2.endswith(".xlsx")
    n3 = exports.efd_parcel_invoice_download_name(
        "All Accounts",
        parent_number=None,
        customer_number=None,
        end_date="2026-01-31",
    )
    assert "(ALL)" in n3


def test_export_efd_parcel_invoice_xlsx_layout_and_formulas(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_invoice.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()

        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Acme')"
        )
        conn.execute(
            "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'x.csv', 1)"
        )
        import_id = conn.execute(
            "SELECT id FROM billing_imports WHERE billing_id = 'b1'"
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO billing_records (
                billing_import_id, custom_account_code, time_stamp, weight_oz,
                billing_amount, zone, imb_tracking_code, impb
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (import_id, 200, "4/1/2026 15:34", 16.0, 5.22, "5", "trk", "impb1"),
        )
        conn.commit()
        conn.close()

        monkeypatch.setattr(
            db,
            "compute_retail_cost_for_piece",
            lambda *a, **k: {"retail": 10.2, "zone": 5},
        )

        out, title = exports.export_efd_parcel_invoice_xlsx(
            "2026-04-01",
            "2026-04-30",
            200,
            None,
            show_parents=True,
            show_main=True,
        )
        assert title == "Acme"
        assert out.is_file()
        wb = load_workbook(out, data_only=False)
        ws = wb.active
        assert ws.cell(1, 1).value == "Acme"
        assert ws.cell(1, 2).value == "4-1-2026"
        assert ws.cell(1, 3).value == "4-30-2026"
        assert ws.cell(2, 1).value == "Parcel Cost"
        assert ws.cell(6, 1).value == "Total parcel quantities"
        assert ws.cell(9, 24).value == "billing_amount"
        assert ws.cell(9, 25).value == "Price to EFD"
        assert ws.cell(9, 26).value == "EFD Revenue"
        assert ws.cell(9, 27).value == "retail_cost"
        assert ws.cell(10, 24).value == 5.22
        assert ws.cell(10, 25).value == "=X10+1.25"
        assert ws.cell(10, 26).value == "=AA10-Y10"
        assert ws.cell(10, 27).value == 10.2
        b2 = ws.cell(2, 2).value
        assert isinstance(b2, str) and b2.startswith("=SUM(X10:")
        assert ws.cell(6, 2).value == "=COUNTA(A10:A10)"
        assert ws.cell(5, 2).value == "=B4-B3"
        wb.close()
        out.unlink(missing_ok=True)


def test_export_efd_parcel_invoice_xlsx_custom_efd_fee_formula(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_invoice_custom.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()

        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Acme')"
        )
        conn.execute(
            "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'x.csv', 1)"
        )
        import_id = conn.execute(
            "SELECT id FROM billing_imports WHERE billing_id = 'b1'"
        ).fetchone()["id"]
        conn.execute(
            """
            INSERT INTO billing_records (
                billing_import_id, custom_account_code, time_stamp, weight_oz,
                billing_amount, zone, imb_tracking_code, impb
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (import_id, 200, "4/1/2026 15:34", 16.0, 5.22, "5", "trk", "impb1"),
        )
        conn.commit()
        conn.close()

        monkeypatch.setattr(
            db,
            "compute_retail_cost_for_piece",
            lambda *a, **k: {"retail": 10.2, "zone": 5},
        )

        out, _title = exports.export_efd_parcel_invoice_xlsx(
            "2026-04-01",
            "2026-04-30",
            200,
            None,
            show_parents=True,
            show_main=True,
            efd_parcel_fee=0.33,
        )
        wb = load_workbook(out, data_only=False)
        ws = wb.active
        assert ws.cell(10, 25).value == "=X10+0.33"
        wb.close()
        out.unlink(missing_ok=True)


def test_export_efd_parcel_invoice_xlsx_empty_rows(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_empty.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()
        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (300, 'Solo')"
        )
        conn.commit()
        conn.close()

        out, _title = exports.export_efd_parcel_invoice_xlsx(
            "2026-04-01",
            "2026-04-30",
            300,
            None,
            show_parents=True,
            show_main=True,
        )
        wb = load_workbook(out, data_only=False)
        ws = wb.active
        assert ws.cell(1, 2).value == "4-1-2026"
        assert ws.cell(2, 1).value == "Parcel Cost"
        assert ws.cell(2, 2).value == 0
        assert ws.cell(3, 2).value == 0
        assert ws.cell(6, 1).value == "Total parcel quantities"
        assert ws.cell(6, 2).value == 0
        wb.close()
        out.unlink(missing_ok=True)


def test_efd_weekly_invoice_download_name() -> None:
    assert exports.efd_weekly_invoice_download_name(
        "2026-04-27", "2026-05-01"
    ) == "EFD 4-27 to 5-1.xlsx"
    assert exports.efd_weekly_invoice_download_name(
        "2026-04-01", "2026-04-01"
    ) == "EFD 4-1-2026.xlsx"


def test_export_efd_weekly_invoice_xlsx_workbook_and_summary_formulas(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_weekly.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()

        conn = db.get_connection()
        for pn, name in (
            (3901, "Blue Cross Blue Shield -EFD Mailing"),
            (3899, "GEHA -EFD Mailing"),
            (3900, "Security Benefit/Zinnia -EFD Mailing"),
        ):
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (?, ?)",
                (pn, name),
            )
        conn.commit()
        conn.close()

        out = exports.export_efd_weekly_invoice_xlsx(
            "2026-04-27",
            "2026-05-01",
            discount=0.1,
            efd_parcel_fee=1.25,
        )
        assert out.is_file()
        wb = load_workbook(out, data_only=False)
        assert len(wb.sheetnames) == 7
        assert wb.sheetnames[0] == "Summary"
        for pn in (3901, 3899, 3900):
            assert f"{pn} Postage" in wb.sheetnames
            assert f"{pn} Parcel" in wb.sheetnames

        ws = wb["Summary"]
        assert ws["A1"].value == "BCBS"
        assert ws["B1"].value == "Flats"
        assert ws["A2"].value == "BCBS"
        assert ws["B2"].value == "Parcels"
        assert ws["A3"].value == "BCBS"
        assert ws["B3"].value == "Total"
        assert ws["A4"].value == "GEHA"
        assert ws["B4"].value == "Flats"
        assert ws["A5"].value == "GEHA"
        assert ws["B5"].value == "Parcels"
        assert ws["A6"].value == "GEHA"
        assert ws["B6"].value == "Total"
        assert ws["A7"].value == "Zinnia"
        assert ws["B7"].value == "Flats"
        assert ws["A8"].value == "Zinnia"
        assert ws["B8"].value == "Parcels"
        assert ws["A9"].value == "Zinnia"
        assert ws["B9"].value == "Total"
        assert ws["A10"].value == "Grand Total"
        assert ws["B10"].value == "Grand Total"

        assert ws["C1"].value == "='3901 Postage'!L31"
        assert ws["D1"].value == "='3901 Postage'!I32"
        assert ws["C2"].value == "='3901 Parcel'!B3"
        assert ws["D2"].value == "='3901 Parcel'!B6"
        assert ws["C3"].value == "=SUM(C1:C2)"
        assert ws["D3"].value == "=SUM(D1:D2)"
        assert ws["C4"].value == "='3899 Postage'!L31"
        assert ws["D6"].value == "=SUM(D4:D5)"
        assert ws["C10"].value == "=SUM(C3,C6,C9)"
        assert ws["D10"].value == "=SUM(D3,D6,D9)"

        parcel_ws = wb["3900 Parcel"]
        assert parcel_ws.cell(3, 2).value == 0
        assert parcel_ws.cell(6, 2).value == 0

        wb.close()
        out.unlink(missing_ok=True)


def test_efd_weekly_account_download_name() -> None:
    assert exports.efd_weekly_account_download_name(
        "BCBS", "2026-04-27", "2026-05-01"
    ) == "EFD BCBS 4-27 to 5-1.xlsx"
    assert exports.efd_weekly_account_download_name(
        "Zinnia", "2026-04-01", "2026-04-01"
    ) == "EFD Zinnia 4-1-2026.xlsx"


def test_efd_report_date_range_label() -> None:
    assert exports.efd_report_date_range_label("2026-06-08", "2026-06-12") == "6-8 to 6-12"
    assert exports.efd_report_date_range_label("2026-06-08", "2026-06-08") == "6-8-2026"
    assert exports.efd_report_date_range_label("2025-12-31", "2026-01-02") == "12-31-2025 to 1-2-2026"


def test_efd_account_report_download_name() -> None:
    n = exports.efd_account_report_download_name(
        title_name="KC Presort LLC",
        parent_number=3906,
        report_label="Parcel invoice",
        start_date="2026-06-08",
        end_date="2026-06-12",
        ext="xlsx",
    )
    assert n == "KC Presort LLC -EFD (3906) Parcel invoice 6-8 to 6-12.xlsx"

    n2 = exports.efd_account_report_download_name(
        title_name="Blue Cross Blue Shield -EFD Mailing",
        parent_number=3901,
        report_label="Flats Report",
        start_date="2026-04-01",
        end_date="2026-04-01",
        ext="csv",
    )
    assert n2 == "Blue Cross Blue Shield -EFD Mailing (3901) Flats Report 4-1-2026.csv"


def test_postage_invoice_download_name() -> None:
    n = exports.postage_invoice_download_name(
        title_name="Blue Cross Blue Shield",
        parent_number=3901,
        end_date="2026-06-12",
    )
    assert n == "Blue Cross Blue Shield (3901) Postage invoice 6-12-2026.xlsx"


def test_efd_weekly_bundle_folder_name() -> None:
    assert exports.efd_weekly_bundle_folder_name("2026-06-12") == "Weekly EFD 6-12-26"
    assert exports.efd_weekly_bundle_folder_name("2026-01-05") == "Weekly EFD 1-5-26"


def test_efd_weekly_bundle_output_dir(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(exports, "POSTAGE_REPORTS_DIR", tmp_path / "PostageReports")
    out_dir = exports.efd_weekly_bundle_output_dir("2026-06-12")
    assert out_dir.is_dir()
    assert out_dir.name == "Weekly EFD 6-12-26"
    assert out_dir.parent == tmp_path / "PostageReports"


def test_export_efd_weekly_invoice_single_account_workbook(monkeypatch) -> None:
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_weekly_single.db"
        monkeypatch.setattr(db, "DB_PATH", p)
        db.init_db()

        conn = db.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (3901, 'BCBS')"
        )
        conn.commit()
        conn.close()

        out = exports.export_efd_weekly_invoice_xlsx(
            "2026-04-27",
            "2026-05-01",
            discount=0.1,
            efd_parcel_fee=1.25,
            parent_number=3901,
        )
        assert out.is_file()
        wb = load_workbook(out, data_only=False)
        assert wb.sheetnames == ["Summary", "Postage", "Parcel"]

        ws = wb["Summary"]
        assert ws["A1"].value == "BCBS"
        assert ws["B1"].value == "Flats"
        assert ws["B2"].value == "Parcels"
        assert ws["B3"].value == "Total"
        assert ws["C1"].value == "='Postage'!L31"
        assert ws["D2"].value == "='Parcel'!B6"
        assert ws["C3"].value == "=SUM(C1:C2)"
        assert ws["D3"].value == "=SUM(D1:D2)"
        assert ws["A4"].value is None
        assert ws["A10"].value is None

        wb.close()
        out.unlink(missing_ok=True)


def test_efd_weekly_summary_label_invalid_parent() -> None:
    with pytest.raises(ValueError, match="3901"):
        exports.efd_weekly_summary_label(9999)
