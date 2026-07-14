"""Tests for WS3 Customer Mail Detail parsing."""

import sqlite3
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

import db as dbmod
import ws3_mail_detail as ws3


def _make_ws3_workbook(path: Path, bi8: str, *, customer: str = "Acme Corp") -> None:
    """Write a minimal WS3 workbook (BI8 mail id + one customer/profile/rate row)."""
    wb = Workbook()
    ws = wb.active
    ws["BI8"] = bi8
    ws.cell(row=1, column=2, value=customer)
    ws.cell(row=1, column=15, value="$ 0.970")
    ws.cell(row=1, column=17, value="Metered")
    ws.cell(row=2, column=8, value="301079 Acme Profile .970")
    ws.cell(row=3, column=20, value="ADC Auto")
    ws.cell(row=3, column=54, value="9.70")
    ws.cell(row=3, column=67, value="10")
    ws.cell(row=3, column=71, value="10")
    wb.save(path)
    wb.close()


@pytest.mark.parametrize(
    "name,expected",
    [
        ("WS3_FCFL_CustomerMailDetail_4-6-26.xls", "2026-04-06"),
        ("WS3_FCFL_CustomerMailDetail 4-6-26.xls", "2026-04-06"),
        ("x/WS3_FCFL_CustomerMailDetail_12-1-26.xlsx", "2026-12-01"),
    ],
)
def test_parse_date_from_filename(name, expected):
    assert ws3.parse_date_from_filename(name) == expected


def test_parse_date_from_filename_bad():
    assert ws3.parse_date_from_filename("BM_1_2_26.xls") is None


def test_parse_ws3_xlsx_minimal(tmp_path):
    """Minimal sheet matching parser state machine: customer → profile → one rate row."""
    wb = Workbook()
    ws = wb.active
    # Customer header row: col1 name, col14 $..., col16 Metered
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 0.970")
    ws.cell(row=1, column=17, value="Metered")
    # Profile line col H = col 8 -> index 7 in 0-based is column 8? openpyxl column 8 is H
    # Column index 7 in 0-based = Excel column 8 (H)
    ws.cell(row=2, column=8, value="301079 Acme Profile .970")
    # Data row: rate col 20 = T -> index 19
    ws.cell(row=3, column=20, value="ADC Auto")
    ws.cell(row=3, column=41, value="10.00")  # postage claimed col 41
    ws.cell(row=3, column=54, value="9.70")  # postage applied
    ws.cell(row=3, column=67, value="10")  # num pieces
    ws.cell(row=3, column=71, value="10")  # pcs accepted
    p = tmp_path / "t.xlsx"
    wb.save(p)

    mid, rdt, customers, rows = ws3.parse_ws3_xlsx(str(p))
    assert customers.get("301079") == "Acme Corp"
    assert len(rows) == 1
    assert rows[0]["customer_code"] == "301079"
    assert rows[0]["rate_type"] == "ADC Auto"
    assert rows[0]["num_pieces"] == 10
    assert rows[0]["pcs_rejected"] == 0


def test_parse_ws3_xlsx_rate_on_profile_row_is_captured(tmp_path):
    """This NetSort layout puts the first rate type on the profile-name row."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 0.970")
    ws.cell(row=1, column=17, value="Metered")
    # Profile name (col H = 8) AND first rate type (col T = 20) on the SAME row.
    ws.cell(row=2, column=8, value="301079 Acme Profile .970")
    ws.cell(row=2, column=20, value="ThreeDigitAuto")
    ws.cell(row=2, column=54, value="9.70")
    ws.cell(row=2, column=67, value="5")
    ws.cell(row=2, column=71, value="5")
    # Second rate type on its own row.
    ws.cell(row=3, column=20, value="ADC Auto")
    ws.cell(row=3, column=54, value="4.85")
    ws.cell(row=3, column=67, value="5")
    ws.cell(row=3, column=71, value="5")
    p = tmp_path / "t.xlsx"
    wb.save(p)
    wb.close()

    _, _, customers, rows = ws3.parse_ws3_xlsx(str(p))
    assert customers.get("301079") == "Acme Corp"
    assert len(rows) == 2
    assert {r["rate_type"] for r in rows} == {"ThreeDigitAuto", "ADC Auto"}
    assert all(r["customer_code"] == "301079" for r in rows)


def test_parse_ws3_xlsx_new_rate_labels_captured(tmp_path):
    """New-rate NetSort export renamed labels (MXDAuto/ADCAuto); capture them as-is."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 1.025")
    ws.cell(row=1, column=17, value="Metered")
    ws.cell(row=2, column=8, value="301079 Acme Profile 1.025")
    # New Mixed-ADC label.
    ws.cell(row=3, column=20, value="MXDAuto")
    ws.cell(row=3, column=41, value="15.80")  # postage claimed
    ws.cell(row=3, column=54, value="10.25")  # postage applied
    ws.cell(row=3, column=67, value="10")  # num pieces
    ws.cell(row=3, column=71, value="10")  # pcs accepted
    # New ADC label (defensive).
    ws.cell(row=4, column=20, value="ADCAuto")
    ws.cell(row=4, column=41, value="13.26")
    ws.cell(row=4, column=54, value="10.25")
    ws.cell(row=4, column=67, value="10")
    ws.cell(row=4, column=71, value="10")
    p = tmp_path / "t.xlsx"
    wb.save(p)
    wb.close()

    _, _, _, rows = ws3.parse_ws3_xlsx(str(p))
    by_type = {r["rate_type"]: r for r in rows}
    assert set(by_type) == {"MXDAuto", "ADCAuto"}
    mxd = by_type["MXDAuto"]
    assert mxd["rate_type"] == "MXDAuto"  # stored as-is, not normalized
    assert mxd["num_pieces"] == 10
    assert mxd["postage_claimed"] == 15.80
    assert ws3.calc_usps_cost_per_piece(mxd["postage_claimed"], mxd["num_pieces"]) == 1.58
    assert by_type["ADCAuto"]["pcs_rejected"] == 0


def test_parse_ws3_xlsx_unrecognized_rate_type_skipped_and_warns(tmp_path, caplog):
    """A rate-like row with an unknown label is skipped but logged (no silent drop)."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 1.025")
    ws.cell(row=1, column=17, value="Metered")
    ws.cell(row=2, column=8, value="301079 Acme Profile 1.025")
    ws.cell(row=3, column=20, value="BrandNewSort")
    ws.cell(row=3, column=41, value="12.00")  # postage claimed
    ws.cell(row=3, column=54, value="10.25")  # postage applied
    ws.cell(row=3, column=67, value="10")  # num pieces
    ws.cell(row=3, column=71, value="10")  # pcs accepted
    p = tmp_path / "t.xlsx"
    wb.save(p)
    wb.close()

    with caplog.at_level("WARNING", logger="ws3_mail_detail"):
        _, _, _, rows = ws3.parse_ws3_xlsx(str(p))
    assert rows == []
    assert any("BrandNewSort" in rec.getMessage() for rec in caplog.records)


def test_parse_ws3_xlsx_known_ignored_rate_type_no_warn(tmp_path, caplog):
    """MXD ADC Machinable is intentionally excluded and must not warn."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 1.025")
    ws.cell(row=1, column=17, value="Metered")
    ws.cell(row=2, column=8, value="301079 Acme Profile 1.025")
    ws.cell(row=3, column=20, value="MXD ADC Machinable")
    ws.cell(row=3, column=41, value="12.00")
    ws.cell(row=3, column=54, value="10.25")
    ws.cell(row=3, column=67, value="10")
    ws.cell(row=3, column=71, value="10")
    p = tmp_path / "t.xlsx"
    wb.save(p)
    wb.close()

    with caplog.at_level("WARNING", logger="ws3_mail_detail"):
        _, _, _, rows = ws3.parse_ws3_xlsx(str(p))
    assert rows == []
    assert not any("Machinable" in rec.getMessage() for rec in caplog.records)


def test_process_ws3_mislabeled_xls_imports_rows(monkeypatch, tmp_path):
    """An OOXML workbook saved with a .xls name must import without LibreOffice."""
    import db as dbmod
    import importer

    wb = Workbook()
    ws = wb.active
    ws["BI8"] = "040626_F"
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 0.970")
    ws.cell(row=1, column=17, value="Metered")
    ws.cell(row=2, column=8, value="301079 Acme Profile .970")
    ws.cell(row=2, column=20, value="ADC Auto")
    ws.cell(row=2, column=54, value="9.70")
    ws.cell(row=2, column=67, value="10")
    ws.cell(row=2, column=71, value="10")
    src_xlsx = tmp_path / "src.xlsx"
    wb.save(src_xlsx)
    wb.close()

    db_path = tmp_path / "t.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    # Fail loudly if conversion is attempted: a real OOXML file must not be converted.
    def _no_convert(*_a, **_k):
        raise AssertionError("convert_xls_to_xlsx should not be called for OOXML content")

    monkeypatch.setattr(importer, "convert_xls_to_xlsx", _no_convert)

    dest = tmp_path / "WS3_FCFL_CustomerMailDetail(9).xls"
    dest.write_bytes(src_xlsx.read_bytes())

    out = ws3.process_ws3_mail_detail_file(str(dest), db_path)
    assert out.get("skipped") is False
    assert out.get("mail_date") == "2026-04-06"
    assert out.get("rows_imported", 0) >= 1


def test_process_ws3_same_filename_different_mailing_both_import(monkeypatch, tmp_path):
    """Reused download names (e.g. '(15)') with different data must both import."""
    db_path = tmp_path / "t.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    d1 = tmp_path / "a"
    d2 = tmp_path / "b"
    d1.mkdir()
    d2.mkdir()
    f1 = d1 / "WS3_FCFL_CustomerMailDetail(15).xlsx"
    f2 = d2 / "WS3_FCFL_CustomerMailDetail(15).xlsx"
    _make_ws3_workbook(f1, "050726_F")
    _make_ws3_workbook(f2, "071326_F")

    out1 = ws3.process_ws3_mail_detail_file(str(f1), db_path)
    out2 = ws3.process_ws3_mail_detail_file(str(f2), db_path)
    assert out1.get("skipped") is False and out1.get("mail_date") == "2026-05-07"
    assert out2.get("skipped") is False and out2.get("mail_date") == "2026-07-13"

    conn = dbmod.get_connection()
    try:
        n_runs = conn.execute("SELECT COUNT(*) FROM ws3_mail_runs").fetchone()[0]
        n_imports = conn.execute("SELECT COUNT(*) FROM ws3_imports").fetchone()[0]
    finally:
        conn.close()
    assert n_runs == 2
    assert n_imports == 2


def test_process_ws3_same_mailing_different_filename_skips(monkeypatch, tmp_path):
    """Identical mailing (mail_date, mail_id) from a differently named file is skipped."""
    db_path = tmp_path / "t.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    f1 = tmp_path / "WS3_FCFL_CustomerMailDetail(14).xlsx"
    f2 = tmp_path / "WS3_FCFL_CustomerMailDetail(54).xlsx"
    _make_ws3_workbook(f1, "062926_F")
    _make_ws3_workbook(f2, "062926_F")

    out1 = ws3.process_ws3_mail_detail_file(str(f1), db_path)
    out2 = ws3.process_ws3_mail_detail_file(str(f2), db_path)
    assert out1.get("skipped") is False
    assert out2.get("skipped") is True
    assert out2.get("reason") == "duplicate_mail_run"

    conn = dbmod.get_connection()
    try:
        n_runs = conn.execute("SELECT COUNT(*) FROM ws3_mail_runs").fetchone()[0]
    finally:
        conn.close()
    assert n_runs == 1


def test_process_ws3_reimport_same_file_is_idempotent(monkeypatch, tmp_path):
    db_path = tmp_path / "t.db"
    monkeypatch.setattr(dbmod, "DB_PATH", db_path)
    dbmod.init_db()

    f = tmp_path / "WS3_FCFL_CustomerMailDetail(1).xlsx"
    _make_ws3_workbook(f, "040626_F")

    out1 = ws3.process_ws3_mail_detail_file(str(f), db_path)
    out2 = ws3.process_ws3_mail_detail_file(str(f), db_path)
    assert out1.get("skipped") is False
    assert out2.get("skipped") is True
    assert out2.get("reason") == "duplicate_mail_run"


def test_migrate_ws3_imports_drops_filename_unique(tmp_path):
    """Legacy DBs with UNIQUE(file_name) are rebuilt without it, preserving rows."""
    db_path = tmp_path / "legacy.db"
    conn = sqlite3.connect(str(db_path))
    conn.executescript(
        """
        CREATE TABLE ws3_mail_runs (
            run_id INTEGER PRIMARY KEY AUTOINCREMENT,
            mail_date TEXT NOT NULL,
            mail_id TEXT NOT NULL DEFAULT '',
            run_datetime TEXT,
            source_file_name TEXT,
            UNIQUE (mail_date, mail_id)
        );
        CREATE TABLE ws3_imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT NOT NULL UNIQUE,
            mail_date TEXT NOT NULL,
            run_id INTEGER NOT NULL REFERENCES ws3_mail_runs (run_id) ON DELETE CASCADE,
            row_count INTEGER,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    conn.execute("INSERT INTO ws3_mail_runs (mail_date, mail_id) VALUES ('2026-05-07', '050726_F')")
    conn.execute("INSERT INTO ws3_mail_runs (mail_date, mail_id) VALUES ('2026-07-13', '071326_F')")
    conn.execute(
        "INSERT INTO ws3_imports (file_name, mail_date, run_id, row_count) VALUES ('dup.xls', '2026-05-07', 1, 66)"
    )
    conn.commit()

    assert dbmod._ws3_imports_has_filename_unique(conn) is True
    dbmod._migrate_ws3_imports_drop_filename_unique(conn)
    conn.commit()
    assert dbmod._ws3_imports_has_filename_unique(conn) is False

    # A second row with the same filename now succeeds.
    conn.execute(
        "INSERT INTO ws3_imports (file_name, mail_date, run_id, row_count) VALUES ('dup.xls', '2026-07-13', 2, 20)"
    )
    conn.commit()

    rows = conn.execute(
        "SELECT file_name, mail_date, row_count FROM ws3_imports ORDER BY id"
    ).fetchall()
    conn.close()
    assert rows == [("dup.xls", "2026-05-07", 66), ("dup.xls", "2026-07-13", 20)]


def test_parse_ws3_xlsx_negative_rejected_clamped_to_zero(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Acme Corp")
    ws.cell(row=1, column=15, value="$ 0.970")
    ws.cell(row=1, column=17, value="Metered")
    ws.cell(row=2, column=8, value="301079 Acme Profile .970")
    ws.cell(row=3, column=20, value="ADC Auto")
    ws.cell(row=3, column=54, value="9.70")  # postage applied
    ws.cell(row=3, column=67, value="10")  # num pieces
    ws.cell(row=3, column=71, value="12")  # pcs accepted (bad data -> negative rejected)
    p = tmp_path / "t.xlsx"
    wb.save(p)

    _, _, _, rows = ws3.parse_ws3_xlsx(str(p))
    assert len(rows) == 1
    assert rows[0]["pcs_rejected"] == 0


def test_list_ws3_assignment_accounts_parent_vs_main(monkeypatch):
    """Parent = has children; main = no customers use this number as parent_number."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "t.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        conn.executemany(
            """
            INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
            VALUES (?, ?, ?, ?)
            """,
            [
                (100, "Parent Co", None, None),
                (101, "Child A", 100, "Parent Co"),
                (200, "Solo Inc", None, None),
            ],
        )
        conn.commit()
        accts = dbmod.list_ws3_assignment_accounts(conn)
        conn.close()
    kinds = {a["customer_number"]: a["kind"] for a in accts}
    assert kinds[100] == "parent"
    assert kinds[101] == "main"
    assert kinds[200] == "main"


def test_query_postage_includes_ws3_reject_row(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "t.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        conn.execute(
            "INSERT INTO customers (customer_number, customer_name) VALUES (100, 'Parent Co')"
        )
        conn.execute(
            """
            INSERT INTO ws3_parent_daily_rejects (mail_date, parent_customer_number, reject_count)
            VALUES ('2026-04-01', 100, 7)
            """
        )
        conn.commit()
        data = dbmod.query_postage(
            conn,
            "2026-04-01",
            "2026-04-30",
            None,
            None,
            True,
            True,
            False,
            False,
            False,
        )
        conn.close()
    rej = [r for r in data["rows"] if r.get("mail_class") == dbmod.WS3_REJECT_MAIL_CLASS]
    assert len(rej) == 1
    assert rej[0]["total_qty"] == 7
    assert rej[0]["child_number"] == 100
