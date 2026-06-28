"""Smoke tests for database init and summary query."""

import tempfile
from pathlib import Path

import pytest

import db as dbmod
import exports as exports_mod


def _seed_retail_matrix_rates(
    conn,
    *,
    table: str,
    rows: list[tuple[int, int, float]],
    effective_date: str | None = "2000-01-01",
) -> None:
    """
    Insert minimal retail matrix rows for tests.

    Each row is (zone, lb, price) and is stored as row_type='matrix', weight_unit='lb', weight_max=<lb>.
    ``effective_date`` is set for Priority Mail versioned lookups (default is on or before typical test dates).
    """
    conn.executemany(
        f"""
        INSERT INTO {table} (effective_date, row_type, zone, weight_unit, weight_max, price)
        VALUES (?, 'matrix', ?, 'lb', ?, ?)
        """,
        [
            (effective_date, int(z), float(lb), float(price))
            for (z, lb, price) in rows
        ],
    )


def test_get_priority_mail_retail_rates_tariff_as_of_effective_dates(monkeypatch):
    """Latest matrix row wins per zone×lb when effective_date ≤ as_of."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "pm_tariff.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.executemany(
                """
                INSERT INTO priority_mail_retail (
                    effective_date, row_type, zone, weight_unit, weight_max, price,
                    sort_group, sort_order
                ) VALUES (?, 'matrix', ?, 'lb', ?, ?, 1, ?)
                """,
                [
                    ("2020-01-01", 2, 1.0, 10.0, 0),
                    ("2026-04-27", 2, 1.0, 20.0, 1),
                ],
            )
            conn.commit()
            assert dbmod.get_priority_mail_retail_rates(conn, as_of_date="2026-04-01")[0][
                (2, 1)
            ] == 10.0
            assert dbmod.get_priority_mail_retail_rates(conn, as_of_date="2026-04-27")[0][
                (2, 1)
            ] == 20.0
            assert dbmod.get_priority_mail_retail_rates(conn, as_of_date="2027-01-01")[0][
                (2, 1)
            ] == 20.0
        finally:
            conn.close()


def test_init_and_summary_empty(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "test.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            s = dbmod.query_summary(conn, "2026-01-01", "2026-01-31")
        finally:
            conn.close()

    assert s["date_range"] == {"start": "2026-01-01", "end": "2026-01-31"}
    assert s["postage"]["total_pieces"] == 0
    assert s["postage"]["total_cost"] == 0.0
    assert s["parcels"]["total_pieces"] == 0
    assert s["parcels"]["total_billed"] == 0.0
    assert s["imports"] == []
    assert s["postage"]["by_customer"] == []
    assert s["parcels"]["by_customer"] == []


def test_customer_hierarchy(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "hier.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (100, "Parent Co", None, None),
                    (101, "Child A", 100, "Parent Co"),
                    (102, "Child B", 100, "Parent Co"),
                    (200, "Solo Inc", None, None),
                ],
            )
            conn.commit()
            h = dbmod.query_customer_hierarchy(conn)
        finally:
            conn.close()

    assert len(h["parents"]) == 1
    assert h["parents"][0]["customer_number"] == 100
    assert h["parents"][0]["child_count"] == 2
    assert {c["customer_number"] for c in h["parents"][0]["children"]} == {101, 102}
    assert h["standalone"] == [{"customer_number": 200, "customer_name": "Solo Inc"}]


def test_query_summary_parcel_billing_timestamp(monkeypatch):
    """Parcel summary must parse M/D/YYYY HH:MM (year is not the last 4 chars of the string)."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "parcel_sum.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (9999, 'TestAcct', NULL, NULL)
                """
            )
            conn.execute(
                """
                INSERT INTO billing_imports (billing_id, file_name, row_count)
                VALUES ('t1', 'test.csv', 1)
                """
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, unmatched_account
                ) VALUES (1, 9999, 'TestAcct', '4/1/2026 15:34', 1.0, 'USPS GROUND ADVANTAGE', 5.25, 0)
                """
            )
            conn.commit()
            s = dbmod.query_summary(conn, "2026-04-01", "2026-04-30")
        finally:
            conn.close()

    assert s["parcels"]["total_pieces"] == 1
    assert s["parcels"]["total_billed"] == 5.25
    by_class = {r["mail_class"]: r for r in s["parcels"]["by_class"]}
    assert by_class["USPS GROUND ADVANTAGE"]["pieces"] == 1
    cust = {r["customer_number"]: r for r in s["parcels"]["by_customer"] if r["customer_number"] is not None}
    assert cust[9999]["pieces"] == 1
    assert cust[9999]["cost"] == 5.25
    assert cust[9999]["unmatched"] is False


def test_query_summary_postage_unmatched_orphan(monkeypatch):
    """Postage for unknown account appears as Unmatched and rolls into totals."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "post_orphan.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (100, 'Known', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('r.csv', '2026-02-10', 2)"
            )
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-02-10', ?, 'C', 1.0, 3, 1.5, 0)
                """,
                [(100,), (7777,)],
            )
            conn.commit()
            s = dbmod.query_summary(conn, "2026-02-01", "2026-02-28")
        finally:
            conn.close()

    by_num = {r["customer_number"]: r for r in s["postage"]["by_customer"]}
    assert by_num[100]["unmatched"] is False
    assert by_num[100]["pieces"] == 3
    orphan = by_num[7777]
    assert orphan["customer_name"] == "Unmatched"
    assert orphan["unmatched"] is True
    assert orphan["pieces"] == 3
    assert s["postage"]["total_pieces"] == 6
    assert s["postage"]["total_cost"] == 3.0


def test_query_summary_parcel_unmatched_orphan(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "par_orphan.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (100, 'Known', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 3)"
            )
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, unmatched_account
                ) VALUES (1, ?, 'x', '2/15/2026 10:00', 2.0, 'CLS', 2.0, 1)
                """,
                [(100,), (8888,), (None,)],
            )
            conn.commit()
            s = dbmod.query_summary(conn, "2026-02-01", "2026-02-28")
        finally:
            conn.close()

    rows = s["parcels"]["by_customer"]
    matched = next(r for r in rows if r["customer_number"] == 100)
    assert matched["unmatched"] is False
    assert matched["pieces"] == 1
    unk = next(r for r in rows if r["customer_number"] == 8888)
    assert unk["customer_name"] == "Unmatched"
    assert unk["unmatched"] is True
    assert unk["pieces"] == 1
    null_cac = next(r for r in rows if r["customer_number"] is None)
    assert null_cac["unmatched"] is True
    assert null_cac["pieces"] == 1
    assert s["parcels"]["total_pieces"] == 3
    assert s["parcels"]["total_billed"] == 6.0


def test_list_unmatched_accounts_all_time_and_upsert_customer(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "unmatched_assign.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            # One known parent account.
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (100, 'Parent Co', NULL, NULL)
                """
            )
            # Create unmatched usage in both postage and parcels.
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('r.csv', '2026-02-10', 1)"
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-02-10', 7777, 'C', 1.0, 3, 1.5, 1)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 1)"
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, unmatched_account
                ) VALUES (1, 8888, 'x', '2/15/2026 10:00', 2.0, 'CLS', 2.0, 1)
                """
            )
            conn.commit()

            unm = dbmod.list_unmatched_accounts_all_time(conn)
            by_code = {r["account_code"]: r for r in unm}
            assert 7777 in by_code
            assert 8888 in by_code
            assert by_code[7777]["sources"] == "postage"
            assert by_code[7777]["postage_pieces"] == 3
            assert by_code[7777]["postage_cost"] == 1.5
            assert by_code[8888]["sources"] == "parcels"
            assert by_code[8888]["parcel_pieces"] == 1
            assert by_code[8888]["parcel_cost"] == 2.0

            # Assign customer 7777 to Parent Co; it should disappear from unmatched list.
            with conn:
                dbmod.upsert_customer(conn, 7777, "New Child", parent_number=100)
            unm2 = dbmod.list_unmatched_accounts_all_time(conn)
            codes2 = {r["account_code"] for r in unm2}
            assert 7777 not in codes2
            assert 8888 in codes2

            row = conn.execute(
                "SELECT customer_name, parent_number, parent_name FROM customers WHERE customer_number = 7777"
            ).fetchone()
            assert row is not None
            assert row["customer_name"] == "New Child"
            assert row["parent_number"] == 100
            assert row["parent_name"] == "Parent Co"
        finally:
            conn.close()


def test_query_parcel_report_rows_parent_name_coalesce(monkeypatch):
    """Column C uses COALESCE(parent_name, customer_name) for matched accounts."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "parcel_export.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (200, 'Solo Inc', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 1)"
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, piece_id
                ) VALUES (1, 200, 'Solo Inc', '4/1/2026 10:00', 16.0, 'CLS', 1.0, 'P1')
                """
            )
            conn.commit()
            rows = dbmod.query_parcel_report_rows(
                conn, "2026-04-01", "2026-04-30", None, None, True, True
            )
        finally:
            conn.close()

    assert len(rows) == 1
    assert rows[0]["parent_name"] == "Solo Inc"


def test_query_parcel_report_rows_customer_number_filter(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "parcel_export2.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (100, "Parent Co", None, None),
                    (101, "Child A", 100, "Parent Co"),
                    (102, "Child B", 100, "Parent Co"),
                ],
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 2)"
            )
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, piece_id
                ) VALUES (1, ?, 'x', '4/2/2026 11:00', 8.0, 'CLS', 2.0, ?)
                """,
                [(101, "PA"), (102, "PB")],
            )
            conn.commit()
            rows = dbmod.query_parcel_report_rows(
                conn, "2026-04-01", "2026-04-30", 100, 101, True, True
            )
        finally:
            conn.close()

    assert len(rows) == 1
    assert rows[0]["custom_account_code"] == 101
    assert rows[0]["piece_id"] == "PA"


def test_export_parcel_report_workbook(monkeypatch):
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "parcel_xlsx.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (100, "Parent Co", None, None),
                    (101, "Child A", 100, "Parent Co"),
                ],
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 1)"
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, piece_id
                ) VALUES (1, 101, 'Child A', '4/2/2026 11:00', 8.0, 'CLS', 2.0, 2.5, 'PX')
                """
            )
            conn.commit()
        finally:
            conn.close()

        out = exports_mod.export_parcel_report("2026-04-01", "2026-04-30", 100, None, True, True)
        try:
            wb = load_workbook(out)
            ws = wb.active
            assert ws.cell(1, 1).value == "Customer #"
            assert ws.cell(1, 3).value == "Parent Name"
            assert ws.cell(2, 1).value == 101
            assert ws.cell(2, 3).value == "Parent Co"
            assert ws.cell(2, 9).value == "=H2/16"
            assert ws.cell(1, 13).value == "IMPB"
            assert ws.cell(2, 13).value in (None, "")
            tot_row = 3
            assert ws.cell(tot_row, 1).value == "TOTALS"
            assert ws.cell(tot_row, 10).value == "=SUM(J2:J2)"
            assert ws.max_row == tot_row
        finally:
            out.unlink(missing_ok=True)

    assert exports_mod.parcel_report_download_name("2026-04-01", "2026-04-02", 100, 101) == (
        "Parcel_Report_100_c101_2026-04-01_2026-04-02.xlsx"
    )


def test_export_parcel_zone_summary_includes_af_hm_sections(monkeypatch, tmp_path):
    """Parcel Invoice workbook: over-10lb block, then per-customer invoice table below the zone grid."""
    from openpyxl import load_workbook

    # Isolate from the ambient postage.db: the export queries billing_records via
    # db.get_connection(), so the schema must exist (it is empty here).
    monkeypatch.setattr(dbmod, "DB_PATH", tmp_path / "afhm.db")
    dbmod.init_db()

    def fake_af_hm(*_a, **_k):
        return {
            "heavy_rows": [
                {
                    "customer_name": "Heavy Child",
                    "count": 2,
                    "lbs": 15,
                    "zone": 2,
                    "base": 20.5,
                    "efd": 20.0,
                    "savings": 0.5,
                },
            ],
            "customers": [
                {
                    "customer_number": 101,
                    "name": "Child",
                    "qty": 3,
                    "cost": 10.0,
                    "discount": 9.25,
                    "savings": 0.75,
                }
            ],
            "grand_total_qty": 5,
        }

    monkeypatch.setattr(dbmod, "compute_parcel_report_af_hm_sections", fake_af_hm)

    def row_template(lb: int) -> dict:
        return {
            "weight_label": f"{lb} lb",
            "zone_a": {"priority": 1.0, "efd": 1.0, "count": 0},
            "zone_b": {"priority": 1.0, "efd": 1.0, "count": 0},
            "costs": None,
            "savings": None,
        }

    rows = [row_template(lb) for lb in range(1, 11)]
    blocks = [
        {"zone_a": za, "zone_b": zb, "rows": rows}
        for za, zb in ((1, 3), (2, 4), (5, 6), (7, 8))
    ]
    summary = {
        "report_date": "01-Apr-2026",
        "title_name": "Test Co",
        "total_pieces": 0,
        "total_cost": 0.0,
        "total_savings": 0.0,
        "blocks": blocks,
    }
    out = exports_mod.export_parcel_zone_summary_xlsx(
        summary,
        start_date="2026-04-01",
        end_date="2026-04-30",
        parent_number=None,
        customer_number=None,
        show_parents=True,
        show_main=True,
    )
    try:
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2
        ws = wb.active
        assert wb.worksheets[1].cell(1, 1).value == "Customer #"
        assert wb.worksheets[1].cell(1, 13).value == "IMPB"
        assert ws.cell(56, 1).value == "Customer Name"
        assert ws.cell(57, 1).value == "Heavy Child"
        assert ws.cell(57, 2).value == 2
        assert ws.cell(58, 1).value == "Total"
        assert str(ws.cell(58, 2).value or "").startswith("=SUM(")
        assert "SUMPRODUCT" in str(ws.cell(58, 6).value or "")
        assert str(ws.cell(58, 7).value or "").startswith("=SUM(")
        assert ws.cell(60, 1).value == "Customer #"
        assert ws.cell(61, 1).value == 101
        assert ws.cell(61, 3).value == 3
        assert ws.cell(62, 1).value == "Total"
        assert str(ws.cell(62, 3).value or "").startswith("=SUM(")
    finally:
        out.unlink(missing_ok=True)


def test_query_parcel_zone_summary_matrix_and_totals(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "zone_sum.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            # Seed minimal Priority Mail retail matrix cells used by this test.
            _seed_retail_matrix_rates(
                conn,
                table="priority_mail_retail",
                rows=[
                    (1, 1, 11.0),
                    (3, 2, 13.0),
                    # For zone 9 we rely on fallback-to-fully_paid_postage in totals.
                ],
            )
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (9999, 'ZoneTest', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('z1', 'z.csv', 3)"
            )
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, zone
                ) VALUES (1, 9999, 'ZoneTest', '4/1/2026 10:00', ?, 'USPS GROUND ADVANTAGE', ?, ?, ?)
                """,
                [
                    (16.0, 10.80, 11.05, "1"),
                    (32.0, 12.20, 12.45, "3"),
                    (16.0, 5.0, 5.5, "9"),
                ],
            )
            conn.commit()
            s = dbmod.query_parcel_zone_summary(
                conn, "2026-04-01", "2026-04-30", None, None, True, True, False
            )
        finally:
            conn.close()

    assert s["total_pieces"] == 3
    # Retail model totals: sum(retail) with fallback-to-fully_paid_postage when rate missing.
    # Pieces:
    # - zone 1, 1 lb => 11.0 (seeded)
    # - zone 3, 2 lb => 13.0 (seeded)
    # - zone 9, 1 lb => fallback fully_paid_postage 5.5 (not seeded)
    assert s["total_cost"] == pytest.approx((11.0 - 0.25) + (13.0 - 0.25) + (5.5 - 0.25), rel=1e-9)
    assert s["total_savings"] == pytest.approx(0.25 + 0.25 + 0.25, rel=1e-9)
    b0 = s["blocks"][0]
    assert b0["rows"][0]["zone_a"]["count"] == 1
    assert b0["rows"][0]["zone_a"]["priority"] == 11.0
    assert b0["rows"][0]["zone_a"]["efd"] == 10.75
    assert b0["rows"][0]["zone_b"]["count"] == 0
    assert b0["rows"][1]["zone_a"]["count"] == 0
    assert b0["rows"][1]["zone_b"]["count"] == 1
    assert b0["rows"][1]["zone_b"]["priority"] == 13.0
    assert b0["rows"][1]["zone_b"]["efd"] == 12.75
    assert b0["rows"][0]["costs"] == pytest.approx(10.75, rel=1e-9)
    assert b0["rows"][0]["savings"] == 0.25
    assert b0["rows"][1]["costs"] == pytest.approx(12.75, rel=1e-9)
    assert b0["rows"][1]["savings"] == 0.25


def test_compute_parcel_report_af_hm_sections(monkeypatch):
    # Per-customer cost/savings follow parcel summary (zone × lb_row≤10), same as zone grid footer.
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "afhm.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            # Seed minimal Priority Mail retail matrix cells used by this test (lbs > 10).
            _seed_retail_matrix_rates(
                conn,
                table="priority_mail_retail",
                rows=[
                    (2, 13, 19.2),
                    (3, 16, 22.05),
                ],
            )
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (500, 'Heavy Co', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 2)"
            )
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, zone
                ) VALUES (1, 500, 'Heavy Co', '4/1/2026 10:00', ?, 'CLS', 1.0, 1.0, ?)
                """,
                [
                    (200.0, "2"),
                    (256.0, "3"),
                ],
            )
            conn.commit()
            s = dbmod.compute_parcel_report_af_hm_sections(
                conn, "2026-04-01", "2026-04-30", None, None, True, True
            )
        finally:
            conn.close()

    assert s["grand_total_qty"] == 2
    assert len(s["heavy_rows"]) == 2
    assert len(s["customers"]) == 1
    assert s["customers"][0]["customer_number"] == 500
    assert s["customers"][0]["name"] == "Heavy Co"
    assert s["customers"][0]["qty"] == 2
    assert s["customers"][0]["cost"] == pytest.approx(19.2 + 22.05, rel=1e-6)
    assert s["customers"][0]["discount"] == pytest.approx((19.2 - 0.25) + (22.05 - 0.25), rel=1e-6)
    assert s["customers"][0]["savings"] == pytest.approx(0.25 + 0.25, rel=1e-6)
    lbs_z = {(r["lbs"], r["zone"]): r["count"] for r in s["heavy_rows"]}
    assert lbs_z[(13, 2)] == 1
    assert lbs_z[(16, 3)] == 1
    assert s["heavy_rows"][0]["customer_name"] == "Heavy Co"
    assert s["heavy_rows"][1]["customer_name"] == "Heavy Co"


def test_compute_parcel_report_af_hm_matches_zone_summary_totals(monkeypatch):
    """Customer invoice cost/savings roll up to the same totals as the zone summary footer."""
    fake_parcel = {(1, 1): (11.0, 10.0), (2, 10): (25.0, 20.0)}
    monkeypatch.setattr(dbmod, "get_parcel_summary_rates", lambda: fake_parcel)
    monkeypatch.setattr(dbmod, "get_heavy_parcel_rates", lambda: {})
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "afhm_align.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (700, 'Align Co', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 2)"
            )
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, zone
                ) VALUES (1, 700, 'Align Co', '4/1/2026 10:00', ?, 'CLS', 1.0, 1.0, ?)
                """,
                [
                    (16.0, "1"),
                    (200.0, "2"),
                ],
            )
            conn.commit()
            af = dbmod.compute_parcel_report_af_hm_sections(
                conn, "2026-04-01", "2026-04-30", None, None, True, True
            )
            zn = dbmod.query_parcel_zone_summary(
                conn, "2026-04-01", "2026-04-30", None, None, True, True, False
            )
        finally:
            conn.close()

    c0 = af["customers"][0]
    assert c0["discount"] == zn["total_cost"]
    assert c0["savings"] == zn["total_savings"]


def test_compute_parcel_report_af_hm_uses_customer_master_name_not_billing(monkeypatch):
    """H–M Customer Name column uses customers.customer_name, not billing_records.account_name."""
    monkeypatch.setattr(dbmod, "get_heavy_parcel_rates", lambda: {})
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "afhm_name.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (600, 'Master Name From Customers', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 1)"
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, zone
                ) VALUES (1, 600, 'Wrong Name On Billing Row', '4/1/2026 10:00', 8.0, 'CLS', 1.0, 1.0, '1')
                """
            )
            conn.commit()
            s = dbmod.compute_parcel_report_af_hm_sections(
                conn, "2026-04-01", "2026-04-30", None, None, True, True
            )
        finally:
            conn.close()

    assert len(s["customers"]) == 1
    assert s["customers"][0]["name"] == "Master Name From Customers"


def test_postage_row_edit_preview_and_apply_merge(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "postage_edit.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, NULL, NULL)
                """,
                [(100, "FromAcct"), (200, "ToAcct")],
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('r.csv', '2026-02-10', 3)"
            )
            # Source account rows (account_code=100) for one mail_class on one day.
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class, weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-02-10', 100, 'C', ?, ?, ?, 0)
                """,
                [
                    (1.0, 10, 5.0),
                    (2.0, 4, 2.0),
                ],
            )
            # Destination already has weight_oz=1.0, so update must merge into it.
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class, weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-02-10', 200, 'C', 1.0, 3, 1.5, 0)
                """
            )
            conn.commit()

            src = conn.execute(
                """
                SELECT id, weight_oz, pieces, total_cost
                FROM postage_data
                WHERE file_date='2026-02-10' AND account_code=100 AND mail_class='C'
                ORDER BY weight_oz
                """
            ).fetchall()
            assert len(src) == 2
            id_w1 = int(src[0]["id"])
            id_w2 = int(src[1]["id"])
            assert float(src[0]["weight_oz"]) == 1.0
            assert float(src[1]["weight_oz"]) == 2.0

            pieces_by_id = {str(id_w1): 12, str(id_w2): 4}

            prev = dbmod.preview_postage_row_update(
                conn,
                file_date="2026-02-10",
                from_account_code=100,
                mail_class="C",
                to_account_code=200,
                pieces_by_id=pieces_by_id,
            )
            assert prev["ok"] is True
            assert prev["summary"]["source_rows"] == 2
            assert prev["summary"]["merged"] == 1
            assert prev["summary"]["updated"] == 1

            with conn:
                out = dbmod.apply_postage_row_update(
                    conn,
                    file_date="2026-02-10",
                    from_account_code=100,
                    mail_class="C",
                    to_account_code=200,
                    pieces_by_id=pieces_by_id,
                    reason="Fix wrong account",
                )
            assert out["ok"] is True
            assert out["summary"]["merged"] == 1
            assert out["summary"]["updated"] == 1

            # Source rows should be gone.
            left = conn.execute(
                "SELECT COUNT(*) AS n FROM postage_data WHERE file_date='2026-02-10' AND account_code=100 AND mail_class='C'"
            ).fetchone()
            assert int(left["n"]) == 0

            # Weight 1.0 should be merged into existing destination row:
            dest_w1 = conn.execute(
                """
                SELECT pieces, total_cost
                FROM postage_data
                WHERE file_date='2026-02-10' AND account_code=200 AND mail_class='C' AND weight_oz=1.0
                """
            ).fetchone()
            assert dest_w1 is not None
            assert int(dest_w1["pieces"]) == 3 + 12
            # Source cost 5.0 scaled from 10->12 pieces => 6.0, added to dest 1.5 => 7.5
            assert float(dest_w1["total_cost"]) == pytest.approx(7.5, rel=1e-9)

            # Weight 2.0 should now exist under destination.
            dest_w2 = conn.execute(
                """
                SELECT pieces, total_cost
                FROM postage_data
                WHERE file_date='2026-02-10' AND account_code=200 AND mail_class='C' AND weight_oz=2.0
                """
            ).fetchone()
            assert dest_w2 is not None
            assert int(dest_w2["pieces"]) == 4
            assert float(dest_w2["total_cost"]) == pytest.approx(2.0, rel=1e-9)

            # Audit tables should be populated.
            edits = conn.execute("SELECT * FROM postage_edits").fetchall()
            assert len(edits) == 1
            assert edits[0]["from_account_code"] == 100
            assert edits[0]["to_account_code"] == 200
            assert edits[0]["mail_class"] == "C"
            assert edits[0]["reason"] == "Fix wrong account"

            lines = conn.execute(
                "SELECT action, weight_oz, old_account_code, new_account_code FROM postage_edit_lines ORDER BY weight_oz"
            ).fetchall()
            assert len(lines) == 2
            assert {ln["action"] for ln in lines} == {"merged", "updated"}
            assert all(ln["old_account_code"] == 100 for ln in lines)
            assert all(ln["new_account_code"] == 200 for ln in lines)
        finally:
            conn.close()


def test_billing_row_edit_preview_and_apply_partial_buckets(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "billing_edit.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, NULL, NULL)
                """,
                [(100, "FromAcct"), (200, "ToAcct")],
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 5)"
            )
            # 3 pieces at 1 lb (16 oz), 2 pieces at 2 lb (20 oz) — same day/account/class/zone.
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, zone
                ) VALUES (1, 100, '2/10/2026 10:00', ?, 'Priority', 5.0, '3')
                """,
                [(16.0,), (16.0,), (16.0,), (20.0,), (20.0,)],
            )
            conn.commit()

            details = dbmod.get_billing_row_details(
                conn,
                bill_date="2026-02-10",
                account_code=100,
                mail_class="Priority",
                zone="3",
            )
            assert len(details) == 2
            by_lb = {int(d["lb_bucket"]): int(d["pieces"]) for d in details}
            assert by_lb[1] == 3
            assert by_lb[2] == 2

            pieces_by_bucket = {"1": 2, "2": 2}

            prev = dbmod.preview_billing_row_update(
                conn,
                bill_date="2026-02-10",
                from_account_code=100,
                mail_class="Priority",
                zone="3",
                to_account_code=200,
                pieces_by_bucket=pieces_by_bucket,
            )
            assert prev["ok"] is True
            assert prev["summary"]["source_rows"] == 5
            assert prev["summary"]["updated"] == 4

            with conn:
                out = dbmod.apply_billing_row_update(
                    conn,
                    bill_date="2026-02-10",
                    from_account_code=100,
                    mail_class="Priority",
                    zone="3",
                    to_account_code=200,
                    pieces_by_bucket=pieces_by_bucket,
                    reason="Wrong account on parcel billing",
                )
            assert out["ok"] is True
            assert out["summary"]["updated"] == 4

            on_200 = conn.execute(
                """
                SELECT COUNT(*) AS n FROM billing_records
                WHERE custom_account_code = 200 AND usps_mail_class = 'Priority'
                """
            ).fetchone()
            assert int(on_200["n"]) == 4

            on_100 = conn.execute(
                """
                SELECT COUNT(*) AS n FROM billing_records
                WHERE custom_account_code = 100 AND usps_mail_class = 'Priority'
                """
            ).fetchone()
            assert int(on_100["n"]) == 1

            edits = conn.execute("SELECT * FROM billing_edits").fetchall()
            assert len(edits) == 1
            assert edits[0]["from_account_code"] == 100
            assert edits[0]["to_account_code"] == 200
            assert edits[0]["mail_class"] == "Priority"
            assert edits[0]["zone"] == "3"
            assert edits[0]["reason"] == "Wrong account on parcel billing"
            assert int(edits[0]["updated_rows"]) == 4

            lines = conn.execute(
                "SELECT lb_bucket, old_account_code, new_account_code FROM billing_edit_lines ORDER BY id"
            ).fetchall()
            assert len(lines) == 4
            assert all(ln["old_account_code"] == 100 for ln in lines)
            assert all(ln["new_account_code"] == 200 for ln in lines)
            lb_counts = {}
            for ln in lines:
                lb = int(ln["lb_bucket"])
                lb_counts[lb] = lb_counts.get(lb, 0) + 1
            assert lb_counts[1] == 2
            assert lb_counts[2] == 2
        finally:
            conn.close()


def test_query_parcel_over_10lb_lines_filters_and_base(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "o10lb.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            _seed_retail_matrix_rates(
                conn,
                table="priority_mail_retail",
                rows=[
                    (2, 13, 19.2),
                    (3, 16, 22.05),
                ],
            )
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (700, 'Co', NULL, NULL)
                """
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 3)"
            )
            conn.executemany(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, zone
                ) VALUES (1, 700, 'Co', '4/1/2026 10:00', ?, 'CLS', 1.0, ?, ?)
                """,
                [
                    (16.0, 2.5, "2"),
                    (200.0, 19.2, "2"),
                    (256.0, 30.0, "3"),
                ],
            )
            conn.commit()
            rows = dbmod.query_parcel_over_10lb_lines(
                conn, "2026-04-01", "2026-04-30", None, None, True, True
            )
        finally:
            conn.close()

    assert len(rows) == 2
    by_lbs = {r["lbs"]: r for r in rows}
    assert by_lbs[13]["zone"] == 2
    assert by_lbs[13]["customer_number"] == 700
    assert by_lbs[13]["child_name"] == "Co"
    assert by_lbs[13]["base"] == pytest.approx(19.2, rel=1e-9)
    assert by_lbs[16]["zone"] == 3
    assert by_lbs[16]["base"] == pytest.approx(22.05, rel=1e-9)


def test_presort_reject_unit_cost_and_invoice_reject_count(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "rej.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            assert dbmod.get_presort_reject_unit_cost(conn) == pytest.approx(0.66, rel=1e-9)
            dbmod.set_presort_reject_unit_cost(conn, 0.55)
            assert dbmod.get_presort_reject_unit_cost(conn) == pytest.approx(0.55, rel=1e-9)

            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (100, "Parent Co", None, None),
                    (101, "Child A", 100, "Parent Co"),
                ],
            )
            conn.execute(
                """
                INSERT INTO ws3_parent_daily_rejects (mail_date, parent_customer_number, reject_count)
                VALUES ('2026-04-01', 100, 4), ('2026-04-02', 100, 6)
                """
            )
            conn.execute(
                """
                INSERT INTO postage_presort_rejects (file_date, account_code, reject_count, source)
                VALUES ('2026-04-01', 101, 3, 'bm_uplift_1120_othercls_backfill')
                """
            )
            conn.commit()

            n = dbmod.query_ws3_presort_reject_count_for_invoice(
                conn, "2026-04-01", "2026-04-30", 100, None, True, True
            )
            assert n == 10

            pn = dbmod.query_postage_presort_reject_count_for_invoice(
                conn, "2026-04-01", "2026-04-30", 100, None, True, True
            )
            assert pn == 3

            combined = dbmod.query_total_presort_reject_count_for_invoice(
                conn, "2026-04-01", "2026-04-30", 100, None, True, True
            )
            assert combined == 13

            n_child = dbmod.query_ws3_presort_reject_count_for_invoice(
                conn, "2026-04-01", "2026-04-30", 100, 101, True, True
            )
            assert n_child == 10

            combined_child = dbmod.query_total_presort_reject_count_for_invoice(
                conn, "2026-04-01", "2026-04-30", 100, 101, True, True
            )
            assert combined_child == 13
        finally:
            conn.close()


def test_clamp_negative_ws3_reject_counts(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "clamp.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()

        conn = dbmod.get_connection()
        try:
            # Minimal WS3 run + detail with negative pcs_rejected
            conn.execute(
                "INSERT INTO ws3_netsort_customers (customer_code, customer_name) VALUES ('1','A')"
            )
            cur = conn.execute(
                "INSERT INTO ws3_mail_runs (mail_date, mail_id) VALUES ('2026-04-01', '')"
            )
            run_id = int(cur.lastrowid)
            cur = conn.execute("INSERT INTO ws3_profiles (profile_name) VALUES ('P1')")
            profile_id = int(cur.lastrowid)
            conn.execute(
                """
                INSERT INTO ws3_mail_detail (run_id, profile_id, customer_code, rate_type, pcs_rejected)
                VALUES (?, ?, '1', 'ADC Auto', -5)
                """,
                (run_id, profile_id),
            )
            # ws3_parent_daily_rejects has FK to customers, so insert a customer row.
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (1, 'A')"
            )
            conn.execute(
                """
                INSERT INTO ws3_parent_daily_rejects (mail_date, parent_customer_number, reject_count)
                VALUES ('2026-04-01', 1, -2)
                """
            )
            conn.commit()

            dbmod.clamp_negative_ws3_reject_counts(conn)
            conn.commit()

            d = conn.execute("SELECT pcs_rejected FROM ws3_mail_detail").fetchone()[0]
            pval = conn.execute(
                "SELECT reject_count FROM ws3_parent_daily_rejects"
            ).fetchone()[0]
            assert int(d) == 0
            assert int(pval) == 0
        finally:
            conn.close()


def test_backfill_postage_uplift_othercls_1120_moves_to_presort_rejects(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "uplift.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute("INSERT INTO customers (customer_number, customer_name) VALUES (200, 'P')")
            conn.execute("INSERT INTO customers (customer_number, customer_name, parent_number, parent_name) VALUES (201, 'C', 200, 'P')")
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('x.csv', '2026-04-23', 1)"
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('y.csv', '2026-04-24', 1)"
            )
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (?, ?, 201, 'OtherCls', 1120.0, ?, 0.66, 0)
                """,
                [
                    (1, "2026-04-23", 1),
                    (2, "2026-04-24", 2),
                ],
            )
            conn.commit()

            out = dbmod.backfill_postage_uplift_othercls_1120_as_presort_rejects(
                conn, start_date="2026-04-20", end_date="2026-04-25"
            )
            conn.commit()

            assert out["postage_rows_deleted"] == 2
            assert out["reject_pieces_moved"] == 3
            n_postage = conn.execute(
                "SELECT COUNT(*) FROM postage_data WHERE weight_oz=1120.0 AND UPPER(mail_class)='OTHERCLS'"
            ).fetchone()[0]
            assert int(n_postage) == 0
            n_rej = conn.execute(
                "SELECT COALESCE(SUM(reject_count),0) FROM postage_presort_rejects"
            ).fetchone()[0]
            assert int(n_rej) == 3
        finally:
            conn.close()


def test_flats_over_13_oz_excluded_from_postage_and_merged_into_parcels(monkeypatch):
    """Flat-class >13 oz leaves query_postage totals; appears on parcels with real mail_class."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "flat13.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            dbmod.upsert_flat_retail_rates(conn, dbmod.DEFAULT_FLATS_RETAIL_RATES)
            _seed_retail_matrix_rates(
                conn,
                table="priority_mail_retail",
                rows=[(2, 1, 8.5)],
            )
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (100, 'Parent Co', NULL, NULL), (101, 'Child A', 100, 'Parent Co')
                """
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-05-01', 2)"
            )
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-05-01', 101, '1ClFlat', ?, ?, ?, 0)
                """,
                [
                    (13.0, 1, 5.04),
                    (14.0, 3, 30.0),
                ],
            )
            conn.execute(
                "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES ('b1', 'b.csv', 1)"
            )
            conn.execute(
                """
                INSERT INTO billing_records (
                    billing_import_id, custom_account_code, account_name, time_stamp,
                    weight_oz, usps_mail_class, billing_amount, fully_paid_postage, zone
                ) VALUES (1, 101, 'Child A', '5/1/2026 9:00', 14.0, 'USPS GROUND ADVANTAGE', 8.0, 8.5, '2')
                """
            )
            conn.commit()

            post = dbmod.query_postage(
                conn,
                "2026-05-01",
                "2026-05-31",
                100,
                None,
                True,
                True,
                False,
                False,
                False,
            )
            flat_row = next(r for r in post["rows"] if r.get("mail_class") == "1ClFlat")
            assert flat_row["oz_13"] == 1
            assert flat_row["oz_13plus"] == 0
            assert flat_row["total_qty"] == 1
            assert flat_row["total_cost"] == pytest.approx(5.04, rel=1e-9)
            assert flat_row["metered_cost"] == pytest.approx(5.04, rel=1e-9)
            assert flat_row["retail_cost"] == pytest.approx(5.04, rel=1e-9)

            par = dbmod.query_parcels(
                conn,
                "2026-05-01",
                "2026-05-31",
                100,
                None,
                True,
                True,
                False,
                False,
                False,
            )
            syn = next(r for r in par["rows"] if r.get("mail_class") == "1ClFlat")
            assert syn["zone"] == "2"
            assert syn["lb_1"] == 3
            assert syn["total_qty"] == 3
            assert syn["total_retail"] == pytest.approx(3 * 8.5, rel=1e-9)

            bill = next(
                r
                for r in par["rows"]
                if r.get("mail_class") == "USPS GROUND ADVANTAGE"
            )
            assert bill["lb_1"] == 1
            assert par["total_pieces"] == 4
        finally:
            conn.close()


def test_parcel_zone_summary_includes_postage_flat_over_13_zone_2(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "flat13z.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            _seed_retail_matrix_rates(
                conn,
                table="priority_mail_retail",
                rows=[(2, 1, 10.0)],
            )
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Solo')"
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-06-01', 1)"
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-06-01', 200, '1ClFlat', 14.0, 2, 20.0, 0)
                """
            )
            conn.commit()
            s = dbmod.query_parcel_zone_summary(
                conn, "2026-06-01", "2026-06-30", None, None, True, True, False
            )
        finally:
            conn.close()

    assert s["total_pieces"] == 2
    assert s["total_cost"] == pytest.approx((10.0 - 0.25) * 2, rel=1e-9)
    assert s["total_savings"] == pytest.approx(0.5, rel=1e-9)
    block_2_4 = next(b for b in s["blocks"] if b["zone_a"] == 2 and b["zone_b"] == 4)
    r0 = block_2_4["rows"][0]
    assert r0["weight_label"] == "1 lb"
    assert r0["zone_a"]["count"] == 2
    assert r0["zone_a"]["priority"] == 10.0
    assert r0["costs"] == pytest.approx(19.5, rel=1e-9)
    assert r0["savings"] == 0.5


def test_postage_over_13_othercls_merged_to_parcels_not_in_postage(monkeypatch):
    """Non-flat BM-style class >13 oz: not in query_postage totals; appears on parcels."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "othercls13.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            _seed_retail_matrix_rates(
                conn, table="priority_mail_retail", rows=[(2, 1, 7.0)]
            )
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (400, 'Acme')"
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('bm.csv', '2026-08-01', 1)"
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-08-01', 400, 'OtherCls', 14.0, 2, 14.0, 0)
                """
            )
            conn.commit()
            post = dbmod.query_postage(
                conn,
                "2026-08-01",
                "2026-08-31",
                None,
                None,
                True,
                True,
                False,
                False,
                False,
            )
            oc = next(r for r in post["rows"] if r.get("mail_class") == "OtherCls")
            assert oc["total_qty"] == 0
            assert oc["oz_13plus"] == 0

            par = dbmod.query_parcels(
                conn,
                "2026-08-01",
                "2026-08-31",
                None,
                None,
                True,
                True,
                False,
                False,
                False,
            )
            syn = next(r for r in par["rows"] if r.get("mail_class") == "OtherCls")
            assert syn["lb_1"] == 2
            assert syn["total_qty"] == 2

            rpt = dbmod.query_parcel_report_rows(
                conn,
                "2026-08-01",
                "2026-08-31",
                None,
                None,
                True,
                True,
            )
            post_rows = [x for x in rpt if str(x.get("piece_id") or "").startswith("POSTAGE:")]
            assert len(post_rows) == 2
        finally:
            conn.close()


def test_kc_presort_keeps_othercls_over_13_in_postage_not_parcels(monkeypatch):
    """KC presort: do not shift OtherCls/NOCLASS >13 oz into parcels; keep in metered/postage."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "kc_othercls13.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (400, 'Acme')"
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('bm.csv', '2026-08-01', 1)"
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-08-01', 400, 'OtherCls', 14.0, 2, 14.0, 0)
                """
            )
            conn.commit()

            post = dbmod.query_postage(
                conn,
                "2026-08-01",
                "2026-08-31",
                None,
                None,
                True,
                True,
                False,
                False,
                False,
                kc_presort=True,
            )
            oc = next(r for r in post["rows"] if r.get("mail_class") == "OtherCls")
            assert oc["total_qty"] == 2
            assert oc["oz_13plus"] == 2

            par = dbmod.query_parcels(
                conn,
                "2026-08-01",
                "2026-08-31",
                None,
                None,
                True,
                True,
                False,
                False,
                False,
                kc_presort=True,
            )
            assert not any(r.get("mail_class") == "OtherCls" for r in (par.get("rows") or []))
        finally:
            conn.close()


def test_postage_invoice_other_excludes_flat_over_13(monkeypatch):
    """Letter / other bucket must not count flat mail > 13 oz (moved to parcel logic)."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "inv13.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (300, 'InvCo')"
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-07-01', 1)"
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-07-01', 300, '1ClFlat', 14.0, 5, 50.0, 0)
                """
            )
            conn.commit()
        finally:
            conn.close()

        out = exports_mod.export_postage_invoice(
            parent_number=300,
            start_date="2026-07-01",
            end_date="2026-07-31",
            discount=0.1,
            customer_number=None,
            show_parents=True,
            show_main=True,
        )
        try:
            from openpyxl import load_workbook

            wb = load_workbook(out)
            ws = wb.active
            assert ws.cell(28, 9).value == 0
        finally:
            out.unlink(missing_ok=True)


def test_query_postage_reject_flats_use_retail_efd_class_uses_imported_cost(monkeypatch):
    """Non–1CA5DFlt invoice flats 1–13 oz: total_cost from retail tiers; 1CA5DFlt keeps imported total_cost."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "reject_flat_retail.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            dbmod.upsert_flat_retail_rates(conn, dbmod.DEFAULT_FLATS_RETAIL_RATES)
            conn.execute(
                "INSERT INTO customers (customer_number, customer_name) VALUES (200, 'Solo')"
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('x.csv', '2026-03-01', 1)"
            )
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-03-01', 200, ?, 2.0, 5, 1.0, 0)
                """,
                [("1ClFlat",), ("1CSPiece",)],
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-03-01', 200, '1CA5DFlt', 2.0, 2, 99.0, 0)
                """
            )
            conn.commit()
            post = dbmod.query_postage(
                conn,
                "2026-03-01",
                "2026-03-31",
                200,
                None,
                True,
                True,
                False,
                False,
                False,
            )
            post_hidden = dbmod.query_postage(
                conn,
                "2026-03-01",
                "2026-03-31",
                200,
                None,
                True,
                True,
                False,
                False,
                True,
            )
        finally:
            conn.close()

    assert "total_cost" not in post_hidden
    assert "total_metered_cost" not in post_hidden
    assert "total_retail_cost" not in post_hidden
    assert all("metered_cost" not in r and "retail_cost" not in r for r in post_hidden["rows"])

    by_class = {r.get("mail_class"): r for r in post["rows"]}
    r2 = 1.90  # 2 oz retail tier from DEFAULT_FLATS_RETAIL_RATES
    for cls in ("1ClFlat", "1CSPiece"):
        assert by_class[cls]["total_cost"] == pytest.approx(5 * r2, rel=1e-9)
        assert by_class[cls]["metered_cost"] == pytest.approx(5 * r2, rel=1e-9)
        assert by_class[cls]["retail_cost"] == pytest.approx(5 * r2, rel=1e-9)
    assert by_class["1CA5DFlt"]["total_cost"] == pytest.approx(99.0, rel=1e-9)
    assert by_class["1CA5DFlt"]["metered_cost"] == pytest.approx(99.0, rel=1e-9)
    assert by_class["1CA5DFlt"]["retail_cost"] == pytest.approx(2 * r2, rel=1e-9)
    assert post["total_metered_cost"] == pytest.approx(5 * r2 + 5 * r2 + 99.0, rel=1e-9)
    assert post["total_cost"] == post["total_metered_cost"]
    assert post["total_retail_cost"] == pytest.approx(5 * r2 + 5 * r2 + 2 * r2, rel=1e-9)


def test_query_postage_allocate_presort_rejects_by_day(monkeypatch):
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "alloc.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            conn.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (100, 'Parent', NULL, NULL),
                       (101, 'Child A', 100, 'Parent'),
                       (102, 'Child B', 100, 'Parent')
                """
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-04-01', 4)"
            )
            # Child A: 100 @ 5oz, 200 @ 6oz
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-04-01', ?, '1CA5DFlt', ?, ?, 0.0, 0)
                """,
                [
                    (101, 5.0, 100),
                    (101, 6.0, 200),
                ],
            )
            # Child B: 100 @ 6oz
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-04-01', 102, '1CA5DFlt', 6.0, 100, 0.0, 0)
                """
            )
            # 4 presort rejects for the parent on that day
            conn.execute(
                """
                INSERT INTO ws3_parent_daily_rejects (mail_date, parent_customer_number, reject_count)
                VALUES ('2026-04-01', 100, 4)
                """
            )
            conn.commit()

            base = dbmod.query_postage(
                conn,
                "2026-04-01",
                "2026-04-01",
                100,
                None,
                True,
                True,
                False,
                False,
                True,
            )
            assert any(r.get("mail_class") == "Presort rejects" for r in base["rows"])

            alloc = dbmod.query_postage(
                conn,
                "2026-04-01",
                "2026-04-01",
                100,
                None,
                True,
                True,
                False,
                False,
                True,
                allocate_presort_rejects=True,
            )
        finally:
            conn.close()

    rows = alloc["rows"]
    assert not any(r.get("mail_class") == "Presort rejects" for r in rows)
    by_key = {(r["child_number"], r["mail_class"]): r for r in rows}

    # Allocation is proportional across all `1CA5DFlt` cells for the day:
    # Total weights: ChildA 5oz=100, ChildA 6oz=200, ChildB 6oz=100 => 400 total.
    # Rejects=4 => exact allocation: ChildA 5oz=1, ChildA 6oz=2, ChildB 6oz=1.
    a_efd = by_key[(101, "1CA5DFlt")]
    b_efd = by_key[(102, "1CA5DFlt")]
    a_rej = by_key[(101, "Rejects")]
    b_rej = by_key[(102, "Rejects")]

    assert int(a_rej["oz_5"]) == 1
    assert int(a_rej["oz_6"]) == 2
    assert int(a_rej["total_qty"]) == 3
    assert int(b_rej["oz_6"]) == 1
    assert int(b_rej["total_qty"]) == 1
    assert int(a_efd["oz_5"]) + int(a_efd["oz_6"]) == int(a_efd["total_qty"]) == 297
    assert int(b_efd["oz_6"]) == int(b_efd["total_qty"]) == 99


def test_postage_invoice_efd_1ca5dflt_in_i_reject_flats_in_k_and_cost_center(monkeypatch):
    """Column I = max(0, EFD − WS3); K = IMB + WS3; J = G; N = I×(G−H); CC savings = (C−D)×($G$14−$H$14)."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "efd_reject_inv.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            dbmod.upsert_flat_retail_rates(conn, dbmod.DEFAULT_FLATS_RETAIL_RATES)
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (100, "Parent Co", None, None),
                    (101, "Child A", 100, "Parent Co"),
                ],
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-08-01', 1)"
            )
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-08-01', 101, ?, 2.0, ?, ?, 0)
                """,
                [
                    ("1CA5DFlt", 3, 5.0),
                    ("1ClFlat", 5, 1.0),
                ],
            )
            conn.commit()
        finally:
            conn.close()

        out = exports_mod.export_postage_invoice(
            parent_number=100,
            start_date="2026-08-01",
            end_date="2026-08-31",
            discount=0.1,
            customer_number=None,
            show_parents=True,
            show_main=True,
        )
        try:
            from openpyxl import load_workbook

            wb = load_workbook(out)
            ws = wb.active
            r_2oz = 15
            assert ws.cell(r_2oz, 9).value == 3
            assert ws.cell(r_2oz, 11).value == 5
            row_child = None
            for r in range(35, 50):
                if ws.cell(r, 1).value == 101:
                    row_child = r
                    break
            assert row_child is not None
            efd_rate = round(max(0.0, 1.90 - 0.1), 4)
            expect_charge = round(3 * efd_rate + 5 * 1.90, 2)
            assert int(ws.cell(row_child, 4).value) == 5
            assert float(ws.cell(row_child, 5).value) == pytest.approx(
                expect_charge, rel=1e-9
            )
            assert ws.cell(r_2oz, 14).value == f"=I{r_2oz}*(G{r_2oz}-H{r_2oz})"
            assert (
                ws.cell(row_child, 6).value
                == f"=(C{row_child}-D{row_child})*($G$14-$H$14)"
            )
        finally:
            out.unlink(missing_ok=True)


def test_postage_invoice_ws3_distributed_per_oz_and_cost_centers(monkeypatch):
    """WS3 rejects allocated by-day across 1CA5DFlt; J = G; N = I×(G−H); CC savings use (C−D)×($G$14−$H$14)."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "ws3_inv.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            dbmod.upsert_flat_retail_rates(conn, dbmod.DEFAULT_FLATS_RETAIL_RATES)
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (600, "Parent Six", None, None),
                    (601, "Dept Low", 600, "Parent Six"),
                    (602, "Dept High", 600, "Parent Six"),
                ],
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-09-01', 1)"
            )
            conn.executemany(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-09-01', ?, ?, 2.0, ?, ?, 0)
                """,
                [
                    (601, "1ClFlat", 10, 19.0),
                    (602, "1CA5DFlt", 90, 162.0),
                ],
            )
            conn.execute(
                """
                INSERT INTO ws3_parent_daily_rejects (mail_date, parent_customer_number, reject_count)
                VALUES ('2026-09-01', 600, 50)
                """
            )
            conn.commit()
        finally:
            conn.close()

        out = exports_mod.export_postage_invoice(
            parent_number=600,
            start_date="2026-09-01",
            end_date="2026-09-30",
            discount=0.1,
            customer_number=None,
            show_parents=True,
            show_main=True,
        )
        try:
            from openpyxl import load_workbook

            wb = load_workbook(out)
            ws = wb.active
            r_2oz = 15
            assert ws.cell(r_2oz, 9).value == 40
            assert ws.cell(r_2oz, 11).value == 60
            retail_2 = 1.90
            assert ws.cell(r_2oz, 10).value == f"=G{r_2oz}"
            assert float(ws.cell(r_2oz, 7).value) == pytest.approx(retail_2, rel=1e-9)
            assert ws.cell(r_2oz, 13).value == f"=G{r_2oz}*(I{r_2oz}+K{r_2oz})"
            assert ws.cell(r_2oz, 14).value == f"=I{r_2oz}*(G{r_2oz}-H{r_2oz})"
            # No separate presort rejects line item; presort rejects are allocated into the weight grid.
            assert ws.cell(27, 6).value == "Letter"
            assert "Presort rejects" not in [ws.cell(r, 6).value for r in range(13, 45)]

            names = [ws.cell(r, 2).value for r in range(34, 45)]
            assert "Presort rejects" not in names

            row_601 = row_602 = None
            for r in range(34, 50):
                if ws.cell(r, 1).value == 601:
                    row_601 = r
                if ws.cell(r, 1).value == 602:
                    row_602 = r
            assert row_601 is not None and row_602 is not None
            assert int(ws.cell(row_601, 3).value) == 10
            assert int(ws.cell(row_602, 3).value) == 90
            efd_rate = round(max(0.0, retail_2 - 0.1), 4)
            # Dept 601 has only non-EFD flats: all 10 pieces stay as IMB rejects; WS3 allocates only to 1CA5DFlt.
            raw_601 = retail_2 * 10.0
            raw_602 = efd_rate * 40.0 + retail_2 * 50.0
            expect_601 = round(raw_601, 2)
            expect_602 = round(raw_602, 2)
            assert int(ws.cell(row_601, 4).value) == 10
            assert int(ws.cell(row_602, 4).value) == 50
            assert float(ws.cell(row_601, 5).value) == pytest.approx(expect_601, rel=1e-9)
            assert float(ws.cell(row_602, 5).value) == pytest.approx(expect_602, rel=1e-9)
            f601 = f"=(C{row_601}-D{row_601})*($G$14-$H$14)"
            f602 = f"=(C{row_602}-D{row_602})*($G$14-$H$14)"
            assert ws.cell(row_601, 6).value == f601
            assert ws.cell(row_602, 6).value == f602
        finally:
            out.unlink(missing_ok=True)


def test_postage_invoice_postage_derived_presort_rejects_allocated_into_weight_grid(monkeypatch):
    """Postage-derived presort rejects (postage_presort_rejects) are allocated into weight K (no separate line item)."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "post_presort_alloc.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            dbmod.upsert_flat_retail_rates(conn, dbmod.DEFAULT_FLATS_RETAIL_RATES)
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (700, "Parent Seven", None, None),
                    (701, "Dept One", 700, "Parent Seven"),
                ],
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-10-01', 1)"
            )
            # Only 2oz EFD flats volume so all allocated rejects should land on the 2oz row.
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-10-01', 701, '1ClFlat', 2.0, 100, 190.0, 0)
                """
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-10-01', 701, '1CA5DFlt', 2.0, 100, 0.0, 0)
                """
            )
            # Postage-derived presort rejects (no WS3 rows).
            conn.execute(
                """
                INSERT INTO postage_presort_rejects (file_date, account_code, reject_count, source, import_id)
                VALUES ('2026-10-01', 701, 12, 'test', 1)
                """
            )
            conn.commit()
        finally:
            conn.close()

        out = exports_mod.export_postage_invoice(
            parent_number=700,
            start_date="2026-10-01",
            end_date="2026-10-31",
            discount=0.1,
            customer_number=None,
            show_parents=True,
            show_main=True,
        )
        try:
            from openpyxl import load_workbook

            wb = load_workbook(out)
            ws = wb.active
            r_2oz = 15
            # K includes both postage_data non-EFD flats (IMB bucket) and allocated presort rejects.
            assert ws.cell(r_2oz, 11).value == 112
            assert ws.cell(27, 6).value == "Letter"
            assert "Presort rejects" not in [ws.cell(r, 6).value for r in range(13, 45)]

            row_701 = None
            for r in range(34, 60):
                if ws.cell(r, 1).value == 701:
                    row_701 = r
                    break
            assert row_701 is not None
            assert int(ws.cell(row_701, 4).value) == 112
        finally:
            out.unlink(missing_ok=True)


def test_postage_invoice_remove_zeros_filters_cost_centers(monkeypatch):
    """remove_zeros=True omits all-zero cost center rows; False keeps every child."""
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "cc_remove_zeros.db"
        monkeypatch.setattr(dbmod, "DB_PATH", p)
        dbmod.init_db()
        conn = dbmod.get_connection()
        try:
            dbmod.upsert_flat_retail_rates(conn, dbmod.DEFAULT_FLATS_RETAIL_RATES)
            conn.executemany(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (800, "Parent Eight", None, None),
                    (801, "Child Active", 800, "Parent Eight"),
                    (802, "Child Empty", 800, "Parent Eight"),
                ],
            )
            conn.execute(
                "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES ('p.csv', '2026-11-01', 1)"
            )
            conn.execute(
                """
                INSERT INTO postage_data (
                    import_id, file_date, account_code, mail_class,
                    weight_oz, pieces, total_cost, unmatched_account
                ) VALUES (1, '2026-11-01', 801, '1CA5DFlt', 2.0, 5, 8.50, 0)
                """
            )
            conn.commit()
        finally:
            conn.close()

        def cost_center_numbers(remove_zeros: bool) -> set[int]:
            out = exports_mod.export_postage_invoice(
                parent_number=800,
                start_date="2026-11-01",
                end_date="2026-11-30",
                discount=0.1,
                customer_number=None,
                show_parents=True,
                show_main=True,
                remove_zeros=remove_zeros,
            )
            try:
                from openpyxl import load_workbook

                wb = load_workbook(out)
                ws = wb.active
                found: set[int] = set()
                for r in range(34, 80):
                    v = ws.cell(r, 1).value
                    if v in (None, "Totals", "—"):
                        continue
                    try:
                        found.add(int(v))
                    except (TypeError, ValueError):
                        pass
                return found
            finally:
                out.unlink(missing_ok=True)

        assert cost_center_numbers(remove_zeros=True) == {801}
        assert cost_center_numbers(remove_zeros=False) == {800, 801, 802}
