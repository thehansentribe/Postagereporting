#!/usr/bin/env python
"""Back-test harness for report regressions.

Snapshots the bytes/values of every report endpoint against the live postage.db,
then re-runs the same requests after code changes and diffs the results.

    python scripts/backtest_reports.py --capture
    python scripts/backtest_reports.py --compare [--allow REGEX ...]

All pricing knobs are passed explicitly at their legacy values, so stored-terms
defaults never influence these snapshots: explicit request params must always
win, and reports for pre-rate-case dates must stay byte-value identical unless
a diff is explicitly allowed with --allow.

xlsx files are compared cell-by-cell (values and formulas) via openpyxl, so
zip/timestamp noise inside the container is ignored. JSON and CSV are compared
on parsed values.
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

DEFAULT_DIR = ROOT / "reports" / "backtest-baseline"
MAX_DIFFS_PER_FILE = 40

# Explicit legacy knob values — see module docstring.
KNOBS = {
    "discount": "0.10",
    "discount_efd": "0.23",
    "parcel_fee": "1.25",
    "efd_parcel_fee": "1.25",
    "parcel_discount": "0.25",
}

WEEK_PRE = {"start_date": "2026-06-08", "end_date": "2026-06-12"}
SPAN_RATECASE = {"start_date": "2026-07-01", "end_date": "2026-07-16"}
MAY = {"start_date": "2026-05-01", "end_date": "2026-05-31"}
BCBS = {"parent_number": "3901", "profit_accounts": "3901"}


def _requests() -> list[dict]:
    reqs: list[dict] = []

    def add(scenario: str, slug: str, path: str, dates: dict, extra: dict | None = None) -> None:
        params = dict(KNOBS)
        params.update(dates)
        if extra:
            params.update(extra)
        reqs.append({"id": f"{scenario}__{slug}", "path": path, "params": params})

    s1 = "week_pre_ratecase"
    for slug, path in [
        ("profit_flats", "/api/profit/flats"),
        ("profit_parcels", "/api/profit/parcels"),
        ("profit_report", "/api/export/profit-report-xlsx"),
        ("efd_parcel_invoice", "/api/export/efd-parcel-invoice-xlsx"),
        ("efd_weekly_invoice", "/api/export/efd-weekly-invoice-xlsx"),
        ("flats_grid_xlsx", "/api/export/flats-grid-xlsx"),
        ("flats_grid_csv", "/api/export/flats-grid-csv"),
        ("parcel_report", "/api/export/parcel-report"),
        ("parcel_counts", "/api/export/parcel-counts-xlsx"),
        ("consolidated_parcel_csv", "/api/export/consolidated-parcel-csv"),
        ("consolidated_volumes", "/api/export/consolidated-volumes-xlsx"),
        ("efd_daily_volumes", "/api/export/efd-daily-volumes-xlsx"),
        ("parcel_zone_summary_xlsx", "/api/export/parcel-zone-summary"),
        ("parcels_json", "/api/parcels"),
        ("parcels_zone_summary_json", "/api/parcels/zone-summary"),
    ]:
        add(s1, slug, path, WEEK_PRE)
    add(s1, "postage_invoice_3901", "/api/export/postage-invoice", WEEK_PRE, {"parent_number": "3901"})
    add(s1, "ga_zone_pricing_csv", "/api/export/ground-advantage-zone-pricing-csv", {})

    s2 = "span_ratecase"
    for slug, path in [
        ("profit_flats", "/api/profit/flats"),
        ("profit_parcels", "/api/profit/parcels"),
        ("profit_report", "/api/export/profit-report-xlsx"),
        ("flats_grid_xlsx", "/api/export/flats-grid-xlsx"),
    ]:
        add(s2, slug, path, SPAN_RATECASE)
    add(s2, "postage_invoice_3901", "/api/export/postage-invoice", SPAN_RATECASE, {"parent_number": "3901"})

    s3 = "may_parcels"
    for slug, path in [
        ("profit_parcels", "/api/profit/parcels"),
        ("efd_parcel_invoice", "/api/export/efd-parcel-invoice-xlsx"),
        ("consolidated_parcel_csv", "/api/export/consolidated-parcel-csv"),
        ("parcel_zone_summary_xlsx", "/api/export/parcel-zone-summary"),
        ("parcels_zone_summary_json", "/api/parcels/zone-summary"),
    ]:
        add(s3, slug, path, MAY)

    s4 = "bcbs3901_scoped"
    for slug, path in [
        ("profit_flats", "/api/profit/flats"),
        ("profit_parcels", "/api/profit/parcels"),
        ("profit_report", "/api/export/profit-report-xlsx"),
        ("efd_parcel_invoice", "/api/export/efd-parcel-invoice-xlsx"),
        ("postage_invoice", "/api/export/postage-invoice"),
    ]:
        add(s4, slug, path, WEEK_PRE, BCBS)

    return reqs


def _ext_for(content_type: str) -> str:
    if "spreadsheetml" in content_type:
        return "xlsx"
    if "csv" in content_type:
        return "csv"
    if "json" in content_type:
        return "json"
    return "bin"


def _client():
    import app as app_module

    app_module.app.config["TESTING"] = True
    # Keep the folder watcher and report scheduler off during snapshots so a
    # background import can't mutate the DB between capture and compare.
    app_module._watcher_started = True
    app_module._scheduler_started = True
    return app_module.app.test_client()


def _fetch(client, req: dict) -> tuple[int, str, bytes]:
    resp = client.get(req["path"], query_string=req["params"])
    return resp.status_code, resp.content_type or "", resp.get_data()


def capture(out_dir: Path) -> int:
    out_dir.mkdir(parents=True, exist_ok=True)
    client = _client()
    manifest: dict[str, dict] = {}
    for req in _requests():
        status, ctype, body = _fetch(client, req)
        ext = _ext_for(ctype)
        fname = f"{req['id']}.{ext}"
        (out_dir / fname).write_bytes(body)
        manifest[req["id"]] = {
            "path": req["path"],
            "params": req["params"],
            "status": status,
            "content_type": ctype,
            "file": fname,
        }
        print(f"  captured {req['id']} [{status}] -> {fname} ({len(body)} bytes)")
    (out_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, sort_keys=True))
    print(f"\nBaseline: {len(manifest)} snapshots in {out_dir}")
    return 0


def _xlsx_diffs(old: bytes, new: bytes) -> list[str]:
    import openpyxl

    diffs: list[str] = []
    wb_old = openpyxl.load_workbook(io.BytesIO(old), data_only=False)
    wb_new = openpyxl.load_workbook(io.BytesIO(new), data_only=False)
    names_old, names_new = wb_old.sheetnames, wb_new.sheetnames
    if names_old != names_new:
        diffs.append(f"sheets: {names_old} != {names_new}")
    for name in [n for n in names_old if n in names_new]:
        ws_o, ws_n = wb_old[name], wb_new[name]
        rows = max(ws_o.max_row, ws_n.max_row)
        cols = max(ws_o.max_column, ws_n.max_column)
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                v_o = ws_o.cell(row=r, column=c).value
                v_n = ws_n.cell(row=r, column=c).value
                if v_o != v_n:
                    coord = ws_o.cell(row=r, column=c).coordinate
                    diffs.append(f"{name}!{coord}: {v_o!r} -> {v_n!r}")
    return diffs


def _json_diffs(old: bytes, new: bytes) -> list[str]:
    diffs: list[str] = []

    def walk(a, b, path):
        if type(a) is not type(b):
            diffs.append(f"{path}: {a!r} -> {b!r}")
        elif isinstance(a, dict):
            for k in sorted(set(a) | set(b)):
                if k not in a:
                    diffs.append(f"{path}.{k}: <absent> -> {b[k]!r}")
                elif k not in b:
                    diffs.append(f"{path}.{k}: {a[k]!r} -> <absent>")
                else:
                    walk(a[k], b[k], f"{path}.{k}")
        elif isinstance(a, list):
            if len(a) != len(b):
                diffs.append(f"{path}: length {len(a)} -> {len(b)}")
            for i, (x, y) in enumerate(zip(a, b)):
                walk(x, y, f"{path}[{i}]")
        elif a != b:
            diffs.append(f"{path}: {a!r} -> {b!r}")

    try:
        walk(json.loads(old or b"null"), json.loads(new or b"null"), "$")
    except json.JSONDecodeError as e:
        diffs.append(f"json parse error: {e}")
    return diffs


def _csv_diffs(old: bytes, new: bytes) -> list[str]:
    import csv as csv_mod

    rows_o = list(csv_mod.reader(io.StringIO(old.decode("utf-8", "replace"))))
    rows_n = list(csv_mod.reader(io.StringIO(new.decode("utf-8", "replace"))))
    diffs: list[str] = []
    if len(rows_o) != len(rows_n):
        diffs.append(f"row count: {len(rows_o)} -> {len(rows_n)}")
    for i, (ro, rn) in enumerate(zip(rows_o, rows_n)):
        if ro != rn:
            diffs.append(f"row {i + 1}: {ro} -> {rn}")
    return diffs


def compare(base_dir: Path, allow: list[str]) -> int:
    manifest = json.loads((base_dir / "manifest.json").read_text())
    allow_res = [re.compile(p) for p in allow]
    client = _client()
    unexpected_files: list[str] = []
    expected_files: list[str] = []
    identical = 0
    for rid, entry in sorted(manifest.items()):
        status, ctype, body = _fetch(client, {"path": entry["path"], "params": entry["params"]})
        old = (base_dir / entry["file"]).read_bytes()
        diffs: list[str] = []
        if status != entry["status"]:
            diffs.append(f"status: {entry['status']} -> {status}")
        ext = entry["file"].rsplit(".", 1)[-1]
        if status == entry["status"]:
            if ext == "xlsx" and status == 200:
                diffs += _xlsx_diffs(old, body)
            elif ext == "json":
                diffs += _json_diffs(old, body)
            elif ext == "csv":
                diffs += _csv_diffs(old, body)
            elif old != body:
                diffs.append("binary content differs")
        if not diffs:
            identical += 1
            continue
        unexpected = [d for d in diffs if not any(rx.search(f"{rid}:{d}") for rx in allow_res)]
        label = "UNEXPECTED" if unexpected else "expected"
        (unexpected_files if unexpected else expected_files).append(rid)
        print(f"\n[{label}] {rid} — {len(diffs)} diff(s)")
        for d in diffs[:MAX_DIFFS_PER_FILE]:
            allowed = any(rx.search(f"{rid}:{d}") for rx in allow_res)
            print(f"    {d}{' (allowed)' if allowed else ''}")
        if len(diffs) > MAX_DIFFS_PER_FILE:
            print(f"    ... {len(diffs) - MAX_DIFFS_PER_FILE} more")
    print(
        f"\n{identical} identical, {len(expected_files)} with allowed diffs, "
        f"{len(unexpected_files)} with UNEXPECTED diffs"
    )
    if unexpected_files:
        print("UNEXPECTED:", ", ".join(unexpected_files))
        return 1
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    mode = ap.add_mutually_exclusive_group(required=True)
    mode.add_argument("--capture", action="store_true", help="snapshot baseline")
    mode.add_argument("--compare", action="store_true", help="diff current output vs baseline")
    ap.add_argument("--dir", type=Path, default=DEFAULT_DIR, help=f"baseline dir (default {DEFAULT_DIR})")
    ap.add_argument(
        "--allow",
        action="append",
        default=[],
        metavar="REGEX",
        help="diff lines matching 'id:diff' are expected (repeatable)",
    )
    args = ap.parse_args()
    if args.capture:
        return capture(args.dir)
    return compare(args.dir, args.allow)


if __name__ == "__main__":
    raise SystemExit(main())
