"""File import logic: Pitney postage, parcel billing, customers, flat rates."""

from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import db

# Column A: skip title / header rows (case-insensitive prefix).
SKIP_A_PREFIXES = (
    "pitney bowes",
    "dm weight break",
    "account code",
    "custom field",
    "report on working database",
    "business manager",
    "custom",
    "report",
    "business",
)


def _bm_skip_row_a(a: str) -> bool:
    if not a:
        return False
    low = a.lower()
    return any(low.startswith(p) for p in SKIP_A_PREFIXES)


def _bm_is_g_column_label_not_class_code(g: str) -> bool:
    """True when column G is the sheet header word 'Class', not a Pitney mail class."""
    return g.strip().lower() == "class"


def _bm_cell_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    t = str(val).strip().replace(",", "")
    if not t:
        return None
    try:
        return float(t)
    except (ValueError, TypeError):
        return None


def _bm_cell_int(val: Any) -> int:
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    t = str(val).strip().replace(",", "")
    if not t:
        return 0
    try:
        return int(float(t))
    except (ValueError, TypeError):
        return 0


def strip_zeros(value: Any) -> int | None:
    try:
        return int(str(value).strip().lstrip("0") or "0")
    except (ValueError, TypeError):
        return None


def parse_bm_date(filename: str) -> str:
    """
    BM_3_20_26_report.csv / BM_3_20_26.xls -> 2026-03-20
    BM 3.19.26.xls / BM 3_19_26.xls -> 2026-03-19
    """
    base = os.path.basename(filename)
    m = re.search(r"BM[_\s]+(\d+)[_.\s/-]+(\d+)[_.\s/-]+(\d+)", base, re.IGNORECASE)
    if not m:
        raise ValueError(f"Cannot parse date from filename: {filename}")
    month, day, year_2d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    year = 2000 + year_2d if year_2d < 100 else year_2d
    return date(year, month, day).isoformat()


def read_bm_report_date_from_xlsx(xlsx_path: str) -> str:
    """
    Read the Business Manager report date from the sheet itself.

    The BM/DM Weight Break report embeds the date in cells P3 and S3.
    Those cells must match; otherwise the import is rejected.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        p3_raw = ws["P3"].value
        s3_raw = ws["S3"].value
    finally:
        wb.close()

    p3 = _parse_effective_date_any(p3_raw)
    s3 = _parse_effective_date_any(s3_raw)
    base = os.path.basename(xlsx_path)

    if not p3 or not s3:
        raise ValueError(
            f"Cannot parse BM report date from P3/S3 in {base}: P3={p3_raw!r} S3={s3_raw!r}"
        )
    if p3 != s3:
        raise ValueError(f"BM report date mismatch in {base}: P3={p3} S3={s3}")
    return p3


def _find_libreoffice_executable() -> str | None:
    for cmd in ("libreoffice", "soffice"):
        path = shutil.which(cmd)
        if path:
            return path
    pf = os.environ.get("PROGRAMFILES", r"C:\Program Files")
    pf86 = os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)")
    for base in (pf, pf86):
        for exe in (
            os.path.join(base, "LibreOffice", "program", "soffice.exe"),
            os.path.join(base, "LibreOffice 5", "program", "soffice.exe"),
        ):
            if os.path.isfile(exe):
                return exe
    return None


def convert_xls_to_xlsx(xls_path: str, out_dir: str | None = None) -> str:
    lo = _find_libreoffice_executable()
    if not lo:
        raise RuntimeError(
            "LibreOffice required for XLS conversion. Install LibreOffice and ensure "
            "'soffice' or 'libreoffice' is on PATH."
        )
    out_dir = out_dir or tempfile.gettempdir()
    os.makedirs(out_dir, exist_ok=True)
    result = subprocess.run(
        [lo, "--headless", "--convert-to", "xlsx", "--outdir", out_dir, xls_path],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr or result.stdout}")
    # LibreOffice names its output after the input stem with a .xlsx extension. Strip only
    # the trailing extension; a naive ".xls"->".xlsx" replace turns a misnamed ".xlsx" input
    # (OLE2/BIFF bytes with an .xlsx name) into a bogus ".xlsxx" lookup that never exists.
    stem = re.sub(r"\.(xls|xlsx)$", "", os.path.basename(xls_path), flags=re.IGNORECASE)
    xlsx_path = os.path.join(out_dir, f"{stem}.xlsx")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Expected output not found: {xlsx_path}")
    return xlsx_path


# Magic bytes: OOXML (.xlsx) files are ZIP archives; legacy BIFF (.xls) are OLE2.
_OOXML_MAGIC = b"PK\x03\x04"
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _sniff_spreadsheet_magic(path: str) -> str | None:
    """Return 'ooxml', 'ole2', or None based on the file's leading bytes."""
    try:
        with open(path, "rb") as f:
            head = f.read(8)
    except OSError:
        return None
    if head.startswith(_OOXML_MAGIC):
        return "ooxml"
    if head.startswith(_OLE2_MAGIC):
        return "ole2"
    return None


def resolve_xlsx_path(path: str, out_dir: str | None = None) -> str:
    """
    Return a path that openpyxl can open, deciding by file content not extension.

    - OOXML (ZIP) content: already an .xlsx workbook. If the on-disk name does not
      end in .xlsx, copy it to a temp .xlsx (openpyxl rejects any path ending in
      .xls even when the bytes are valid OOXML), otherwise return it unchanged.
    - Legacy BIFF (OLE2) content: convert via LibreOffice.
    - Unknown content: fall back to extension (.xls -> convert, else as-is).
    """
    kind = _sniff_spreadsheet_magic(path)
    low = path.lower()

    if kind == "ooxml":
        if low.endswith(".xlsx"):
            return path
        out_dir = out_dir or tempfile.gettempdir()
        os.makedirs(out_dir, exist_ok=True)
        stem = re.sub(r"\.(xls|xlsx)$", "", os.path.basename(path), flags=re.IGNORECASE)
        dest = os.path.join(out_dir, f"{stem}.xlsx")
        shutil.copyfile(path, dest)
        return dest

    if kind == "ole2":
        return convert_xls_to_xlsx(path, out_dir)

    if low.endswith(".xls") and not low.endswith(".xlsx"):
        return convert_xls_to_xlsx(path, out_dir)
    return path


def parse_bm_xlsx(xlsx_path: str) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_out: list[dict[str, Any]] = []
    current_account: str | None = None
    current_class: str | None = None

    for row in ws.iter_rows(values_only=False):
        cells = {cell.column_letter: cell.value for cell in row if cell.value is not None}

        a = str(cells.get("A", "") or "").strip()
        g = str(cells.get("G", "") or "").strip()
        j = str(cells.get("J", "") or "").strip()
        m = cells.get("M")
        p = cells.get("P")
        r = cells.get("R")

        if _bm_skip_row_a(a):
            continue
        if j == "Sub Total":
            continue
        if a and re.match(r"^\d{4}$", a):
            current_account = a
            current_class = None
            continue
        if g:
            if _bm_is_g_column_label_not_class_code(g):
                continue
            current_class = g
            continue
        if current_account is None:
            continue
        weight = _bm_cell_float(m)
        if weight is None:
            continue
        pieces = _bm_cell_int(p)
        cost = _bm_cell_float(r) or 0.0
        rows_out.append(
            {
                "account_code": current_account,
                "mail_class": current_class or "UNKNOWN",
                "weight_oz": weight,
                "pieces": pieces,
                "total_cost": round(cost, 3),
            }
        )

    wb.close()
    return rows_out


def parse_bm_raw_csv(csv_path: str) -> list[dict[str, Any]]:
    """Parse Pitney Bowes raw BM export CSV (class G, subtotals J, weight M, pieces P, cost R).

    Pitney CSV uses the same layout as XLSX: weight is column M (index 12), not N (often empty).
    Numeric cells may include thousands separators (e.g. \"1,120.0\", \"5,786\").
    """
    rows_out: list[dict[str, Any]] = []
    current_account: str | None = None
    current_class: str | None = None

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            while len(row) < 18:
                row.append("")

            a = row[0].strip()
            g = row[6].strip()
            j = row[9].strip()
            m_raw = row[12]
            p_raw = row[15]
            r_raw = row[17]

            if _bm_skip_row_a(a):
                continue
            if j == "Sub Total":
                continue
            if a and re.match(r"^\d{4}$", a):
                current_account = a
                current_class = None
                continue
            if g:
                if _bm_is_g_column_label_not_class_code(g):
                    continue
                current_class = g
                continue
            if current_account is None:
                continue
            weight = _bm_cell_float(m_raw)
            if weight is None:
                continue
            pieces = _bm_cell_int(p_raw)
            cost = _bm_cell_float(r_raw) or 0.0
            rows_out.append(
                {
                    "account_code": current_account,
                    "mail_class": current_class or "UNKNOWN",
                    "weight_oz": weight,
                    "pieces": pieces,
                    "total_cost": round(cost, 3),
                }
            )

    return rows_out


def write_report_csv(rows: list[dict[str, Any]], source_path: str, out_dir: str) -> str:
    base = os.path.basename(source_path)
    csv_name = re.sub(r"\.(xlsx|xls|csv)$", "_report.csv", base, flags=re.IGNORECASE)
    if not csv_name.endswith("_report.csv"):
        csv_name = base + "_report.csv"
    out_path = os.path.join(out_dir, csv_name)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Account Code", "Class", "Weight  (oz.)", "Pieces", "Total Cost"])
        for row in rows:
            w.writerow(
                [
                    row["account_code"],
                    row["mail_class"],
                    row["weight_oz"],
                    row["pieces"],
                    f"{row['total_cost']:.3f}",
                ]
            )
    return out_path


def import_bm_csv(
    csv_path: str,
    db_path: Path | str,
    *,
    file_date_override: str | None = None,
) -> dict[str, Any]:
    file_name = os.path.basename(csv_path)
    file_date = file_date_override or parse_bm_date(file_name)

    data_rows: list[dict[str, str]] = []
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            data_rows.append(row)

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    conflict = cur.execute(
        "SELECT file_name FROM postage_imports WHERE file_date = ? AND file_name != ?",
        (file_date, file_name),
    ).fetchone()
    if conflict:
        other = conflict[0]
        conn.close()
        raise ValueError(
            f"Postage for {file_date} is already imported from {other}. "
            f"Delete that import from the database or remove the conflicting file before importing {file_name}."
        )

    existing = cur.execute(
        "SELECT id FROM postage_imports WHERE file_name = ?", (file_name,)
    ).fetchone()
    if existing:
        cur.execute("DELETE FROM postage_imports WHERE file_name = ?", (file_name,))

    cur.execute(
        "INSERT INTO postage_imports (file_name, file_date, row_count) VALUES (?, ?, ?)",
        (file_name, file_date, len(data_rows)),
    )
    import_id = cur.lastrowid

    valid_accounts = {r[0] for r in cur.execute("SELECT customer_number FROM customers").fetchall()}

    unmatched: set[int] = set()
    inserted = 0
    diverted_reject_pieces = 0
    for row in data_rows:
        account_code = strip_zeros(row.get("Account Code", ""))
        if account_code is None:
            continue
        mail_class = str(row.get("Class", "")).strip()
        try:
            weight_oz = float(row.get("Weight  (oz.)", 0) or 0)
            pieces = int(row.get("Pieces", 0) or 0)
            total_cost = float(row.get("Total Cost", 0) or 0)
        except (ValueError, TypeError):
            continue

        is_unmatched = 0 if account_code in valid_accounts else 1
        if is_unmatched:
            unmatched.add(account_code)

        # KC presort uplift artifact: OtherCls @ 1120 oz (70 lb) is not real weight; treat as
        # presort rejects and exclude from postage totals/volumes.
        if mail_class.strip().upper() == "OTHERCLS" and float(weight_oz) == 1120.0:
            pcs = int(pieces or 0)
            if pcs > 0:
                cur.execute(
                    """
                    INSERT INTO postage_presort_rejects (file_date, account_code, reject_count, source, import_id)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(file_date, account_code, source) DO UPDATE SET
                        reject_count = postage_presort_rejects.reject_count + excluded.reject_count,
                        import_id = COALESCE(postage_presort_rejects.import_id, excluded.import_id)
                    """,
                    (
                        file_date,
                        account_code,
                        pcs,
                        "bm_uplift_1120_othercls",
                        import_id,
                    ),
                )
                diverted_reject_pieces += pcs
            continue

        try:
            cur.execute(
                """
                INSERT INTO postage_data
                    (import_id, file_date, account_code, mail_class,
                     weight_oz, pieces, total_cost, unmatched_account)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    file_date,
                    account_code,
                    mail_class,
                    weight_oz,
                    pieces,
                    total_cost,
                    is_unmatched,
                ),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    conn.close()

    return {
        "file_name": file_name,
        "file_date": file_date,
        "rows_imported": inserted,
        "diverted_presort_reject_pieces": diverted_reject_pieces,
        "unmatched": sorted(unmatched),
    }


def process_bm_file(
    xls_path: str, db_path: Path | str, csv_out_dir: str | None = None
) -> dict[str, Any]:
    csv_out_dir = csv_out_dir or os.path.dirname(xls_path)
    low = xls_path.lower()
    file_date_override: str | None = None
    if low.endswith(".csv"):
        rows = parse_bm_raw_csv(xls_path)
    else:
        xlsx_path = resolve_xlsx_path(xls_path)
        file_date_override = read_bm_report_date_from_xlsx(xlsx_path)
        rows = parse_bm_xlsx(xlsx_path)
    csv_path = write_report_csv(rows, xls_path, csv_out_dir)
    return import_bm_csv(csv_path, db_path, file_date_override=file_date_override)


def safe_real(value: Any) -> float | None:
    try:
        s = str(value).strip()
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


def import_billing_csv(csv_path: str, db_path: Path | str) -> dict[str, Any]:
    file_name = os.path.basename(csv_path)

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [h.strip() for h in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            rows.append({(k or "").strip(): (v or "").strip() if isinstance(v, str) else v for k, v in row.items()})

    if not rows:
        raise ValueError("CSV file is empty")

    billing_id = str(rows[0].get("BillingID", "") or "").strip()
    if not billing_id:
        raise ValueError("Cannot determine BillingID from file")

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    existing = cur.execute(
        "SELECT id FROM billing_imports WHERE billing_id = ?", (billing_id,)
    ).fetchone()
    if existing:
        cur.execute("DELETE FROM billing_imports WHERE billing_id = ?", (billing_id,))

    cur.execute(
        "INSERT INTO billing_imports (billing_id, file_name, row_count) VALUES (?, ?, ?)",
        (billing_id, file_name, len(rows)),
    )
    import_id = cur.lastrowid

    valid_customers = {r[0] for r in cur.execute("SELECT customer_number FROM customers").fetchall()}
    unmatched: set[int] = set()
    inserted = 0

    _cols = 98
    insert_sql = (
        """
        INSERT INTO billing_records (
            billing_import_id,
            piece_id, machine_serial, time_stamp, weight_oz,
            handling_type, usps_mail_class, usps_mail_prep_type,
            routing_category, routing_string, bundle_qualification, bundle_zip,
            account_id, account_name, custom_account_code,
            customer_barcode_symbology, customer_barcode,
            department_id, department_name, manifest_id,
            piece_postage, lbs_postage, final_postage,
            fully_paid_postage, billing_amount, imb_tracking_code,
            sack_level, sack_zip, destination_entry_level, zone,
            irregular, custom1, custom2, driver_route, adc,
            schemed_3d, schemed_5d, manifest_date,
            length_in, width_in, height_in, girth_in,
            is_flat_rate_conversion, nonrectangular, sub_type,
            ocr, bmc, asf, scf, master_mail_class,
            ezconfirm_pic, ezconfirm_processing_type,
            ezconfirm_name, ezconfirm_company,
            ezconfirm_address1, ezconfirm_address2, ezconfirm_city,
            ezconfirm_state, ezconfirm_zip, ezconfirm_zip4,
            ezconfirm_record_case_number, ezconfirm_is_uploaded,
            wabcr_symbology1, wabcr_data1, wabcr_symbology2, wabcr_data2,
            wabcr_symbology3, wabcr_data3, wabcr_symbology4, wabcr_data4,
            wabcr_symbology5, wabcr_data5,
            job_name, billing_id_ref, permit_origin, permit_number, permit_name,
            ezconfirm_special_services, mail_piece_tag_data, is_open_and_distribute,
            payment_method, premeter_qual_level, key_line, impb, impb_normalized, efn,
            surcharge_postage, fss, tub_number, postal_discounts,
            hr_address, hr_city, hr_state, hr_zip,
            label_list_installer_version, is_move, is_catalog,
            unmatched_account
        ) VALUES ("""
        + ",".join(["?"] * _cols)
        + ")"
    )

    for row in rows:
        cac = strip_zeros(row.get("Custom Account Code"))
        is_unmatched = 0 if (cac is not None and cac in valid_customers) else 1
        if is_unmatched and cac is not None:
            unmatched.add(cac)

        cur.execute(
            insert_sql,
            (
                import_id,
                row.get("Piece ID"),
                row.get("Machine Serial"),
                row.get("Time Stamp"),
                safe_real(row.get("Weight")),
                row.get("Handling Type"),
                row.get("USPS Mail Class"),
                row.get("USPS Mail Prep Type"),
                row.get("Routing Category"),
                row.get("Routing String"),
                row.get("Bundle Qualification"),
                row.get("Bundle Zip"),
                row.get("Account ID"),
                row.get("AccountName"),
                cac,
                row.get("Customer Barcode Symbology"),
                row.get("Customer Barcode"),
                row.get("Department ID"),
                row.get("Department Name"),
                row.get("Manifest ID"),
                safe_real(row.get("Piece Postage")),
                safe_real(row.get("LBS Postage")),
                safe_real(row.get("Final Postage")),
                safe_real(row.get("Fully Paid Postage")),
                safe_real(row.get("Billing Amount")),
                row.get("IMB Tracking Code"),
                row.get("SackLevel"),
                row.get("SackZip"),
                row.get("DestinationEntryLevel"),
                row.get("Zone"),
                row.get("Irregular"),
                row.get("Custom1"),
                row.get("Custom2"),
                row.get("Driver Route"),
                row.get("ADC"),
                row.get("Schemed 3D"),
                row.get("Schemed 5D"),
                row.get("Manifest Date"),
                safe_real(row.get("Length")),
                safe_real(row.get("Width")),
                safe_real(row.get("Height")),
                safe_real(row.get("Girth")),
                row.get("IsFlatRateConversion"),
                row.get("Nonrectangular"),
                row.get("SubType"),
                row.get("OCR"),
                row.get("BMC"),
                row.get("ASF"),
                row.get("SCF"),
                row.get("Master Mail Class"),
                row.get("EZConfirm PIC"),
                row.get("EZConfirm ProcessingType"),
                row.get("EZConfirm Name"),
                row.get("EZConfirm Company"),
                row.get("EZConfirm Address1"),
                row.get("EZConfirm Address2"),
                row.get("EZConfirm City"),
                row.get("EZConfirm State"),
                row.get("EZConfirm Zip"),
                row.get("EZConfirm Zip4"),
                row.get("EZConfirm RecordCaseNumber"),
                row.get("EZConfirm IsUploaded"),
                row.get("WABCR Symbology1"),
                row.get("WABCR Data1"),
                row.get("WABCR Symbology2"),
                row.get("WABCR Data2"),
                row.get("WABCR Symbology3"),
                row.get("WABCR Data3"),
                row.get("WABCR Symbology4"),
                row.get("WABCR Data4"),
                row.get("WABCR Symbology5"),
                row.get("WABCR Data5"),
                row.get("Job Name"),
                row.get("BillingID"),
                row.get("Permit Origin"),
                row.get("Permit Number"),
                row.get("Permit Name"),
                row.get("EZConfirm Special Services"),
                row.get("MailPieceTagData"),
                row.get("IsOpenAndDistribute"),
                row.get("Payment Method"),
                row.get("Premeter Qual Level"),
                row.get("KeyLine"),
                row.get("IMPB"),
                db.normalize_billing_impb(row.get("IMPB")),
                row.get("EFN"),
                safe_real(row.get("Surcharge Postage")),
                row.get("FSS"),
                row.get("TubNumber"),
                safe_real(row.get("Postal Discounts")),
                row.get("HRAddress"),
                row.get("HRCity"),
                row.get("HRState"),
                row.get("HRZip"),
                row.get("LabelListInstallerVersion"),
                row.get("Is Move"),
                row.get("Is Catalog"),
                is_unmatched,
            ),
        )
        inserted += 1

    conn.commit()
    conn.close()

    return {
        "billing_id": billing_id,
        "file_name": file_name,
        "rows_imported": inserted,
        "unmatched_accounts": sorted(unmatched),
    }


def import_customers_csv(csv_path: str, db_path: Path | str) -> dict[str, Any]:
    conn = sqlite3.connect(str(db_path))
    # Parent account rows often appear after children in the source CSV; allow any order.
    conn.execute("PRAGMA foreign_keys = OFF")
    cur = conn.cursor()
    cur.execute("DELETE FROM customers")

    warnings: list[str] = []
    count = 0
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("Customer") or row.get("customer") or "").strip()
            code = strip_zeros(row.get("Customer Code") or row.get("Customer Code "))
            parent_name = (row.get("Parent Co") or "").strip()
            parent_num_raw = row.get("Parent Co Number") or row.get("Parent Co Number ")
            parent_num = strip_zeros(parent_num_raw) if str(parent_num_raw or "").strip() else None

            if code is None:
                continue
            if not parent_name and not str(parent_num_raw or "").strip():
                pnum, pname = None, None
            elif parent_name and parent_num is not None:
                pname, pnum = parent_name, parent_num
            elif parent_name and parent_num is None:
                warnings.append(f"Parent name without number for customer {code}: {parent_name}")
                pname, pnum = parent_name, None
            else:
                pname, pnum = None, parent_num

            cur.execute(
                """
                INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
                VALUES (?, ?, ?, ?)
                """,
                (code, name or f"Account {code}", pnum, pname),
            )
            count += 1

    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")
    conn.close()
    return {"rows_imported": count, "warnings": warnings}


def _raw_export_normalize_lookup_key(name: str) -> str:
    """Normalize customer display name for Parent Company → column A VLOOKUP-style match."""
    return " ".join(name.split()).strip().casefold()


def import_customers_raw_export_xlsx(xlsx_path: str, db_path: Path | str) -> dict[str, Any]:
    """
    WS3 Raw Export workbook: column A = Customer, B = Code, N = Parent Company.
    Resolves parent_number by exact name lookup in column A → code in B (Excel VLOOKUP semantics: first match wins).
    """
    warnings: list[str] = []
    _WARN_CAP = 50

    def _finalize_warnings(ws: list[str]) -> list[str]:
        if len(ws) <= _WARN_CAP:
            return ws
        return ws[:_WARN_CAP] + [f"... and {len(ws) - _WARN_CAP} more warnings"]

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        data_rows: list[tuple[Any, ...]] = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not data_rows:
        return {"rows_imported": 0, "warnings": []}

    def _cell_str(v: Any) -> str:
        if v is None:
            return ""
        return str(v).strip()

    def _abn_padded(row: tuple[Any, ...]) -> tuple[Any, Any, Any]:
        need = 14
        if len(row) < need:
            row = tuple(list(row) + [None] * (need - len(row)))
        return row[0], row[1], row[13]

    start = 0
    hdr_a, hdr_b, _ = _abn_padded(data_rows[0])
    if _cell_str(hdr_a).casefold() == "customer" and _cell_str(hdr_b).casefold() == "code":
        start = 1

    name_to_code: dict[str, int] = {}
    for row in data_rows[start:]:
        a, b, _n = _abn_padded(row)
        code = strip_zeros(b)
        name = _cell_str(a)
        if code is None or not name:
            continue
        k = _raw_export_normalize_lookup_key(name)
        if not k:
            continue
        if k not in name_to_code:
            name_to_code[k] = code
        elif name_to_code[k] != code:
            warnings.append(
                f"Duplicate customer name in column A (keeping first code {name_to_code[k]}): "
                f"{name!r} also has code {code}"
            )

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = OFF")
    cur = conn.cursor()
    cur.execute("DELETE FROM customers")

    count = 0
    for row in data_rows[start:]:
        a, b, n_col = _abn_padded(row)
        code = strip_zeros(b)
        if code is None:
            continue
        cust_name = _cell_str(a)
        display_name = cust_name or f"Account {code}"

        parent_raw = _cell_str(n_col)
        if not parent_raw:
            pnum, pname = None, None
        else:
            pk = _raw_export_normalize_lookup_key(parent_raw)
            pnum = name_to_code.get(pk)
            if pnum is not None:
                pname = parent_raw
            else:
                warnings.append(
                    f"Parent Company unknown in column A (no matching name) for customer {code}: {parent_raw!r}"
                )
                pname, pnum = parent_raw, None

        cur.execute(
            """
            INSERT INTO customers (customer_number, customer_name, parent_number, parent_name)
            VALUES (?, ?, ?, ?)
            """,
            (code, display_name, pnum, pname),
        )
        count += 1

    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")
    conn.close()

    return {"rows_imported": count, "warnings": _finalize_warnings(warnings)}


def import_flat_rate_costs(csv_path: str, db_path: Path | str) -> dict[str, Any]:
    def parse_rate(val: Any) -> float | None:
        try:
            return float(str(val).replace("$", "").strip())
        except (ValueError, TypeError):
            return None

    rows: list[dict[str, Any]] = []
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fn = [h.strip().replace("\n", " ") for h in (reader.fieldnames or [])]
        reader.fieldnames = fn
        for row in reader:
            norm = {(k or "").strip().replace("\n", " "): v for k, v in row.items()}
            rows.append(
                {
                    "weight_not_over_oz": float(str(norm.get("Weight Not Over (oz.)", "0")).strip()),
                    "rate_5digit": parse_rate(norm.get("5-Digit")),
                    "rate_3digit": parse_rate(norm.get("3 digit")),
                    "rate_aadc": parse_rate(norm.get("AADC")),
                    "rate_mixed_adc": parse_rate(norm.get("Mixed ADC")),
                    "rate_machinable_pres": parse_rate(
                        norm.get("Machinable Presorted") or norm.get("Machinable\nPresorted")
                    ),
                    "rate_retail": parse_rate(norm.get("Retail Cost")),
                }
            )

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()
    cur.execute("DELETE FROM flat_rate_costs")
    for row in rows:
        row["effective_date"] = db.FLAT_RATE_BASELINE_DATE
    cur.executemany(
        """
        INSERT INTO flat_rate_costs
            (weight_not_over_oz, rate_5digit, rate_3digit, rate_aadc,
             rate_mixed_adc, rate_machinable_pres, rate_retail, effective_date)
        VALUES
            (:weight_not_over_oz, :rate_5digit, :rate_3digit, :rate_aadc,
             :rate_mixed_adc, :rate_machinable_pres, :rate_retail, :effective_date)
        """,
        rows,
    )
    conn.commit()
    conn.close()
    return {"rows_imported": len(rows)}


def _parse_effective_date_any(cell: Any) -> str | None:
    """
    Try to parse M/D/YYYY or similar into YYYY-MM-DD.
    Returns None if parsing fails.
    """
    if cell is None:
        return None
    s = str(cell).strip()
    if not s:
        return None
    s = s.replace("\\", "/").replace("-", "/")
    # Prefer month/day/year (matches the sample files).
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat()
        except ValueError:
            continue
    return None


def parse_priority_mail_filename_effective_date(file_name: str) -> str | None:
    """Parse M-D-YY or M-D-YYYY from basename (e.g. 'Priority mail zones 4-27-26')."""
    stem = Path(file_name).stem
    m = re.search(r"\b(\d{1,2})-(\d{1,2})-(\d{2,4})\b", stem)
    if not m:
        return None
    month, day, y_raw = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    if y_raw < 100:
        year = 2000 + y_raw if y_raw < 70 else 1900 + y_raw
    else:
        year = y_raw
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def _cell_money_float(val: Any) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _cell_float(val: Any) -> float | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _normalize_row_cells(row: list[Any]) -> list[str]:
    return [str(c).strip() if c is not None else "" for c in row]


def _header_first_nine_cells_are_usps_zones_1_through_9(header: list[str]) -> bool:
    """True when A1–I1 hold the digits 1–9 in some order (zone labels only)."""
    if len(header) < 9:
        return False
    zones: list[int] = []
    for i in range(9):
        zf = _cell_float(header[i])
        if zf is None:
            return False
        zi = int(round(zf))
        if abs(zf - zi) > 1e-6:
            return False
        if not (1 <= zi <= 9):
            return False
        zones.append(zi)
    return sorted(zones) == [1, 2, 3, 4, 5, 6, 7, 8, 9]


def _compact_matrix_zone_cols_from_header_row(header: list[str]) -> list[tuple[int, int]] | None:
    """
    Map column index -> USPS zone for compact zone matrix xlsx.
    Layout A: Row 1 col A texts like '(lbs.)', cols B+ are zone labels; data column A is weight lb.
    Layout B: Row 1 cols A-I are zones 1–9; data column A is weight lb and cols B+ are zone prices aligned
             so zone for header[col-1] is read from data[col].

    Each tuple is (zone_number, price_column_index).
    Returns None when the sheet does not match either layout.
    """
    if not header:
        return None
    if "lb" in header[0].lower():
        zone_cols: list[tuple[int, int]] = []
        for idx in range(1, min(len(header), 12)):
            zf = _cell_float(header[idx])
            if zf is None:
                continue
            zi = int(zf)
            if 1 <= zi <= 9:
                zone_cols.append((zi, idx))
        if len(zone_cols) < 9:
            return None
        return zone_cols
    if not _header_first_nine_cells_are_usps_zones_1_through_9(header):
        return None
    out: list[tuple[int, int]] = []
    for j in range(1, 10):
        zf = _cell_float(header[j - 1])
        if zf is None:
            return None
        zi = int(round(zf))
        if not (1 <= zi <= 9):
            return None
        out.append((zi, j))
    return out


def _parse_compact_priority_mail_zone_matrix_xlsx(
    raw_rows: list[list[Any]],
    file_name: str,
    *,
    effective_date: str | None = None,
) -> list[dict[str, Any]] | None:
    """
    Compact zone-price matrix .xlsx: data rows weight (lb) in column A.

    Supported row-0 layouts:
    - '(lbs.)' or similar in A1 with zone digits 1–9 in adjacent columns (original path).
    - Zones 1–9 (each once) across A–I; prices for zone header[j−1] are in column j (user template).
    """
    if len(raw_rows) < 2:
        return None
    header = _normalize_row_cells(raw_rows[0])
    zone_cols = _compact_matrix_zone_cols_from_header_row(header)
    if zone_cols is None:
        return None

    out_rows: list[dict[str, Any]] = []
    sort_group = 1
    sort_order = 0
    for r in raw_rows[1:]:
        cells = _normalize_row_cells(r)
        line = " ".join(c for c in cells if c).strip()
        if not line:
            continue
        w = _cell_float(cells[0])
        if w is None or w <= 0:
            continue
        for zone, idx in zone_cols:
            if idx >= len(cells):
                continue
            p = _cell_money_float(cells[idx])
            if p is None:
                continue
            out_rows.append(
                {
                    "effective_date": effective_date,
                    "source_file_name": file_name,
                    "row_type": "matrix",
                    "label": None,
                    "zone": zone,
                    "weight_unit": "lb",
                    "weight_max": float(w),
                    "price": float(p),
                    "sort_group": sort_group,
                    "sort_order": sort_order,
                }
            )
            sort_order += 1
    return out_rows if out_rows else None


def _parse_priority_mail_retail_rows(
    raw_rows: list[list[Any]],
    file_name: str,
) -> tuple[list[dict[str, Any]], str | None]:
    """Parse USPS Priority Mail retail CSV-style rows (flat-rate block + zone matrix + fees)."""
    effective_date: str | None = None
    out_rows: list[dict[str, Any]] = []

    sort_group = 0
    sort_order = 0

    in_matrix = False
    zone_cols: list[tuple[int, int]] = []  # (zone_number, col_index)

    for r in raw_rows:
        cells = _normalize_row_cells(r)
        line = " ".join(c for c in cells if c).strip()
        if not line:
            continue

        # Header: "Priority Mail - Retail,,Final,,4/7/2026"
        if not effective_date and any("priority mail" in c.lower() and "retail" in c.lower() for c in cells):
            # Scan for a date-like token anywhere on the row.
            for c in cells:
                if (d := _parse_effective_date_any(c)) is not None:
                    effective_date = d
                    break
            sort_group += 1
            continue

        # Start of matrix section
        if any("weight not over" in c.lower() and "(lbs" in c.lower() for c in cells):
            in_matrix = True
            zone_cols = []
            sort_group += 1
            continue

        # Zone header row (sometimes begins with blank then Zone 1..)
        if in_matrix and any(re.match(r"^zone\s*\d+$", c, re.I) for c in cells):
            zone_cols = []
            for idx, c in enumerate(cells):
                m = re.match(r"^zone\s*(\d+)$", c, re.I)
                if m:
                    zone_cols.append((int(m.group(1)), idx))
            sort_group += 1
            continue

        # Matrix data row: first cell numeric weight, then prices by zone columns
        if in_matrix and zone_cols:
            w = _cell_float(cells[0])
            if w is not None and w > 0:
                for zone, idx in zone_cols:
                    if idx >= len(cells):
                        continue
                    p = _cell_money_float(cells[idx])
                    if p is None:
                        continue
                    out_rows.append(
                        {
                            "effective_date": effective_date,
                            "source_file_name": file_name,
                            "row_type": "matrix",
                            "label": None,
                            "zone": zone,
                            "weight_unit": "lb",
                            "weight_max": float(w),
                            "price": float(p),
                            "sort_group": sort_group,
                            "sort_order": sort_order,
                        }
                    )
                    sort_order += 1
                continue

        # Non-matrix lines: flat rate items and fee lines.
        # Flat-rate items typically: "<Label>,12.90"
        if not in_matrix:
            label = cells[0].strip().strip('"')
            price = _cell_money_float(cells[1] if len(cells) > 1 else "")
            if label and price is not None:
                out_rows.append(
                    {
                        "effective_date": effective_date,
                        "source_file_name": file_name,
                        "row_type": "flat_rate_item",
                        "label": label,
                        "zone": None,
                        "weight_unit": None,
                        "weight_max": None,
                        "price": float(price),
                        "sort_group": sort_group,
                        "sort_order": sort_order,
                    }
                )
                sort_order += 1
                continue

        # Fees/notes: take first non-empty cell as label, and first parseable money as price.
        label = ""
        for c in cells:
            if c:
                label = c.strip().strip('"')
                break
        if label:
            price: float | None = None
            for c in cells[1:]:
                if (p := _cell_money_float(c)) is not None:
                    price = p
                    break
            out_rows.append(
                {
                    "effective_date": effective_date,
                    "source_file_name": file_name,
                    "row_type": "fee" if price is not None else "note",
                    "label": label,
                    "zone": None,
                    "weight_unit": None,
                    "weight_max": None,
                    "price": float(price) if price is not None else None,
                    "sort_group": sort_group,
                    "sort_order": sort_order,
                }
            )
            sort_order += 1

    return out_rows, effective_date


def _priority_mail_retail_rows_from_file(
    source_path: str,
    *,
    effective_date: str | None = None,
) -> tuple[list[dict[str, Any]], str | None]:
    """Load raw grid from CSV or XLSX and parse into priority_mail_retail row dicts."""
    file_name = os.path.basename(source_path)
    suffix = Path(source_path).suffix.lower()
    if suffix == ".xlsx":
        wb = load_workbook(source_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        compact = _parse_compact_priority_mail_zone_matrix_xlsx(
            raw_rows,
            file_name,
            effective_date=effective_date,
        )
        if compact is not None:
            eff_compact = effective_date or parse_priority_mail_filename_effective_date(file_name)
            if eff_compact is None:
                raise ValueError(
                    "Compact Priority zone .xlsx requires an effective_date (System tab upload) "
                    "or a parseable date in the filename (e.g. zones-4-27-26)."
                )
            for r in compact:
                r["effective_date"] = eff_compact
            return compact, eff_compact
        out_rows, eff_parse = _parse_priority_mail_retail_rows(raw_rows, file_name)
        final_eff = effective_date or eff_parse
        if effective_date:
            for r in out_rows:
                r["effective_date"] = effective_date
            final_eff = effective_date
        return out_rows, final_eff
    with open(source_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        raw_rows = [r for r in reader]
    out_rows, eff_parse = _parse_priority_mail_retail_rows(raw_rows, file_name)
    final_eff = effective_date or eff_parse
    if effective_date:
        for r in out_rows:
            r["effective_date"] = effective_date
        final_eff = effective_date
    return out_rows, final_eff


def import_priority_mail_retail(
    source_path: str,
    db_path: Path | str,
    *,
    effective_date: str | None = None,
) -> dict[str, Any]:
    """
    Import Priority Mail Retail rates (matrix by Zone x Weight(Lbs) + flat-rate items + fees).
    Accepts `.csv` (USPS export layout) or `.xlsx` (compact zone matrix: row 0 = lbs + zones 1–9,
    or zones 1–9 only across columns A–I with weights in column A).

    Rows are tagged with ``effective_date``; re-importing the same date replaces only that revision.
    When ``effective_date`` is omitted, uses the date parsed from a USPS CSV header and/or filename
    (compact .xlsx must supply an explicit date or have one in the filename).
    """
    file_name = os.path.basename(source_path)
    out_rows, effective_used = _priority_mail_retail_rows_from_file(
        source_path,
        effective_date=effective_date,
    )
    if effective_used is None:
        raise ValueError(
            "Could not determine effective_date for Priority Mail retail import "
            "(set it on upload or include a dated USPS CSV header)."
        )

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM priority_mail_retail WHERE effective_date IS ?",
        (effective_used,),
    )
    cur.executemany(
        """
        INSERT INTO priority_mail_retail (
            effective_date, source_file_name, row_type, label, zone,
            weight_unit, weight_max, price, sort_group, sort_order
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                r.get("effective_date"),
                r.get("source_file_name"),
                r.get("row_type"),
                r.get("label"),
                r.get("zone"),
                r.get("weight_unit"),
                r.get("weight_max"),
                r.get("price"),
                r.get("sort_group"),
                r.get("sort_order"),
            )
            for r in out_rows
        ],
    )
    conn.commit()
    conn.close()
    result: dict[str, Any] = {
        "rows_imported": len(out_rows),
        "effective_date": effective_used,
        "file_name": file_name,
    }
    suffix = Path(source_path).suffix.lower()
    if suffix == ".xlsx":
        n_matrix = sum(1 for r in out_rows if r.get("row_type") == "matrix")
        # Each intact weight band should produce one cell per USPS zone (9); otherwise the grid is likely misread.
        if n_matrix % 9 != 0:
            result["warnings"] = [
                f"Priority Mail zone matrix imported {n_matrix} cells ({n_matrix} is not divisible by "
                "9 zones). Rows may not match zones 1–9 across columns. Check row 1 headers and "
                "that column A is weight (lb)."
            ]
    return result


def import_ground_advantage_retail(csv_path: str, db_path: Path | str) -> dict[str, Any]:
    """
    Import USPS Ground Advantage Retail rates (oz section + lb section + fees/special lines).
    Full-replaces `ground_advantage_retail`.
    """
    file_name = os.path.basename(csv_path)
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        raw_rows = [r for r in reader]

    effective_date: str | None = None
    out_rows: list[dict[str, Any]] = []

    sort_group = 0
    sort_order = 0

    section_unit: str | None = None  # oz | lb
    in_matrix = False
    zone_cols: list[tuple[int, int]] = []

    for r in raw_rows:
        cells = _normalize_row_cells(r)
        line = " ".join(c for c in cells if c).strip()
        if not line:
            continue

        if not effective_date and any("ground advantage" in c.lower() for c in cells):
            # date sometimes isn't on the same row; still scan this row.
            for c in cells:
                if (d := _parse_effective_date_any(c)) is not None:
                    effective_date = d
                    break
            sort_group += 1
            continue

        if not effective_date:
            for c in cells:
                if (d := _parse_effective_date_any(c)) is not None:
                    effective_date = d
                    break

        if any("weight not over" in c.lower() and "(ounces" in c.lower() for c in cells):
            section_unit = "oz"
            in_matrix = True
            zone_cols = []
            sort_group += 1
            continue

        if any("weight not over" in c.lower() and "(pounds" in c.lower() for c in cells):
            section_unit = "lb"
            in_matrix = True
            # The pounds section in your file does not repeat the Zone header row,
            # so keep the previously detected `zone_cols` (from the ounces section).
            sort_group += 1
            continue

        if in_matrix and any(re.match(r"^zone\s*\d+$", c, re.I) for c in cells):
            zone_cols = []
            for idx, c in enumerate(cells):
                m = re.match(r"^zone\s*(\d+)$", c, re.I)
                if m:
                    zone_cols.append((int(m.group(1)), idx))
            sort_group += 1
            continue

        if in_matrix and zone_cols and section_unit:
            w = _cell_float(cells[0])
            if w is not None and w > 0:
                for zone, idx in zone_cols:
                    if idx >= len(cells):
                        continue
                    p = _cell_money_float(cells[idx])
                    if p is None:
                        continue
                    out_rows.append(
                        {
                            "effective_date": effective_date,
                            "source_file_name": file_name,
                            "row_type": "matrix",
                            "label": None,
                            "zone": zone,
                            "weight_unit": section_unit,
                            "weight_max": float(w),
                            "price": float(p),
                            "sort_group": sort_group,
                            "sort_order": sort_order,
                        }
                    )
                    sort_order += 1
                continue

        # Special/fee/note lines (e.g. Oversized, Nonstandard Length..., notes)
        label = ""
        for c in cells:
            if c:
                label = c.strip().strip('"')
                break
        if label:
            price: float | None = None
            for c in cells[1:]:
                if (p := _cell_money_float(c)) is not None:
                    price = p
                    break
            out_rows.append(
                {
                    "effective_date": effective_date,
                    "source_file_name": file_name,
                    "row_type": "fee" if price is not None else "note",
                    "label": label,
                    "zone": None,
                    "weight_unit": None,
                    "weight_max": None,
                    "price": float(price) if price is not None else None,
                    "sort_group": sort_group,
                    "sort_order": sort_order,
                }
            )
            sort_order += 1

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()
    cur.execute("DELETE FROM ground_advantage_retail")
    cur.executemany(
        """
        INSERT INTO ground_advantage_retail (
            effective_date, source_file_name, row_type, label, zone,
            weight_unit, weight_max, price, sort_group, sort_order
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                r.get("effective_date"),
                r.get("source_file_name"),
                r.get("row_type"),
                r.get("label"),
                r.get("zone"),
                r.get("weight_unit"),
                r.get("weight_max"),
                r.get("price"),
                r.get("sort_group"),
                r.get("sort_order"),
            )
            for r in out_rows
        ],
    )
    conn.commit()
    conn.close()
    return {
        "rows_imported": len(out_rows),
        "effective_date": effective_date,
        "file_name": file_name,
    }


NOTICE123_DIR = db.ROOT / "Notice123"

NOTICE123_PM_RETAIL_CSV = "PM Retail.csv"
NOTICE123_FCM_RETAIL_CSV = "FCM & EDDM - Retail.csv"
NOTICE123_FCM_COMM_FLATS_CSV = "FCM - Comm Flats.csv"


def _parse_notice123_money(val: Any) -> float | None:
    try:
        return float(str(val).replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _parse_notice123_weight_oz(val: Any) -> float | None:
    try:
        w = float(str(val).strip())
    except (ValueError, TypeError):
        return None
    if w <= 0 or w > 13:
        return None
    return w


def parse_notice123_fcm_retail_flats(csv_path: str) -> dict[float, float]:
    """Parse flats retail tiers from FCM & EDDM - Retail.csv (cols C/D)."""
    retail_by_weight: dict[float, float] = {}
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 4:
                continue
            w = _parse_notice123_weight_oz(row[2])
            if w is None:
                continue
            price = _parse_notice123_money(row[3])
            if price is None:
                continue
            retail_by_weight[w] = price
    return retail_by_weight


def parse_notice123_fcm_comm_flats_presort(csv_path: str) -> list[dict[str, Any]]:
    """Parse commercial presort tiers from FCM - Comm Flats.csv."""
    rows: list[dict[str, Any]] = []
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 5:
                continue
            w = _parse_notice123_weight_oz(row[0])
            if w is None:
                continue
            rows.append(
                {
                    "weight_not_over_oz": w,
                    "rate_5digit": _parse_notice123_money(row[1]),
                    "rate_3digit": _parse_notice123_money(row[2]),
                    "rate_mixed_adc": _parse_notice123_money(row[3]),
                    "rate_machinable_pres": _parse_notice123_money(row[4]),
                }
            )
    return rows


def _notice123_csv_rank(path: Path, root: Path) -> tuple[int, int, int, str]:
    """Sort key for duplicate CSV resolution (lower = preferred)."""
    rel = path.relative_to(root)
    depth = len(rel.parts) - 1
    parent_parts = rel.parts[:-1]
    in_notice_folder = any("notice 123" in part.lower() for part in parent_parts)
    return (
        0 if depth > 0 else 1,
        0 if in_notice_folder else 1,
        -len(rel.parts),
        str(rel).lower(),
    )


def _find_notice123_csv(root: Path, basename: str) -> tuple[Path, int]:
    """
    Locate a required Notice 123 CSV under ``root``.

    Returns ``(path, candidate_count)``. When multiple paths match, picks the best
    nested path (prefer folders named like Notice 123). Raises if top candidates
    differ in file content.
    """
    matches = [p for p in root.rglob("*.csv") if p.name.lower() == basename.lower()]
    if not matches:
        raise ValueError(f"Required Notice 123 file not found: {basename}")
    if len(matches) == 1:
        return matches[0], 1

    ranked = sorted(matches, key=lambda p: _notice123_csv_rank(p, root))
    best_rank = _notice123_csv_rank(ranked[0], root)
    candidates = [p for p in ranked if _notice123_csv_rank(p, root) == best_rank]
    hashes = {hashlib.sha256(p.read_bytes()).hexdigest() for p in candidates}
    if len(hashes) > 1:
        paths = ", ".join(str(p.relative_to(root)) for p in candidates[:5])
        raise ValueError(
            f"Multiple distinct {basename} files found (content differs): {paths}"
        )
    return candidates[0], len(matches)


def extract_notice123_zip(zip_path: str | Path, dest_dir: Path | None = None) -> Path:
    """Extract a Notice 123 zip into a fresh staging directory (not Notice123/)."""
    _ = dest_dir  # staging only; dest_dir is applied when replacing Notice123/
    staging = Path(tempfile.mkdtemp(prefix="notice123-stage-"))
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(staging)
    return staging


def _replace_notice123_dir(staged_root: Path, dest_dir: Path | None = None) -> Path:
    """Replace Notice123/ (or ``dest_dir``) with the staged extract tree."""
    target = dest_dir or NOTICE123_DIR
    parent = target.parent
    parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        shutil.rmtree(target)
    shutil.move(str(staged_root), str(target))
    return target


def import_notice123_flat_rates(
    retail_csv: str,
    presort_csv: str,
    db_path: Path | str,
    *,
    effective_date: str,
) -> dict[str, Any]:
    """Import merged flats retail + presort for one effective-date revision."""
    retail_by_weight = parse_notice123_fcm_retail_flats(retail_csv)
    presort_rows = parse_notice123_fcm_comm_flats_presort(presort_csv)
    if not retail_by_weight and not presort_rows:
        raise ValueError("No flats rate tiers found in Notice 123 CSV files")

    merged: dict[float, dict[str, Any]] = {}
    for w, retail in retail_by_weight.items():
        merged[w] = {
            "weight_not_over_oz": w,
            "rate_retail": retail,
            "rate_5digit": None,
            "rate_3digit": None,
            "rate_aadc": None,
            "rate_mixed_adc": None,
            "rate_machinable_pres": None,
            "effective_date": effective_date,
        }
    for row in presort_rows:
        w = float(row["weight_not_over_oz"])
        if w not in merged:
            merged[w] = {
                "weight_not_over_oz": w,
                "rate_retail": None,
                "rate_5digit": None,
                "rate_3digit": None,
                "rate_aadc": None,
                "rate_mixed_adc": None,
                "rate_machinable_pres": None,
                "effective_date": effective_date,
            }
        for col in (
            "rate_5digit",
            "rate_3digit",
            "rate_mixed_adc",
            "rate_machinable_pres",
        ):
            if row.get(col) is not None:
                merged[w][col] = row[col]

    out_rows = [merged[k] for k in sorted(merged.keys())]
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM flat_rate_costs WHERE effective_date IS ?",
        (effective_date,),
    )
    cur.executemany(
        """
        INSERT INTO flat_rate_costs (
            weight_not_over_oz, rate_5digit, rate_3digit, rate_aadc,
            rate_mixed_adc, rate_machinable_pres, rate_retail, effective_date
        ) VALUES (
            :weight_not_over_oz, :rate_5digit, :rate_3digit, :rate_aadc,
            :rate_mixed_adc, :rate_machinable_pres, :rate_retail, :effective_date
        )
        """,
        out_rows,
    )
    conn.commit()
    conn.close()
    return {
        "rows_imported": len(out_rows),
        "effective_date": effective_date,
        "retail_tiers": len(retail_by_weight),
        "presort_tiers": len(presort_rows),
    }


def import_notice123_rate_case(
    zip_path: str | Path,
    db_path: Path | str,
    *,
    effective_date: str,
    dest_dir: Path | None = None,
) -> dict[str, Any]:
    """
    Extract a Notice 123 zip and import PM retail + flats retail/presort rates.
    """
    eff = str(effective_date).strip()
    if not eff:
        raise ValueError("effective_date required (YYYY-MM-DD)")

    staged_root: Path | None = None
    try:
        staged_root = extract_notice123_zip(zip_path)
        pm_csv, pm_candidates = _find_notice123_csv(staged_root, NOTICE123_PM_RETAIL_CSV)
        retail_csv, retail_candidates = _find_notice123_csv(
            staged_root, NOTICE123_FCM_RETAIL_CSV
        )
        presort_csv, presort_candidates = _find_notice123_csv(
            staged_root, NOTICE123_FCM_COMM_FLATS_CSV
        )

        pm_result = import_priority_mail_retail(
            str(pm_csv), db_path, effective_date=eff
        )
        flats_result = import_notice123_flat_rates(
            str(retail_csv),
            str(presort_csv),
            db_path,
            effective_date=eff,
        )
        files_rel = {
            "pm_retail": str(pm_csv.relative_to(staged_root)),
            "fcm_retail": str(retail_csv.relative_to(staged_root)),
            "fcm_comm_flats": str(presort_csv.relative_to(staged_root)),
            "pm_retail_candidates": pm_candidates,
            "fcm_retail_candidates": retail_candidates,
            "fcm_comm_flats_candidates": presort_candidates,
        }
        extract_path = _replace_notice123_dir(staged_root, dest_dir=dest_dir)
        staged_root = None
        return {
            "effective_date": eff,
            "extract_path": str(extract_path),
            "priority_mail": pm_result,
            "flats": flats_result,
            "files": files_rel,
        }
    finally:
        if staged_root is not None and staged_root.exists():
            shutil.rmtree(staged_root, ignore_errors=True)


# --- Pitney Detail Transactions (per-transaction carrier ledger) ---

_PITNEY_REQUIRED_HEADERS = ("transactionType", "amount", "transactionId", "transactionDateTime")
_PITNEY_OPTIONAL_HEADERS = (
    "parcelTrackingNumber",
    "service",
    "zone",
    "weightInOunces",
    "status",
    "postageBalance",
)


def _pitney_iso_date(raw: Any) -> str | None:
    """ISO date from a Pitney transactionDateTime ('2026-05-30 22:04:56' or datetime)."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    s = str(raw).strip()
    if not s:
        return None
    cal = s.split()[0]
    try:
        return datetime.strptime(cal[:10], "%Y-%m-%d").date().isoformat()
    except ValueError:
        pass
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(cal, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def import_pitney_detail_transactions(xlsx_path: str, db_path: Path | str) -> dict[str, Any]:
    """
    Import a Pitney Detail Transactions export (.xlsx) into ``pitney_transactions``.

    Rows are deduplicated by (transactionId, type, datetime): a print and its
    refund/adjustment share a transactionId, and a refund is re-listed at each
    lifecycle stage (REQUESTED/ACCEPTED/DENIED), so all such rows are kept while
    re-importing a file (or overlapping monthly exports) inserts nothing twice.
    All transaction types are stored, including ``Postage Fund`` deposits;
    cost/reconciliation queries exclude funds and count refunds only at ACCEPTED.
    """
    file_name = os.path.basename(xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)

        idx: dict[str, int] | None = None
        header_scan = 0
        for raw in rows_iter:
            header_scan += 1
            cells = [str(c).strip() if c is not None else "" for c in raw]
            if all(h in cells for h in _PITNEY_REQUIRED_HEADERS):
                idx = {h: cells.index(h) for h in cells if h}
                break
            if header_scan >= 10:
                break
        if idx is None:
            raise ValueError(
                "Not a Pitney Detail Transactions export (missing "
                + ", ".join(_PITNEY_REQUIRED_HEADERS)
                + " header row)"
            )

        def col(row: tuple, name: str) -> Any:
            i = idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        conn = sqlite3.connect(str(db_path))
        try:
            conn.execute("PRAGMA foreign_keys = ON")
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO pitney_imports (file_name, row_count) VALUES (?, 0)",
                (file_name,),
            )
            import_id = cur.lastrowid

            row_count = 0
            inserted = 0
            skipped = 0
            for raw in rows_iter:
                tx_id = col(raw, "transactionId")
                tx_type = col(raw, "transactionType")
                if tx_id is None or str(tx_id).strip() == "":
                    continue
                if tx_type is None or str(tx_type).strip() == "":
                    continue
                row_count += 1
                tracking_raw = col(raw, "parcelTrackingNumber")
                tracking = db.normalize_pitney_tracking(tracking_raw)
                dt_raw = col(raw, "transactionDateTime")
                cur.execute(
                    """
                    INSERT INTO pitney_transactions (
                        import_id, transaction_id, transaction_type,
                        transaction_date, transaction_datetime, amount,
                        tracking_number, tracking_normalized,
                        service, zone, weight_oz, status, postage_balance
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(transaction_id, transaction_type, transaction_datetime)
                    DO NOTHING
                    """,
                    (
                        import_id,
                        str(tx_id).strip(),
                        str(tx_type).strip(),
                        _pitney_iso_date(dt_raw),
                        str(dt_raw).strip() if dt_raw is not None else "",
                        safe_real(col(raw, "amount")),
                        tracking,
                        tracking,
                        (str(col(raw, "service")).strip() or None)
                        if col(raw, "service") is not None
                        else None,
                        (str(col(raw, "zone")).strip() or None)
                        if col(raw, "zone") is not None
                        else None,
                        safe_real(col(raw, "weightInOunces")),
                        (str(col(raw, "status")).strip() or None)
                        if col(raw, "status") is not None
                        else None,
                        safe_real(col(raw, "postageBalance")),
                    ),
                )
                if cur.rowcount > 0:
                    inserted += 1
                else:
                    skipped += 1

            cur.execute(
                "UPDATE pitney_imports SET row_count = ?, inserted_count = ?, skipped_count = ? WHERE id = ?",
                (row_count, inserted, skipped, import_id),
            )
            conn.commit()
        finally:
            conn.close()
    finally:
        wb.close()

    return {
        "file": file_name,
        "row_count": row_count,
        "rows_imported": inserted,
        "skipped_duplicates": skipped,
    }
